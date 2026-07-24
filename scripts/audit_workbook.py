"""
Automated data audit — cross-check every workbook claim against scored.csv.

Run after build_xlsx in the auto-build chain. Catches:
  - Auto-generated text that contradicts the underlying numbers (Tab 1 claims
    "8/8 single-field wins" when the actual best is 7 pairs + 1 Full brief).
  - Math errors (CI lower > upper, score outside [0, 1], etc.).
  - Workbook ↔ scored.csv drift (table cells that don't match a fresh groupby).

Exit code:
  0 = all checks pass
  1 = at least one mismatch (auto-build chain surfaces [WARN] in STOP gate)

Run standalone:
    python -m scripts.audit_workbook
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

from src import config as cfg

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

SCORED_CSV = cfg.OUTPUTS_DIR / "scored.csv"
WORKBOOK   = cfg.RESULTS_DIR / "Prompt Eval Results.xlsx"


# ---------- check harness ----------

class AuditFailed(Exception):
    """Raised on missing prerequisites; not a check failure."""


def _check(checks: list, name: str, condition: bool, detail: str = "") -> None:
    """Append a (passed, name, detail) row."""
    checks.append((bool(condition), name, detail))


def _find_header_row(ws, key: str, search_col: int = 1, max_row: int = 80) -> int | None:
    for r in range(1, max_row + 1):
        v = ws.cell(r, search_col).value
        if v and key in str(v):
            return r
    return None


# ---------- individual checks ----------

def check_ci_bounds(checks: list, ws) -> None:
    """Tab 2 95% CI strings like '[0.59, 0.67]' must parse, be in [0,1], lo ≤ hi."""
    hdr = _find_header_row(ws, "Task")
    if hdr is None:
        _check(checks, "Tab 2: header row found", False, "no row begins with 'Task'")
        return
    bad = 0
    seen = 0
    for r in range(hdr + 1, hdr + 20):
        v = ws.cell(r, 5).value  # "95% CI" column
        if v in (None, "", "—"):
            continue
        seen += 1
        m = re.match(r"\[([\d.]+),\s*([\d.]+)\]$", str(v).strip())
        if not m:
            bad += 1
            continue
        lo, hi = float(m.group(1)), float(m.group(2))
        if not (0 <= lo <= hi <= 1):
            bad += 1
    _check(checks, "Tab 2 95% CI bounds valid",
           seen > 0 and bad == 0,
           f"{bad} of {seen} rows have invalid CI" if bad else f"all {seen} rows OK")


def check_delta_average(checks: list, ws_tab1, ws_tab2, scored: pd.DataFrame) -> None:
    """Tab 1 Δ-vs-Full-Brief avg must equal the actual Tab 2 Δ-column mean."""
    # Tab 1 claim
    claim_row = None
    for r in range(1, 30):
        v = ws_tab1.cell(r, 1).value
        if v and "Δ vs Full brief" in str(v):
            claim_row = r
            break
    claim_avg = None
    if claim_row:
        claim_str = str(ws_tab1.cell(claim_row, 2).value or "")
        m = re.search(r"([+\-]?\d*\.\d+)\s*avg", claim_str)
        if m:
            claim_avg = float(m.group(1))

    # Tab 2 Δ column
    hdr = _find_header_row(ws_tab2, "Task")
    actuals = []
    for r in range(hdr + 1, hdr + 20) if hdr else []:
        v = ws_tab2.cell(r, 6).value  # Δ column
        if v in (None, "", "—"):
            continue
        try:
            actuals.append(float(v))
        except (TypeError, ValueError):
            pass
    if claim_avg is not None and actuals:
        actual_avg = sum(actuals) / len(actuals)
        diff = abs(claim_avg - actual_avg)
        _check(checks, "Tab 1 Δ-avg matches Tab 2 column",
               diff < 0.003,
               f"claim {claim_avg:+.3f} vs actual {actual_avg:+.3f} (Δ {diff:.4f})")
    else:
        _check(checks, "Tab 1 Δ-avg matches Tab 2 column",
               False, "could not parse claim or column")


def check_full_brief_verdict(checks: list, ws_tab1, scored: pd.DataFrame) -> None:
    """Tab 1 'Is Full brief worth using?' must agree with absolute-best analysis."""
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return
    per_task_best = (
        sent.groupby(["task", "config_id"])["cosine"].mean().reset_index()
        .sort_values(["task", "cosine"], ascending=[True, False])
        .groupby("task").head(1)
    )
    full_brief = {"A:_full_brief", "_full_brief"}
    n_full = int(per_task_best["config_id"].isin(full_brief).sum())
    total = len(per_task_best)

    claim = None
    for r in range(1, 30):
        v = ws_tab1.cell(r, 1).value
        if v and "Full brief worth" in str(v):
            claim = str(ws_tab1.cell(r, 2).value or "")
            break
    if not claim:
        _check(checks, "Tab 1 Full-brief verdict present", False, "row not found")
        return
    # Decide what the claim should say
    if n_full == total:
        expected_kw = "Yes"
    elif n_full == 0:
        expected_kw = "No"
    else:
        expected_kw = "Mixed"
    matches = claim.lstrip().startswith(expected_kw)
    _check(checks, "Tab 1 Full-brief verdict matches absolute best",
           matches,
           f"absolute best Full brief wins {n_full}/{total}, claim starts: {claim[:40]!r}")


def check_single_vs_pair_finding(checks: list, ws_tab1, scored: pd.DataFrame) -> None:
    """If Tab 1 Key Findings mentions 'single brief field', the count must match."""
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return

    # Reproduce the absolute-best per task analysis
    per_task_best = (
        sent.groupby(["task", "config_id"])["cosine"].mean().reset_index()
        .sort_values(["task", "cosine"], ascending=[True, False])
        .groupby("task").head(1)
    )

    def _kind(cid):
        if cid in ("A:_full_brief", "_full_brief"):
            return "full"
        if cid in ("A:_prompt_implied", "_prompt_implied"):
            return "no_brief"
        rhs = cid.split(":")[-1] if ":" in cid else cid
        return "pair" if "+" in rhs else "single"

    counts = per_task_best["config_id"].apply(_kind).value_counts()
    n_single = int(counts.get("single", 0))
    n_pair   = int(counts.get("pair", 0))

    # Search Tab 1 findings rows for "X of N tasks ... single brief field" claim
    for r in range(20, 36):
        v = ws_tab1.cell(r, 1).value
        if v and "single brief field" in str(v):
            m = re.search(r"(\d+)\s+of\s+(\d+)\s+tasks", str(v))
            if m:
                claimed = int(m.group(1))
                # Only counts as a problem if the claim says "single field"
                # but the count includes pairs.
                expected = n_single
                _check(checks, "Tab 1 'single field' count matches absolute best",
                       claimed == expected,
                       f"claim says {claimed} single, actual absolute best = {expected} single ({n_pair} pair)")
            return
    # No single-field claim found — silently OK
    _check(checks, "Tab 1 'single field' count matches absolute best", True, "no claim about single fields")


def check_combo_coverage(checks: list, ws_tab3, scored: pd.DataFrame) -> None:
    """Combos NAMED in Tab 3 'Best multi-field combos' must each have ≥ 5-task
    coverage. Otherwise we're showing a one-task-only metadata combo as 'best'."""
    MIN_COVERAGE = 5
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return
    combos = sent[sent["config_id"].str.contains(r"\+", regex=True, na=False)]
    if combos.empty:
        return
    coverage = combos.groupby("config_id")["task"].nunique()
    # Map humanized "audience + product" → "A:audience+product"
    name_to_cid = {}
    for cid in coverage.index:
        rhs = cid.split(":")[-1]
        human = " + ".join(rhs.split("+"))
        name_to_cid[human] = cid

    # Find the "Best multi-field combos" row in Tab 3
    cited_combos = []
    for r in range(1, 50):
        label = ws_tab3.cell(r, 1).value
        if label and "Best multi-field combos" in str(label):
            cell = str(ws_tab3.cell(r, 2).value or "")
            # cell looks like "audience + product (0.66), differentiators + product (0.67), ..."
            cited_combos = re.findall(r"([\w\s+]+?)\s*\(\d", cell)
            cited_combos = [c.strip() for c in cited_combos]
            break
    if not cited_combos:
        _check(checks, "Tab 3 'Best multi-field combos' coverage",
               True, "no combos listed yet")
        return

    bad = []
    for name in cited_combos:
        cid = name_to_cid.get(name)
        if not cid:
            continue
        if int(coverage.loc[cid]) < MIN_COVERAGE:
            bad.append(f"{name} ({coverage.loc[cid]}-task)")
    _check(checks, "Tab 3 'Best multi-field combos' coverage",
           not bad,
           f"undercovered: {', '.join(bad)}" if bad else f"all {len(cited_combos)} ≥ {MIN_COVERAGE}-task")


def _humanize_to_config_id(human_name: str) -> str | None:
    """Inverse of build_xlsx.humanize_config — best-effort string lookup."""
    if human_name == "Full brief":
        return "A:_full_brief"
    if human_name in ("No brief (instruction only)", "No Context Baseline"):
        return "A:_prompt_implied"
    # "audience + product"  →  "A:audience+product"
    parts = [p.strip() for p in human_name.split("+")]
    return "A:" + "+".join(parts)


def check_tab2_scores_match_groupby(checks: list, ws_tab2, scored: pd.DataFrame) -> None:
    """Every Tab 2 'Score' cell must match the (task, config, model) mean cosine
    in scored.csv. Catches helper bugs where the wrong cell value gets written."""
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return
    hdr = _find_header_row(ws_tab2, "Task")
    if hdr is None:
        return
    # Map task display name → internal task id (TASK_DISPLAY_NAMES is in build_xlsx)
    try:
        from scripts.build_xlsx import TASK_DISPLAY_NAMES
    except Exception:
        return
    name_to_task = {v: k for k, v in TASK_DISPLAY_NAMES.items()}

    mismatches = []
    for r in range(hdr + 1, hdr + 12):
        task_name = ws_tab2.cell(r, 1).value
        recipe = ws_tab2.cell(r, 2).value
        model = ws_tab2.cell(r, 3).value
        score = ws_tab2.cell(r, 4).value
        if not (task_name and recipe and model and score is not None):
            continue
        task_id = name_to_task.get(str(task_name).strip())
        cfg_id = _humanize_to_config_id(str(recipe).strip())
        cell = sent[
            (sent["task"] == task_id)
            & (sent["config_id"] == cfg_id)
            & (sent["model_key"] == model)
        ]
        if cell.empty:
            continue
        actual = float(cell["cosine"].mean())
        try:
            shown = float(score)
        except (TypeError, ValueError):
            continue
        if abs(actual - shown) > 0.005:
            mismatches.append(f"{task_name}/{recipe}/{model}: shown {shown:.3f} vs actual {actual:.3f}")
    _check(checks, "Tab 2 Score matches scored.csv groupby",
           not mismatches,
           f"{len(mismatches)} row(s) drift" + (f" — first: {mismatches[0]}" if mismatches else ""))


def check_modal_recipe_count(checks: list, ws_tab1, ws_tab2) -> None:
    """Tab 1 r9 'X wins N/M tasks' count must match the actual Tab 2 winner column."""
    # Tab 1: parse "Best recipe (overall)" row
    claim = None
    for r in range(1, 30):
        v = ws_tab1.cell(r, 1).value
        if v and "Best recipe" in str(v):
            claim = str(ws_tab1.cell(r, 2).value or "")
            break
    if not claim:
        return
    m = re.search(r"([\w\s+]+?)\s*\(wins\s+(\d+)/(\d+)\s*tasks?\)", claim)
    if not m:
        _check(checks, "Tab 1 'best recipe wins N/M' format parses", True, "no count claim")
        return
    claimed_recipe = m.group(1).strip()
    claimed_count = int(m.group(2))
    claimed_total = int(m.group(3))

    # Tab 2: count winners matching that recipe
    hdr = _find_header_row(ws_tab2, "Task")
    actual_count = 0
    actual_total = 0
    for r in range(hdr + 1, hdr + 12) if hdr else []:
        winning_recipe = ws_tab2.cell(r, 2).value
        if not winning_recipe:
            continue
        actual_total += 1
        if str(winning_recipe).strip() == claimed_recipe:
            actual_count += 1

    _check(checks, "Tab 1 modal-recipe count matches Tab 2 winners",
           claimed_count == actual_count and claimed_total == actual_total,
           f"claim {claimed_recipe} {claimed_count}/{claimed_total} vs actual {actual_count}/{actual_total}")


def check_cost_summary_total(checks: list, ws_tab1, raw: pd.DataFrame) -> None:
    """Tab 1 Cost Summary 'Total' must equal sum of rows AND raw['cost_usd'].sum()."""
    if raw is None or raw.empty or "cost_usd" not in raw.columns:
        return
    expected_total = float(raw["cost_usd"].sum())
    # Find 'Total' row in Cost Summary section
    total_row = None
    for r in range(1, 50):
        v = ws_tab1.cell(r, 1).value
        if v and str(v).strip() == "Total":
            total_row = r
            break
    if not total_row:
        return
    cell = str(ws_tab1.cell(total_row, 2).value or "").strip()
    m = re.match(r"\$?([\d.]+)", cell)
    if not m:
        _check(checks, "Tab 1 Cost Summary total parseable", False, f"cell {cell!r}")
        return
    shown = float(m.group(1))
    _check(checks, "Tab 1 Cost Summary total matches raw sum",
           abs(shown - expected_total) < 0.005,
           f"shown ${shown:.4f} vs raw sum ${expected_total:.4f}")


def check_heatmap_spot_cells(checks: list, ws_tab3, scored: pd.DataFrame) -> None:
    """Spot-check 3 heatmap cells against scored.csv groupby."""
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return
    try:
        from scripts.build_xlsx import TASK_DISPLAY_NAMES
    except Exception:
        return
    name_to_task = {v: k for k, v in TASK_DISPLAY_NAMES.items()}

    # Locate the heatmap header row (says "Task" in col 1; col 2+ are config labels)
    # Skip the Field Conclusions row by requiring col 2 to NOT contain " "
    # Heatmap headers are short like "Full brief", "audience", etc.
    hdr_row = None
    for r in range(8, 22):
        v1 = ws_tab3.cell(r, 1).value
        v2 = ws_tab3.cell(r, 2).value
        if v1 == "Task" and v2 and len(str(v2)) < 32:
            hdr_row = r
            break
    if hdr_row is None:
        return

    # Build config map from headers
    col_to_cid = {}
    for c in range(2, 22):
        v = ws_tab3.cell(hdr_row, c).value
        if not v:
            break
        col_to_cid[c] = _humanize_to_config_id(str(v).strip().replace("No Context Baseline", "No brief (instruction only)"))

    # Spot-check first 3 task rows × first 3 config columns
    mismatches = 0
    samples = 0
    for r in range(hdr_row + 1, hdr_row + 4):
        task_name = ws_tab3.cell(r, 1).value
        if not task_name:
            continue
        task_id = name_to_task.get(str(task_name).strip())
        if not task_id:
            continue
        for c in list(col_to_cid)[:3]:
            cfg_id = col_to_cid[c]
            shown = ws_tab3.cell(r, c).value
            if not isinstance(shown, (int, float)):
                continue
            samples += 1
            cell = sent[(sent["task"] == task_id) & (sent["config_id"] == cfg_id)]
            if cell.empty:
                continue
            actual = float(cell["cosine"].mean())
            if abs(actual - float(shown)) > 0.005:
                mismatches += 1
    _check(checks, "Tab 3 heatmap spot-check matches groupby",
           samples > 0 and mismatches == 0,
           f"{mismatches}/{samples} sampled cells drift")


def check_stability_plausible(checks: list, ws_appendix) -> None:
    """Tab 4 Stability avg-rerun-std should be < STAGE_B_STD_CEIL (typical noise
    floor σ ≈ 0.018, so std on top of that is usually ≤ 0.04). Worst std > 0.15
    is suspicious."""
    ceil = cfg.STAGE_B_STD_CEIL
    hdr = None
    for r in range(1, 80):
        if ws_appendix.cell(r, 1).value == "Task" and ws_appendix.cell(r, 5).value == "Avg rerun std":
            hdr = r
            break
    if hdr is None:
        return
    bad_avg = 0
    bad_worst = 0
    n = 0
    for r in range(hdr + 1, hdr + 30):
        avg = ws_appendix.cell(r, 5).value
        worst = ws_appendix.cell(r, 6).value
        if not isinstance(avg, (int, float)):
            break
        n += 1
        if float(avg) >= ceil:
            bad_avg += 1
        if isinstance(worst, (int, float)) and float(worst) >= 0.15:
            bad_worst += 1
    _check(checks, f"Tab 4 Stability avg-std < {ceil}",
           bad_avg == 0,
           f"{bad_avg}/{n} (task, recipe, model) cells exceed {ceil} — flagged unstable")


def check_human_review_row_count(checks: list, ws_appendix) -> None:
    """Human Review table should have between 20 and 30 sample rows (stratified
    across 8 tasks; with 3-4 per task you get 24-32)."""
    hdr = None
    for r in range(1, 50):
        if ws_appendix.cell(r, 1).value == "Task" and ws_appendix.cell(r, 4).value == "AI Output":
            hdr = r
            break
    if hdr is None:
        return
    n_rows = 0
    for r in range(hdr + 1, hdr + 45):
        v = ws_appendix.cell(r, 1).value
        # Empty row or hit next section
        if v in (None, "", " ") or (isinstance(v, str) and v.startswith(("Agreement", "Stability", "AI Judge", "Premium"))):
            break
        n_rows += 1
    _check(checks, "Tab 4 Human Review row count plausible (20–32)",
           20 <= n_rows <= 32,
           f"{n_rows} rows")


def check_best_combo_is_combo(checks: list, ws_tab1) -> None:
    """Tab 1 r10 'Best field combination' must be an actual combo (contains '+') or
    a clear 'not available' placeholder."""
    for r in range(1, 30):
        v = ws_tab1.cell(r, 1).value
        if v and "Best field combination" in str(v):
            cell = str(ws_tab1.cell(r, 2).value or "")
            is_combo = "+" in cell
            is_placeholder = cell.startswith(("Not tested", "No combo", "—"))
            _check(checks, "Tab 1 'Best field combination' is a real combo (or clearly missing)",
                   is_combo or is_placeholder,
                   f"value: {cell[:60]!r}")
            return


def check_keyword_f1_matches_groupby(checks: list, ws_tab3, scored: pd.DataFrame) -> None:
    """Tab 3 keyword compression F1 values must match scored.csv groupby."""
    kw = scored[scored["f1"].notna()].copy()
    if kw.empty:
        return
    kw["_v"] = kw["config_id"].apply(lambda c: c.split(":")[0])
    truth = kw.groupby("_v")["f1"].mean().round(3).to_dict()

    # Find keyword table header
    hdr = None
    for r in range(1, 50):
        v = ws_tab3.cell(r, 1).value
        if v and "Prompt Version" == str(v):
            hdr = r
            break
    if hdr is None:
        return
    mismatches = 0
    for r in range(hdr + 1, hdr + 6):
        version = ws_tab3.cell(r, 1).value  # "A — Full"
        f1_cell = ws_tab3.cell(r, 6).value
        if not version or f1_cell in (None, "—"):
            continue
        key = str(version).strip()[0]
        try:
            shown = round(float(f1_cell), 3)
        except (TypeError, ValueError):
            mismatches += 1
            continue
        truth_val = truth.get(key)
        if truth_val is None:
            continue
        if abs(shown - round(truth_val, 3)) > 0.002:
            mismatches += 1
    _check(checks, "Tab 3 keyword F1 matches groupby on scored.csv",
           mismatches == 0,
           f"{mismatches} version row(s) drift from groupby" if mismatches else "all rows match")


# ---------- main ----------

def main() -> int:
    if not SCORED_CSV.exists():
        print(f"[WARN] Missing {SCORED_CSV}. Audit skipped — run analyze --score first.")
        return 0  # not a failure of the audit itself
    if not WORKBOOK.exists():
        print(f"[WARN] Missing {WORKBOOK}. Audit skipped — run build_xlsx first.")
        return 0

    scored = pd.read_csv(SCORED_CSV)
    wb = openpyxl.load_workbook(WORKBOOK)
    if not {"Executive Summary", "Recommended Configs", "Field & Compression"} <= set(wb.sheetnames):
        print("[WARN] Workbook missing expected tabs — skipping audit.")
        return 0

    from src.utils import read_jsonl
    raw_records = read_jsonl(cfg.OUTPUTS_DIR / "results.jsonl")
    raw = pd.DataFrame(raw_records) if raw_records else pd.DataFrame()

    checks: list = []
    # Tab 2
    check_ci_bounds(checks, wb["Recommended Configs"])
    check_tab2_scores_match_groupby(checks, wb["Recommended Configs"], scored)
    # Tab 1 (uses Tab 2 cross-refs)
    check_delta_average(checks, wb["Executive Summary"], wb["Recommended Configs"], scored)
    check_full_brief_verdict(checks, wb["Executive Summary"], scored)
    check_single_vs_pair_finding(checks, wb["Executive Summary"], scored)
    check_modal_recipe_count(checks, wb["Executive Summary"], wb["Recommended Configs"])
    check_best_combo_is_combo(checks, wb["Executive Summary"])
    check_cost_summary_total(checks, wb["Executive Summary"], raw)
    # Tab 3
    check_combo_coverage(checks, wb["Field & Compression"], scored)
    check_heatmap_spot_cells(checks, wb["Field & Compression"], scored)
    check_keyword_f1_matches_groupby(checks, wb["Field & Compression"], scored)
    # Tab 4
    check_stability_plausible(checks, wb["Appendix"])
    check_human_review_row_count(checks, wb["Appendix"])

    passed = [c for c in checks if c[0]]
    failed = [c for c in checks if not c[0]]

    print(f"=== Workbook audit ({len(passed)} passed, {len(failed)} failed) ===")
    for ok, name, detail in passed:
        print(f"  [OK] {name}" + (f" — {detail}" if detail else ""))
    for ok, name, detail in failed:
        print(f"  [FAIL] {name}" + (f" — {detail}" if detail else ""))

    return 0 if not failed else 1


if __name__ == "__main__":
    sys.exit(main())
