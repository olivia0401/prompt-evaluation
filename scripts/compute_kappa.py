"""
Compute agreement score (Cohen's weighted kappa) between human and AI scorer.

Reads the Human 1-5 column (col 8) and the Sonnet 1-5 column (col 7) from the
Human-Review table in the Appendix tab, computes the agreement score, and
prints the decision (HIGH / MEDIUM / LOW).

The Human-Review table moved to the Appendix tab in the lean 4-tab workbook
restructure (CP2, 2026-05-20). Anchor row is found by header text "Human 1-5".

Usage:
    python -m scripts.compute_kappa
"""
import sys
from pathlib import Path

import openpyxl

from src import config as cfg

XLSX_PATH = cfg.RESULTS_DIR / "Prompt Eval Results.xlsx"
SHEET_NAME = "Appendix"
SONNET_COL = 7   # column G — "Sonnet 1-5"
HUMAN_COL = 8    # column H — "Human 1-5"
HUMAN_HEADER_TEXT = "Human 1-5"


def _cohen_weighted_kappa(rater_a: list[int], rater_b: list[int]) -> float:
    """
    Cohen's weighted kappa with quadratic weights.

    Standard formula:
        kappa = 1 - sum(w_ij * O_ij) / sum(w_ij * E_ij)
    where w_ij = (i - j)^2 / (k - 1)^2 for k categories.
    """
    if len(rater_a) != len(rater_b) or not rater_a:
        return float("nan")

    categories = sorted(set(rater_a) | set(rater_b))
    k = len(categories)
    if k < 2:
        return 1.0  # all identical scores

    idx = {c: i for i, c in enumerate(categories)}

    # Observed matrix
    n = len(rater_a)
    O = [[0] * k for _ in range(k)]
    for a, b in zip(rater_a, rater_b):
        O[idx[a]][idx[b]] += 1

    # Marginal counts → expected matrix (chance agreement)
    row_tot = [sum(row) for row in O]
    col_tot = [sum(O[r][c] for r in range(k)) for c in range(k)]
    E = [[row_tot[r] * col_tot[c] / n for c in range(k)] for r in range(k)]

    # Quadratic weights
    denom_w = (k - 1) ** 2
    W = [[((i - j) ** 2) / denom_w for j in range(k)] for i in range(k)]

    num = sum(W[i][j] * O[i][j] for i in range(k) for j in range(k))
    den = sum(W[i][j] * E[i][j] for i in range(k) for j in range(k))
    if den == 0:
        return float("nan")
    return 1 - num / den


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing {XLSX_PATH}. Run `python -m scripts.build_xlsx` first.")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found in {XLSX_PATH}.")
    ws = wb[SHEET_NAME]

    # Locate the header row by searching for the Human-rating column header.
    # Search a wider row range because the table is below several other sections
    # in the Appendix tab.
    header_row = None
    for r in range(1, 60):
        for c in range(1, 12):
            val = ws.cell(r, c).value
            if isinstance(val, str) and HUMAN_HEADER_TEXT in val:
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        raise SystemExit(
            f"Couldn't find the '{HUMAN_HEADER_TEXT}' header in the "
            f"'{SHEET_NAME}' tab. Did the workbook structure change?"
        )

    # Collect paired scores starting just below the header
    sonnet_scores, human_scores = [], []
    r = header_row + 1
    while True:
        first_cell = ws.cell(r, 1).value
        if not first_cell:
            break  # empty row = end of data
        s = ws.cell(r, SONNET_COL).value
        h = ws.cell(r, HUMAN_COL).value
        if isinstance(s, (int, float)) and isinstance(h, (int, float)):
            sonnet_scores.append(int(round(s)))
            human_scores.append(int(round(h)))
        r += 1

    n_pairs = len(sonnet_scores)
    print(f"Found {n_pairs} rows with BOTH human and AI-scorer scores.")
    if n_pairs < 10:
        print("Need at least 10 paired scores to compute a reliable agreement score.")
        print("Fill more rows in Tab 3 col 8 (Human Score) and re-run.")
        return

    kappa = _cohen_weighted_kappa(sonnet_scores, human_scores)
    print(f"\nAgreement score (Cohen's weighted kappa): {kappa:.3f}")

    if kappa >= 0.7:
        decision = "HIGH — automated scorer broadly aligns with human judgement."
    elif kappa >= 0.4:
        decision = "MEDIUM — moderate alignment; collect more samples to confirm."
    else:
        decision = "LOW — AI scorer disagrees with humans; treat AI ratings with caution."
    print(f"Decision: {decision}")

    # Persist so build_xlsx Appendix tab can surface the value automatically.
    import json
    from datetime import datetime
    kappa_path = cfg.OUTPUTS_DIR / "kappa.json"
    kappa_path.write_text(json.dumps({
        "kappa": round(kappa, 4),
        "n_pairs": n_pairs,
        "decision": decision.split(" — ")[0],
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"-> {kappa_path}")


if __name__ == "__main__":
    main()
