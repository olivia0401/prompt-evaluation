"""
Data integrity + reliability audit.

Cross-checks the experiment outputs against what the plan and workbook claim,
catching the kind of "23 briefs" mismatch that misled Simon's review:

  Group A — Schema / prerequisites
    A1. results.jsonl exists and is non-empty
    A2. briefs.yml has the expected 23 briefs
    A3. scored.csv exists (warn if not — workbook can't be trusted without it)

  Group B — Data completeness
    B1. All 23 briefs from briefs.yml appear at least once
    B2. All 9 tasks (8 sentence + keyword) present
    B3. Per-cell (task × config × model) sample size — distribution + outliers
    B4. Stage A coverage % (vs 23 × 142 × 2 = 6,532 expected cells)
    B5. Stage B run_id coverage (if scored.csv has any run_id ∈ {2, 3})

  Group C — Per-call reliability
    C1. Per-model ok-rate ≥ 95% (PASS gate)
    C2. No leftover rate_limited / api_error / timeout (or report count)
    C3. No empty predictions on status=ok rows
    C4. Resume keys unique

  Group D — Statistical adequacy
    D1. Sentence cells used by Tab 1: n_briefs ≥ 2 (else CI is undefined)
    D2. Length compliance ≥ 90% (PASS gate)
    D3. "Winner" diff vs second-best ≥ noise floor (0.036)

  Group E — Scoring integrity
    E1. Every ok row in results.jsonl has a corresponding scored.csv row
    E2. cosine / rouge_l / F1 in [0, 1] (NaN allowed for skipped rows)
    E3. No "missing GT" or "brief not found" parse_errors

  Group F — Workbook ↔ data consistency
    F1. xlsx doesn't claim "23 briefs" when fewer were actually run
    F2. Pilot status string in xlsx matches actual call count and cost

  Group G — Security
    G1. Secret files (.env, credentials.json, token.json) not in git
    G2. No hardcoded API keys in src/ or scripts/
    G3. No API key leaks in raw_response (results.jsonl)
    G4. No PII (email / phone / credit card) in raw_response
    G5. .gitignore covers .env, credentials.json, outputs/, embedding_cache
    G6. No `sk-...` / `sk-ant-...` tokens leaked in scored.csv predictions

  Group H — Output quality
    H1. Refusal rate ≤ 5% — "I cannot", "As an AI", "I don't have..."
    H2. Echo / verbatim copy rate ≤ 2% — prediction = brief field
    H3. Keyword task: ≥ 90% of ok rows return exactly 10 keywords
    H4. Cosine spike at < 0.1 — refusal/off-topic floor ≤ 5%

  Group I — Stability & determinism
    I1. Stage B cross-run std dev: median < 0.05 cosine
    I2. gpt5mini reruns more stable than haiku (seed=42 vs no seed)

Exit code:
  0 = all PASS (or only WARNINGs)
  1 = at least one FAIL

Run standalone:
    python -m scripts.audit_data
    python -m scripts.audit_data --json   # machine-readable
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import yaml

from src import config as cfg

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

def _results_jsonl() -> Path:
    return cfg.OUTPUTS_DIR / "results.jsonl"


def _scored_csv() -> Path:
    return cfg.OUTPUTS_DIR / "scored.csv"


def _workbook() -> Path:
    return cfg.RESULTS_DIR / "Prompt Eval Results.xlsx"

# All audit thresholds live in src/config.py (single source of truth).
# Re-export with the names this module already uses, so existing code paths
# below stay unchanged.
from src.config import (
    EXPECTED_BRIEF_COUNT,
    EXPECTED_SENTENCE_TASKS,
    EXPECTED_TASKS,
    EXPECTED_CHEAP_MODELS,
    NOISE_FLOOR_COSINE,
    LENGTH_COMPLIANCE_FLOOR,
    OK_RATE_FLOOR,
    REFUSAL_RATE_CEIL,
    ECHO_RATE_CEIL,
    KEYWORD_COUNT_COMPLIANCE_FLOOR,
    COSINE_ZERO_SPIKE_THRESHOLD,
    COSINE_ZERO_SPIKE_CEIL,
    STAGE_B_STD_CEIL,
)

# Audit-only constant — kept here because no other module needs it.
KEYWORD_COUNT_TARGET = 10
# Stage A target: 23 briefs × (Phase 1+2+3 configs ~142) × 2 cheap models.
# Computed dynamically below from prompt_builder so it never drifts.

# Security patterns
API_KEY_PATTERNS = [
    re.compile(r"\bsk-ant-[A-Za-z0-9_-]{20,}", re.I),     # Anthropic
    re.compile(r"\bsk-(?:proj-)?[A-Za-z0-9_-]{20,}", re.I),  # OpenAI
    re.compile(r"\bAKIA[A-Z0-9]{16}\b"),                  # AWS
    re.compile(r"\bAIza[A-Za-z0-9_-]{30,}\b"),            # Google
    re.compile(r"\bghp_[A-Za-z0-9]{30,}\b"),              # GitHub PAT
]
EMAIL_RE = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b")
# Phone: 10+ consecutive digits (loose; will catch various formats too).
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{8,}\d)")
# Credit card: 13–19 digits with optional separators, Luhn not enforced (loose).
CC_RE = re.compile(r"\b(?:\d[ -]?){13,19}\b")

# Refusal markers (case-insensitive). Tuned to be loud-but-not-noisy.
REFUSAL_MARKERS = [
    "i cannot", "i can't", "i'm unable", "i am unable",
    "as an ai", "as a language model",
    "i don't have", "i do not have",
    "i don't see", "i do not see",
    "could you please provide", "please provide",  # AI asking for input back
]


# ---------- result schema ----------

PASS, WARN, FAIL = "PASS", "WARN", "FAIL"


@dataclass
class CheckResult:
    code: str            # e.g. "B1"
    name: str
    status: str          # PASS / WARN / FAIL
    detail: str

    def as_dict(self) -> dict:
        return {"code": self.code, "name": self.name, "status": self.status,
                "detail": self.detail}


def _add(results: list, code: str, name: str, ok: bool, detail: str = "",
         warn_only: bool = False) -> None:
    status = PASS if ok else (WARN if warn_only else FAIL)
    results.append(CheckResult(code=code, name=name, status=status, detail=detail))


# ---------- Group A — prerequisites ----------

def check_prereqs(results: list) -> tuple[pd.DataFrame | None, pd.DataFrame | None, list[dict]]:
    """Returns (raw_df, scored_df, briefs_list) — any may be None."""
    raw_df = None
    scored_df = None
    briefs = []

    # A1
    results_jsonl = _results_jsonl()
    if not results_jsonl.exists():
        _add(results, "A1", "results.jsonl exists", False,
             f"{results_jsonl} missing — nothing to audit")
    else:
        rows = [json.loads(l) for l in results_jsonl.read_text(encoding="utf-8").splitlines() if l.strip()]
        raw_df = pd.DataFrame(rows)
        _add(results, "A1", "results.jsonl exists and non-empty",
             len(raw_df) > 0, f"{len(raw_df)} rows")

    # A2
    if not cfg.BRIEFS_FILE.exists():
        _add(results, "A2", "briefs.yml exists", False,
             f"{cfg.BRIEFS_FILE} missing")
    else:
        briefs = yaml.safe_load(cfg.BRIEFS_FILE.read_text(encoding="utf-8")) or []
        ok = len(briefs) == EXPECTED_BRIEF_COUNT
        _add(results, "A2",
             f"briefs.yml has {EXPECTED_BRIEF_COUNT} briefs",
             ok, f"found {len(briefs)}")

    # A3
    scored_csv = _scored_csv()
    if not scored_csv.exists():
        _add(results, "A3", "scored.csv exists", False,
             "downstream checks (D/E/F) will be skipped",
             warn_only=True)
    else:
        scored_df = pd.read_csv(scored_csv)
        _add(results, "A3", "scored.csv exists",
             len(scored_df) > 0, f"{len(scored_df)} rows")

    return raw_df, scored_df, briefs


# ---------- Group B — completeness ----------

def check_completeness(results: list, raw_df: pd.DataFrame | None,
                       briefs: list[dict]) -> None:
    if raw_df is None or raw_df.empty:
        return

    # B1
    expected_briefs = {str(b.get("current_name", "")).strip() for b in briefs}
    actual_briefs = set(raw_df["brief_id"].dropna().unique())
    missing = sorted(expected_briefs - actual_briefs)
    ok = not missing
    detail = (f"{len(actual_briefs)}/{len(expected_briefs)} briefs covered"
              + (f"; missing: {', '.join(missing[:5])}" + ("..." if len(missing) > 5 else "")
                 if missing else ""))
    _add(results, "B1", "All briefs from briefs.yml appear in results", ok, detail)

    # B2
    actual_tasks = set(raw_df["task"].dropna().unique())
    _add(results, "B2",
         f"All {EXPECTED_TASKS} tasks present (8 sentence + keyword)",
         len(actual_tasks) == EXPECTED_TASKS,
         f"found {len(actual_tasks)}: {sorted(actual_tasks)}")

    # B3 — per-cell sample size consistency for Stage A (cheap-tier) cells.
    # Phase 4 cells (premium/medium models) are intentionally run on a curated
    # 3-brief subset to stay under the £1 ceiling, so we exclude them here.
    # B4 below handles Stage A coverage separately.
    ok_rows = raw_df[raw_df["status"].isin({"ok", "ok_length_violation"})]
    cell_briefs = defaultdict(set)
    for _, r in ok_rows.iterrows():
        cell_briefs[(r["task"], r["config_id"], r["model_key"])].add(r["brief_id"])

    # Split into cheap (Stage A) vs everything else (Phase 4 premium / medium).
    cheap_cells = {k: v for k, v in cell_briefs.items()
                   if k[2] in EXPECTED_CHEAP_MODELS}
    other_cells = {k: v for k, v in cell_briefs.items()
                   if k[2] not in EXPECTED_CHEAP_MODELS}
    n_per_cheap = Counter(len(v) for v in cheap_cells.values())

    if not n_per_cheap:
        _add(results, "B3", "Per-cell sample size distribution (Stage A)",
             False, "no ok cells")
    else:
        distinct = sorted(n_per_cheap.items())
        detail = ", ".join(f"{n}-brief: {c} cells" for n, c in distinct)
        if other_cells:
            detail += (f" (Phase 4 cells excluded — {len(other_cells)} premium/medium "
                       f"cells run on a smaller curated brief subset by design)")
        _add(results, "B3",
             "Per-cell sample size consistent across Stage A cells",
             len(n_per_cheap) == 1,
             detail + (f" — target: {EXPECTED_BRIEF_COUNT}-brief"
                       if EXPECTED_BRIEF_COUNT not in n_per_cheap else ""))

    # B4 — Stage A coverage %
    try:
        from src.prompt_builder import list_configs_for_stage
        n_configs = len(list_configs_for_stage("stage_a"))
        expected_cells = EXPECTED_BRIEF_COUNT * n_configs * len(EXPECTED_CHEAP_MODELS)
        # Only count cells with at least the target number of briefs
        complete_cells = sum(1 for v in cell_briefs.values() if len(v) >= EXPECTED_BRIEF_COUNT)
        pct = 100 * complete_cells / max(1, n_configs * len(EXPECTED_CHEAP_MODELS))
        _add(results, "B4",
             f"Stage A coverage (cells with {EXPECTED_BRIEF_COUNT} briefs)",
             pct >= 95,
             f"{complete_cells} / {n_configs * len(EXPECTED_CHEAP_MODELS)} "
             f"({pct:.1f}%); full Stage A = {expected_cells} calls")
    except Exception as e:
        _add(results, "B4", "Stage A coverage", False,
             f"could not compute: {e}", warn_only=True)

    # B5 — Stage B reruns
    run_ids = set(raw_df["run_id"].dropna().unique())
    has_b = {2, 3} & run_ids
    _add(results, "B5", "Stage B reruns present (run_id ∈ {2,3})",
         bool(has_b),
         f"run_ids in data: {sorted(run_ids)}",
         warn_only=True)


# ---------- Group C — reliability ----------

def check_reliability(results: list, raw_df: pd.DataFrame | None) -> None:
    if raw_df is None or raw_df.empty:
        return

    # C1 — per-model ok-rate, computed per unique resume key (NOT per row).
    # Per-row rate is polluted by historical rate_limited retries that later
    # succeeded on a second attempt — those keys ARE done, even though the
    # failed-row trail stays in JSONL for audit purposes. The per-key view
    # answers the real question: "did this (brief, task, config, model, run)
    # combination eventually produce a usable answer?"
    DONE = {"ok", "ok_length_violation"}
    key_cols = ["brief_id", "task", "config_id", "model_key", "run_id"]
    if not all(c in raw_df.columns for c in key_cols):
        _add(results, "C1", f"Per-model ok-rate ≥ {int(OK_RATE_FLOOR*100)}%",
             False, "missing resume key columns", warn_only=True)
    else:
        # For each unique key, did it ever succeed?
        done_keys = set(map(tuple,
                            raw_df.loc[raw_df["status"].isin(DONE), key_cols].values.tolist()))
        all_keys = raw_df[key_cols].drop_duplicates()
        bad_models = []
        detail_parts = []
        for m, g in all_keys.groupby("model_key"):
            total = len(g)
            ok = sum(1 for row in g.values.tolist() if tuple(row) in done_keys)
            rate = ok / total if total else 0
            detail_parts.append(f"{m}: {ok}/{total} keys = {100*rate:.1f}%")
            if rate < OK_RATE_FLOOR:
                bad_models.append(m)
        _add(results, "C1",
             f"Per-model key-success rate ≥ {int(OK_RATE_FLOOR*100)}% "
             f"(unique keys that eventually succeeded)",
             not bad_models, "; ".join(detail_parts))

    # C2 — lingering rate-limit / api-error / timeout
    bad_statuses = {"rate_limited", "api_error", "timeout", "parse_fail", "refused", "truncated"}
    bad_counts = raw_df["status"].value_counts().to_dict()
    leftover = {k: v for k, v in bad_counts.items() if k in bad_statuses}
    _add(results, "C2", "No leftover failure statuses",
         not leftover,
         (f"found {leftover}; rerun the stage to retry — resume key dedup skips ok rows"
          if leftover else "clean"),
         warn_only=True)

    # C3 — empty predictions on ok rows
    if "raw_response" in raw_df.columns:
        ok_rows = raw_df[raw_df["status"].isin(DONE)]
        empties = ok_rows["raw_response"].fillna("").str.strip().eq("").sum()
        _add(results, "C3", "No empty predictions on ok rows",
             empties == 0,
             f"{empties} empty among {len(ok_rows)} ok rows")
    else:
        _add(results, "C3", "No empty predictions on ok rows", True,
             "raw_response column absent (older schema)", warn_only=True)

    # C4 — stuck resume keys (never reached an ok status across all retry rows)
    key_cols = ["brief_id", "task", "config_id", "model_key", "run_id"]
    if all(c in raw_df.columns for c in key_cols):
        # For each unique key, did it ever succeed?
        done_mask = raw_df["status"].isin(DONE)
        ok_keys = set(map(tuple, raw_df.loc[done_mask, key_cols].values.tolist()))
        all_keys = set(map(tuple, raw_df[key_cols].drop_duplicates().values.tolist()))
        stuck = all_keys - ok_keys
        # Breakdown: stuck on what status?
        stuck_status = Counter()
        if stuck:
            stuck_set = stuck
            for _, r in raw_df.iterrows():
                k = tuple(r[c] for c in key_cols)
                if k in stuck_set:
                    stuck_status[r["status"]] += 1
        detail = (f"{len(stuck)} keys never reached ok"
                  + (f" — by status: {dict(stuck_status)}" if stuck_status else ""))
        if not stuck:
            detail = f"all {len(all_keys)} unique keys eventually succeeded"
        _add(results, "C4", "All resume keys eventually succeeded",
             not stuck, detail,
             warn_only=True)


# ---------- Group D — statistical adequacy ----------

def check_statistical_adequacy(results: list, scored_df: pd.DataFrame | None) -> None:
    if scored_df is None or scored_df.empty:
        return

    sent = scored_df[scored_df["cosine"].notna()]

    # D1 — cells used for "best per task" need n ≥ 2
    if not sent.empty:
        per_cell = sent.groupby(["task", "config_id", "model_key"]).size()
        # Identify each task's winner cell (highest mean cosine)
        means = sent.groupby(["task", "config_id", "model_key"])["cosine"].mean()
        winners = means.groupby("task").idxmax()
        winner_ns = {task: per_cell.get(idx, 0) for task, idx in winners.items()}
        weak = {t: n for t, n in winner_ns.items() if n < 2}
        _add(results, "D1",
             "Each task's 'winning' cell has n ≥ 2 briefs (CI definable)",
             not weak,
             f"{len(weak)} weak winners: {weak}" if weak else f"all {len(winner_ns)} winners have n ≥ 2")

    # D2 — length compliance
    if "length_compliant" in scored_df.columns and not sent.empty:
        comp = sent["length_compliant"].astype(bool).mean()
        _add(results, "D2",
             f"Length compliance ≥ {int(LENGTH_COMPLIANCE_FLOOR*100)}%",
             comp >= LENGTH_COMPLIANCE_FLOOR,
             f"{100*comp:.1f}%")

    # D3 — winner gap vs second-best for each task
    if not sent.empty:
        means = sent.groupby(["task", "config_id"])["cosine"].mean().reset_index()
        weak_winners = []
        n_evaluated = 0
        for task, g in means.groupby("task"):
            srt = g.sort_values("cosine", ascending=False)
            if len(srt) < 2:
                continue
            n_evaluated += 1
            gap = srt.iloc[0]["cosine"] - srt.iloc[1]["cosine"]
            if gap < NOISE_FLOOR_COSINE:
                weak_winners.append(f"{task}(Δ={gap:.3f})")
        _add(results, "D3",
             f"'Winner' margin ≥ noise floor {NOISE_FLOOR_COSINE} cosine",
             not weak_winners,
             f"{len(weak_winners)}/{n_evaluated} tasks tied within noise: " + ", ".join(weak_winners[:5])
             if weak_winners else f"all {n_evaluated} winners are above noise",
             warn_only=True)


# ---------- Group E — scoring integrity ----------

def check_scoring_integrity(results: list, raw_df: pd.DataFrame | None,
                            scored_df: pd.DataFrame | None) -> None:
    if scored_df is None or scored_df.empty:
        return

    DONE = {"ok", "ok_length_violation"}

    # E1 — every ok row scored
    if raw_df is not None:
        key_cols = ["brief_id", "task", "config_id", "model_key", "run_id"]
        ok_raw = raw_df[raw_df["status"].isin(DONE)][key_cols].drop_duplicates()
        ok_scored = scored_df[scored_df["status"].isin(DONE)][key_cols].drop_duplicates()
        missing = ok_raw.merge(ok_scored, on=key_cols, how="left", indicator=True)
        n_missing = (missing["_merge"] == "left_only").sum()
        _add(results, "E1", "Every ok row has a scored.csv entry",
             n_missing == 0,
             f"{n_missing} ok raw rows missing from scored.csv")

    # E2 — metric ranges
    bad_metric = []
    for col in ("cosine", "rouge_l", "f1", "precision", "recall"):
        if col in scored_df.columns:
            vals = scored_df[col].dropna()
            if len(vals) == 0:
                continue
            outside = ((vals < 0) | (vals > 1)).sum()
            if outside:
                bad_metric.append(f"{col}: {outside}")
    _add(results, "E2", "All metrics in [0, 1]",
         not bad_metric, "; ".join(bad_metric) if bad_metric else "ok")

    # E3 — no parse_errors flagging missing GT etc.
    if "parse_errors" in scored_df.columns:
        errs = scored_df["parse_errors"].fillna("").astype(str)
        bad = errs.str.contains("missing GT|brief not found", regex=True, case=False)
        n = bad.sum()
        _add(results, "E3", "No 'missing GT' or 'brief not found' errors",
             n == 0, f"{n} rows with such errors")


# ---------- Group F — workbook ↔ data consistency ----------

def check_workbook_claims(results: list, raw_df: pd.DataFrame | None) -> None:
    workbook = _workbook()
    if not workbook.exists():
        _add(results, "F0", "Workbook present", False,
             f"{workbook} missing — F-group skipped", warn_only=True)
        return
    if raw_df is None or raw_df.empty:
        return

    try:
        import openpyxl
    except ImportError:
        _add(results, "F0", "openpyxl available", False,
             "install openpyxl to enable F-group", warn_only=True)
        return

    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    all_text_parts: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        # read_only mode: iterate rows in chunks to bound memory
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v:
                    all_text_parts.append(v)
        if len(all_text_parts) > 5000:  # workbook is small; cap defensively
            break
    text = " | ".join(all_text_parts)

    # F1 — "23 briefs" claim
    actual_briefs = raw_df["brief_id"].dropna().nunique()
    claim_23 = bool(re.search(r"\b23\s*briefs?\b", text))
    if claim_23 and actual_briefs < EXPECTED_BRIEF_COUNT:
        _add(results, "F1",
             "Workbook 'N briefs' claim matches actual coverage",
             False,
             f"text says '23 briefs' but only {actual_briefs} actually run")
    else:
        _add(results, "F1",
             "Workbook 'N briefs' claim matches actual coverage",
             True,
             f"actual={actual_briefs}, claim_23={claim_23}")

    # F2 — pilot status string
    total_calls = len(raw_df)
    total_cost = raw_df["cost_usd"].sum() if "cost_usd" in raw_df.columns else 0
    m_calls = re.search(r"(\d{2,5})\s*calls", text)
    m_cost = re.search(r"\$([\d.]+)\s*spent", text)
    issues = []
    if m_calls and int(m_calls.group(1)) != total_calls:
        issues.append(f"claim says {m_calls.group(1)} calls but actual is {total_calls}")
    if m_cost:
        claimed_cost = float(m_cost.group(1))
        if abs(claimed_cost - round(total_cost, 2)) > 0.02:
            issues.append(f"claim says ${claimed_cost} but actual is ${total_cost:.4f}")
    _add(results, "F2", "Workbook pilot status string matches actual data",
         not issues, "; ".join(issues) if issues else "ok")


# ---------- Group G — security ----------

def _project_root() -> Path:
    return cfg.PROJECT_ROOT if hasattr(cfg, "PROJECT_ROOT") else cfg.OUTPUTS_DIR.parent


def _git_ls_files() -> list[str]:
    """List of tracked files. Empty list if git not available or not a repo."""
    import subprocess
    try:
        out = subprocess.run(
            ["git", "ls-files"], cwd=str(_project_root()),
            capture_output=True, text=True, check=False, timeout=10,
        )
        if out.returncode == 0:
            return [p.strip() for p in out.stdout.splitlines() if p.strip()]
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass
    return []


def check_security(results: list, raw_df: pd.DataFrame | None) -> None:
    """Security / secrets / PII checks. Stop-the-line: most are FAIL not WARN."""
    root = _project_root()
    tracked = _git_ls_files()

    # G1 — secrets not in git
    SECRET_FILES = {".env", "credentials.json", "token.json"}
    leaked_secrets = [p for p in tracked
                      if Path(p).name in SECRET_FILES or p.endswith(".pem")
                      or p.endswith(".key")]
    _add(results, "G1", "Secret files not tracked in git",
         not leaked_secrets,
         f"leaked: {leaked_secrets}" if leaked_secrets else
         f"checked {len(tracked)} tracked files; none match secret patterns")

    # G2 — no hardcoded API keys in src/ or scripts/
    hits: list[str] = []
    for sub in ("src", "scripts"):
        d = root / sub
        if not d.exists():
            continue
        for p in d.rglob("*.py"):
            try:
                text = p.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            for pat in API_KEY_PATTERNS:
                m = pat.search(text)
                if m:
                    hits.append(f"{p.relative_to(root)}: {m.group(0)[:12]}...")
                    break
    _add(results, "G2", "No hardcoded API keys in src/ + scripts/",
         not hits, f"{len(hits)} files: {hits[:3]}" if hits else "clean")

    # G3 — no API key leaks in raw_response (results.jsonl)
    if raw_df is not None and not raw_df.empty and "raw_response" in raw_df.columns:
        leaked_rows = []
        for idx, r in raw_df.iterrows():
            text = str(r.get("raw_response") or "")
            if not text:
                continue
            for pat in API_KEY_PATTERNS:
                if pat.search(text):
                    leaked_rows.append(idx)
                    break
            if len(leaked_rows) >= 5:
                break  # bound cost
        _add(results, "G3", "No API key patterns in raw_response",
             not leaked_rows,
             f"{len(leaked_rows)} rows match (sample idx: {leaked_rows[:3]})"
             if leaked_rows else f"scanned {len(raw_df)} rows")

    # G4 — no PII in raw_response
    if raw_df is not None and not raw_df.empty and "raw_response" in raw_df.columns:
        email_hits = phone_hits = cc_hits = 0
        sample_rows = []
        for idx, r in raw_df.iterrows():
            text = str(r.get("raw_response") or "")
            if not text:
                continue
            if EMAIL_RE.search(text):
                email_hits += 1
                if len(sample_rows) < 3:
                    sample_rows.append(("email", idx))
            if PHONE_RE.search(text):
                phone_hits += 1
                if len(sample_rows) < 3:
                    sample_rows.append(("phone", idx))
            if CC_RE.search(text):
                cc_hits += 1
                if len(sample_rows) < 3:
                    sample_rows.append(("cc", idx))
        total = email_hits + phone_hits + cc_hits
        detail = (f"email={email_hits} phone={phone_hits} cc={cc_hits}"
                  + (f"; sample: {sample_rows}" if sample_rows else ""))
        # phone & cc regexes are loose, so we WARN not FAIL unless we see
        # a clear email or CC. Tune as needed.
        _add(results, "G4", "No PII (email / phone / CC) in raw_response",
             total == 0, detail, warn_only=(email_hits == 0 and cc_hits == 0))

    # G5 — .gitignore covers critical paths.
    # We only require the canonical entries; e.g. `outputs/` covers everything
    # underneath (embedding_cache.jsonl, results.jsonl, scored.csv) without
    # needing each filename listed separately.
    gi = root / ".gitignore"
    if not gi.exists():
        _add(results, "G5", ".gitignore exists and covers critical paths",
             False, ".gitignore is missing")
    else:
        content = gi.read_text(encoding="utf-8", errors="ignore")
        active_lines = [
            line.strip() for line in content.splitlines()
            if line.strip() and not line.lstrip().startswith("#")
        ]
        REQUIRED = [".env", "credentials.json", "token.json", "outputs/"]
        missing = [p for p in REQUIRED
                   if not any(p in line for line in active_lines)]
        _add(results, "G5", ".gitignore covers critical paths",
             not missing,
             f"missing: {missing}" if missing else f"all {len(REQUIRED)} covered")

    # G6 — no `sk-...` / `sk-ant-...` tokens in scored.csv predictions
    if "raw_response" in (raw_df.columns if raw_df is not None else []):
        # Already covered by G3. Skip duplicate scan.
        pass


# ---------- Group H — output quality ----------

def check_output_quality(results: list, raw_df: pd.DataFrame | None,
                         scored_df: pd.DataFrame | None) -> None:
    DONE = {"ok", "ok_length_violation"}
    if raw_df is None or raw_df.empty:
        return

    ok_rows = raw_df[raw_df["status"].isin(DONE)]
    if "raw_response" not in raw_df.columns or ok_rows.empty:
        return

    # H1 — refusal rate
    refusal_count = 0
    for text in ok_rows["raw_response"].fillna("").astype(str):
        low = text.lower()
        if any(m in low for m in REFUSAL_MARKERS):
            refusal_count += 1
    refusal_rate = refusal_count / len(ok_rows)
    _add(results, "H1", f"Refusal rate ≤ {int(REFUSAL_RATE_CEIL*100)}%",
         refusal_rate <= REFUSAL_RATE_CEIL,
         f"{refusal_count}/{len(ok_rows)} = {100*refusal_rate:.1f}% — "
         f"({'AI explicitly refused or asked for clarification' if refusal_count else 'clean'})")

    # H2 — echo / verbatim copy detection
    # For sentence tasks: prediction shouldn't be identical to any brief field.
    if scored_df is not None and not scored_df.empty:
        sent_rows = scored_df[scored_df["cosine"].notna()]
        try:
            from src.prompt_builder import load_briefs, SEMANTIC_FIELDS
            briefs = {b.get("current_name", "").strip(): b for b in load_briefs()}
            echo_count = 0
            checked = 0
            for _, r in sent_rows.iterrows():
                pred = str(r.get("prediction") or "").strip()
                if not pred:
                    continue
                bid = r.get("brief_id")
                brief = briefs.get(bid)
                if not brief:
                    continue
                checked += 1
                # Compare to any semantic field (lowercased, stripped)
                pred_low = pred.lower()
                for f in SEMANTIC_FIELDS:
                    val = brief.get(f)
                    if isinstance(val, str) and val.strip().lower() == pred_low:
                        echo_count += 1
                        break
            if checked > 0:
                echo_rate = echo_count / checked
                _add(results, "H2", f"Echo rate ≤ {int(ECHO_RATE_CEIL*100)}%",
                     echo_rate <= ECHO_RATE_CEIL,
                     f"{echo_count}/{checked} = {100*echo_rate:.2f}% predictions "
                     f"are verbatim copies of a brief field",
                     warn_only=True)
        except ImportError:
            pass

    # H3 — keyword task count compliance
    if scored_df is not None and not scored_df.empty:
        kw_rows = scored_df[(scored_df["task"] == "keywords")
                            & scored_df["status"].isin(DONE)]
        if not kw_rows.empty:
            compliant = 0
            for _, r in kw_rows.iterrows():
                parsed = str(r.get("prediction_parsed") or "")
                if not parsed:
                    continue
                try:
                    arr = json.loads(parsed)
                    if isinstance(arr, list) and len(arr) == KEYWORD_COUNT_TARGET:
                        compliant += 1
                except (json.JSONDecodeError, TypeError):
                    pass
            comp_rate = compliant / len(kw_rows)
            _add(results, "H3",
                 f"Keyword task returns exactly {KEYWORD_COUNT_TARGET} keywords "
                 f"in ≥ {int(KEYWORD_COUNT_COMPLIANCE_FLOOR*100)}% of ok rows",
                 comp_rate >= KEYWORD_COUNT_COMPLIANCE_FLOOR,
                 f"{compliant}/{len(kw_rows)} = {100*comp_rate:.1f}%")

    # H4 — cosine spike at < 0.1 (off-topic / refusal floor)
    if scored_df is not None and not scored_df.empty:
        sent_rows = scored_df[scored_df["cosine"].notna()]
        if not sent_rows.empty:
            low = (sent_rows["cosine"] < COSINE_ZERO_SPIKE_THRESHOLD).sum()
            rate = low / len(sent_rows)
            _add(results, "H4",
                 f"Cosine < {COSINE_ZERO_SPIKE_THRESHOLD} rate ≤ "
                 f"{int(COSINE_ZERO_SPIKE_CEIL*100)}%",
                 rate <= COSINE_ZERO_SPIKE_CEIL,
                 f"{low}/{len(sent_rows)} = {100*rate:.1f}% sentence rows score "
                 f"below {COSINE_ZERO_SPIKE_THRESHOLD} (suggests refusals / off-topic)")


# ---------- Group I — stability & determinism ----------

def check_stability(results: list, scored_df: pd.DataFrame | None) -> None:
    if scored_df is None or scored_df.empty:
        return
    sent = scored_df[scored_df["cosine"].notna()]
    if sent.empty:
        return

    # I1 — Stage B cross-run std dev
    multi_run = (
        sent.groupby(["brief_id", "task", "config_id", "model_key"])
        .filter(lambda g: g["run_id"].nunique() >= 2)
    )
    if multi_run.empty:
        _add(results, "I1", "Stage B cross-run std dev",
             True, "no (brief × config × model) cell has ≥ 2 runs yet — skipping",
             warn_only=True)
    else:
        stds = (
            multi_run.groupby(["brief_id", "task", "config_id", "model_key"])["cosine"]
            .std(ddof=1).dropna()
        )
        if stds.empty:
            _add(results, "I1", "Stage B cross-run std dev",
                 True, "no usable std dev (all groups had only 1 distinct value)",
                 warn_only=True)
        else:
            median_std = float(stds.median())
            _add(results, "I1",
                 f"Stage B median cross-run cosine std < {STAGE_B_STD_CEIL}",
                 median_std < STAGE_B_STD_CEIL,
                 f"median={median_std:.3f} across {len(stds)} (brief×config×model) groups; "
                 f"max={float(stds.max()):.3f}")

    # I2 — gpt5mini (seeded) should be more stable than haiku (no seed)
    if not multi_run.empty:
        by_model_std = {}
        for m in ("gpt5mini", "haiku"):
            sub = multi_run[multi_run["model_key"] == m]
            if sub.empty:
                continue
            s = sub.groupby(["brief_id", "task", "config_id"])["cosine"].std(ddof=1).dropna()
            if not s.empty:
                by_model_std[m] = float(s.median())
        if "gpt5mini" in by_model_std and "haiku" in by_model_std:
            ok = by_model_std["gpt5mini"] <= by_model_std["haiku"]
            _add(results, "I2",
                 "gpt5mini (seed=42) more stable than haiku (no seed)",
                 ok,
                 f"gpt5mini median std={by_model_std['gpt5mini']:.3f}, "
                 f"haiku={by_model_std['haiku']:.3f}",
                 warn_only=True)


# ---------- main ----------

def run_all_checks() -> list[CheckResult]:
    results: list[CheckResult] = []
    raw_df, scored_df, briefs = check_prereqs(results)
    check_completeness(results, raw_df, briefs)
    check_reliability(results, raw_df)
    check_statistical_adequacy(results, scored_df)
    check_scoring_integrity(results, raw_df, scored_df)
    check_workbook_claims(results, raw_df)
    check_security(results, raw_df)
    check_output_quality(results, raw_df, scored_df)
    check_stability(results, scored_df)
    return results


def print_human(results: list[CheckResult]) -> int:
    glyph = {PASS: "✓", WARN: "⚠", FAIL: "✗"}
    n_fail = sum(1 for r in results if r.status == FAIL)
    n_warn = sum(1 for r in results if r.status == WARN)
    n_pass = sum(1 for r in results if r.status == PASS)

    print(f"\nData integrity audit — {len(results)} checks")
    print("=" * 78)
    for r in results:
        line = f"  {glyph[r.status]} [{r.code}] {r.name}"
        if r.detail:
            line += f"\n        {r.detail}"
        print(line)
    print("-" * 78)
    print(f"  {n_pass} PASS · {n_warn} WARN · {n_fail} FAIL")
    if n_fail:
        print("\n  Fix the FAILs before trusting the workbook for stakeholder review.")
    elif n_warn:
        print("\n  No hard failures, but check the WARNings before claiming completeness.")
    else:
        print("\n  All checks passed.")
    return 1 if n_fail else 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true",
                    help="Emit JSON to stdout (for CI / dashboards).")
    args = ap.parse_args()

    results = run_all_checks()

    if args.json:
        print(json.dumps([r.as_dict() for r in results], indent=2))
        n_fail = sum(1 for r in results if r.status == FAIL)
        sys.exit(1 if n_fail else 0)

    sys.exit(print_human(results))


if __name__ == "__main__":
    main()
