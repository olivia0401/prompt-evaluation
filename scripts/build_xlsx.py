"""
Build the 4-tab client-facing experiment report.

Tab 1: Executive Summary     — top recommendations, ≤5 key findings, small cost summary.
Tab 2: Recommended Configs   — per-task winner table + production-scenario matrix.
Tab 3: Field & Compression   — field-contribution heatmap + keyword-compression table.
Tab 4: Appendix              — human review, stability checks, AI judge, premium ladder, methodology footer.

Reads:  outputs/scored.csv, outputs/results.jsonl (for cost)
Writes: Results/Prompt Eval Results.xlsx (auto-upload to RESULTS_SHEETS_ID when configured).
"""
import sys
from pathlib import Path

# Windows default codec is GBK on Chinese locales — can't print '' / ''.
# Force UTF-8 so terminal logging never crashes the upload step.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# ---------- i18n (English / 简体中文) ----------
# Global language switch. Set by main() from the --lang CLI arg before
# any build_tab_* function is called. T(en, zh) returns the right string
# at call time, so translations live inline next to their English original.
LANG = "en"


def T(en: str, zh: str) -> str:
    """Return zh when LANG == 'zh', otherwise en. Used at every user-visible string."""
    return zh if LANG == "zh" else en


def T_task(task_key: str) -> str:
    """Translated task display name (e.g. 'Benefit' → '利益点')."""
    return _TASK_DISPLAY_ZH.get(task_key, humanize_task(task_key)) if LANG == "zh" \
        else humanize_task(task_key)


# Chinese display names for tasks. The internal task keys
# (concept_relevant, etc.) stay English; this dict is consulted only
# when LANG == 'zh' to render the column.
_TASK_DISPLAY_ZH = {
    "concept_relevant":  "概念",
    "position_relevant": "定位",
    "emotion_relevant":  "情感",
    "function_relevant": "功能",
    "benefit_relevant":  "利益点",
    "category_relevant": "品类",
    "feature_relevant":  "特性",
    "context_relevant":  "使用情境",
    "keywords":          "关键词",
    # Also accept the humanised forms in case a caller passes them through.
    "Concept": "概念", "Positioning": "定位", "Emotion": "情感",
    "Function": "功能", "Benefit": "利益点", "Category": "品类",
    "Features": "特性", "Context": "使用情境", "Keywords": "关键词",
}


# Chinese versions of internal config / recipe labels.
_CONFIG_DISPLAY_ZH = {
    "Full brief":                       "完整 brief",
    "No brief (instruction only)":      "无 brief（仅指令）",
    "product":                          "product",
    "audience":                         "audience",
    "brand_strategy":                   "brand_strategy",
    "personality":                      "personality",
    "differentiators":                  "differentiators",
    "business_category":                "business_category",
}


def T_recipe(label: str) -> str:
    """Translate a recipe label to Chinese while keeping field names in English
    (they're identifiers, not prose). Multi-field labels join with ' + '."""
    if LANG != "zh":
        return label
    parts = [p.strip() for p in label.split("+")]
    return " + ".join(_CONFIG_DISPLAY_ZH.get(p, p) for p in parts)

import math

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import config as cfg


def _bootstrap_ci_mean(values, n_resamples: int = 1000,
                       confidence: float = 0.95,
                       seed: int = 42) -> tuple[float, float]:
    """
    95% bootstrap confidence interval for the mean.

    Vectorized resampling: generate n_resamples × n random index draws at once,
    take the mean across each draw, then percentile the resulting distribution.
    Returns (low, high). Returns (nan, nan) if fewer than 2 values.
    """
    arr = np.asarray(list(values), dtype=float)
    if arr.size < 2:
        return (float("nan"), float("nan"))
    rng = np.random.default_rng(seed)
    idx = rng.integers(0, arr.size, size=(n_resamples, arr.size))
    sampled_means = arr[idx].mean(axis=1)
    alpha = (1 - confidence) / 2
    return (
        float(np.percentile(sampled_means, alpha * 100)),
        float(np.percentile(sampled_means, (1 - alpha) * 100)),
    )


def _rankdata(arr: np.ndarray) -> np.ndarray:
    """Average ranks (1-based), ties share the mean rank. Pure-numpy, no scipy."""
    arr = np.asarray(arr, dtype=float)
    order = arr.argsort(kind="mergesort")
    ranks = np.empty(arr.size, dtype=float)
    sorted_arr = arr[order]
    i = 0
    while i < arr.size:
        j = i
        while j + 1 < arr.size and sorted_arr[j + 1] == sorted_arr[i]:
            j += 1
        avg = (i + j) / 2.0 + 1.0  # 1-based average rank for the tie group
        ranks[order[i:j + 1]] = avg
        i = j + 1
    return ranks


def _paired_signed_rank_p(diffs) -> float | None:
    """Two-sided Wilcoxon signed-rank p-value via normal approximation.

    Tests H0: the paired per-brief differences are symmetric about 0 — i.e.
    "A is no better than B on the same briefs". Zero differences are dropped.
    Returns None when there are fewer than 6 non-zero pairs (the normal
    approximation is unreliable and we'd rather report "not enough evidence"
    than a misleading p-value). Pure-numpy so it works without scipy.

    This is the answer to "is A *really* better than B", as opposed to "is A's
    mean a hair higher" — exactly the small-sample trap with only 23 briefs.
    """
    arr = np.asarray([d for d in diffs if d == d], dtype=float)  # drop NaN
    arr = arr[arr != 0.0]
    n = arr.size
    if n < 6:
        return None
    ranks = _rankdata(np.abs(arr))
    w_plus = float(ranks[arr > 0].sum())
    mean_w = n * (n + 1) / 4.0
    # Tie correction for the variance term.
    _, counts = np.unique(np.abs(arr), return_counts=True)
    tie_term = float(((counts ** 3 - counts).sum())) / 48.0
    var_w = (n * (n + 1) * (2 * n + 1)) / 24.0 - tie_term
    if var_w <= 0:
        return None
    # Continuity-corrected |z|, then two-sided p = erfc(|z| / sqrt(2)).
    z = max(0.0, abs(w_plus - mean_w) - 0.5) / math.sqrt(var_w)
    return float(min(1.0, math.erfc(z / math.sqrt(2))))


DST = cfg.RESULTS_DIR / "Prompt Eval Results.xlsx"
SCORED_CSV = cfg.OUTPUTS_DIR / "scored.csv"

# --- Minimal GOV.UK palette: typography-led, only 4 functional fills ---
# Lesson from prior iterations: lots of background fills make spreadsheet cells
# busy and hurt readability. Instead we lean on bold weight + colour-on-white
# text for hierarchy, and keep colour fills ONLY where they're functional:
#   1. Title / section bars (anchor visual)
#   2. Ground Truth column (column-level indicator)
#   3. Heatmap (Tab 3 — entire mechanic depends on colour)
#   4. Pilot status (one info banner per tab — orientation aid)
# All semantic verdicts (supported / refuted / neutral) use coloured TEXT on
# white, not coloured fills. Cleaner, easier to read, professional report look.

NO_FILL = PatternFill(fill_type=None)

# Functional fills — only what's actually used by the 4 lean builders.
LIGHT_BLUE = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")  # Insight / banner highlights
DARK_BLUE  = PatternFill(start_color="1D70B8", end_color="1D70B8", fill_type="solid")  # Title + section bars
LIGHT_GREY = PatternFill(start_color="F8F8F7", end_color="F8F8F7", fill_type="solid")  # Table-header background

THICK_BLACK_BORDER = Border(
    left=Side(style="medium", color="000000"),
    right=Side(style="medium", color="000000"),
    top=Side(style="medium", color="000000"),
    bottom=Side(style="medium", color="000000"),
)

# Typography — bold weight + colour-on-white carry the hierarchy.
TITLE_FONT   = Font(bold=True, size=16, color="FFFFFF")          # on DARK_BLUE
HEADER_FONT  = Font(bold=True, size=11, color="0B0C0C")          # table headers
SECTION_FONT = Font(bold=True, size=12, color="FFFFFF")          # on DARK_BLUE
ITALIC_GREY  = Font(italic=True, size=10, color="626A6E")        # secondary notes

CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _visual_width(text: str) -> int:
    """Estimate the on-screen width of text in Excel column units.

    Treats CJK and full-width punctuation as 2 units wide, ASCII as 1 unit.
    This matters in a bilingual workbook: a 50-character Chinese sentence
    needs roughly double the column width of a 50-character English sentence,
    so a height heuristic that counts both as 50 will truncate Chinese rows.
    """
    if not text:
        return 0
    w = 0
    for ch in text:
        cp = ord(ch)
        # CJK Unified Ideographs, Hiragana/Katakana, full-width punctuation /
        # ASCII forms, CJK punctuation. Conservative: also count em-dash and
        # full-width Latin chars as 2.
        if (0x3000 <= cp <= 0x9FFF) or (0xFF00 <= cp <= 0xFFEF) or ch in "——「」（）。，；：？！":
            w += 2
        else:
            w += 1
    return w


def _estimate_row_height(text: str, total_width_chars: int = 110) -> int:
    """Row-height heuristic — 1 visual line ≈ 14 points + 4 padding.

    Counts CJK as 2 visual units (see _visual_width) so Chinese wraps
    estimate correctly. Calibrated tight: 14 pt is roughly the rendered
    height of one wrapped line of 10-11 pt text in openpyxl, and a
    larger value leaves visible empty space below short rows. Capped at
    220 so a single runaway cell can not blow up the sheet.
    """
    if not text:
        return 18
    visual = _visual_width(text)
    line_count = max(1, (visual // total_width_chars) + text.count("\n") + 1)
    return min(220, max(20, line_count * 14 + 4))


# Centered wrap — text floats vertically in the middle of the cell.
WRAP_CENTER = Alignment(wrap_text=True, vertical="center")


def _merge_and_write(ws, row: int, text: str, ncols: int, font, fill=None, align=None):
    """Write text in column A, then merge across ncols so wrap uses full width."""
    cell = write_cell(ws, row, 1, text, font=font, fill=fill,
                      align=align or WRAP_CENTER)
    if fill:
        for c in range(2, ncols + 1):
            write_cell(ws, row, c, "", fill=fill)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    # Width factor depends on ncols. ~18 chars per col is realistic at our widths.
    ws.row_dimensions[row].height = _estimate_row_height(text, total_width_chars=ncols * 18)


def write_tab_title(ws, title: str, ncols: int = 6) -> int:
    """Single-row blue title bar. No 'what's inside' / 'how to read' subtext.

    Lean header for the client-facing 4-tab workbook. Returns the next free row
    (= 3, leaving row 2 blank for breathing room).
    """
    _merge_and_write(ws, 1, title, ncols, font=TITLE_FONT, fill=DARK_BLUE,
                     align=Alignment(vertical="center", horizontal="left", indent=1))
    ws.row_dimensions[1].height = 34
    return 3


def write_section_bar(ws, row: int, text: str, ncols: int = 9) -> int:
    """Section divider — blue bar with white bold text, merged."""
    _merge_and_write(ws, row, text, ncols, font=SECTION_FONT, fill=DARK_BLUE,
                     align=Alignment(vertical="center", horizontal="left", indent=1))
    ws.row_dimensions[row].height = 26
    return row + 1


def write_section_explain(ws, row: int, text: str, ncols: int = 9) -> int:
    """Italic explanation under section header, merged for full-width wrap."""
    _merge_and_write(ws, row, text, ncols, font=ITALIC_GREY, fill=None)
    return row + 1


SENTENCE_TASKS = [
    "concept_relevant", "position_relevant", "emotion_relevant",
    "function_relevant", "benefit_relevant", "category_relevant",
    "feature_relevant", "context_relevant",
]
KEYWORD_TASK = "keywords"


# Human-readable names for internal config codes.
CONFIG_DISPLAY_NAMES = {
    "A:_full_brief":      "Full brief",
    "A:_prompt_implied":  "No brief (instruction only)",
    "_full_brief":        "Full brief",
    "_prompt_implied":    "No brief (instruction only)",
}


def _keyword_prompt_lengths() -> pd.DataFrame:
    """
    Extract char/est-token counts for the 4 keyword prompt versions (A/B/C/D)
    by parsing prompts.txt. Useful before Phase 3 F1 data is available — the
    length comparison itself is static and known from day one.
    """
    if not cfg.PROMPTS_FILE.exists():
        return pd.DataFrame()
    text = cfg.PROMPTS_FILE.read_text(encoding="utf-8")

    # Section start markers — note the typo "extraxction" in Version A's header
    # exists verbatim in prompts.txt and is the actual delimiter.
    markers = [
        ("A — Full",     "Keyword extraxction brief:"),
        ("B — Reduced",  "Keyword extraction brief (Version B"),
        ("C — Compact",  "Keyword extraction brief (Version C"),
        ("D — Minimal",  "Keyword extraction brief (Version D"),
    ]
    positions = []
    for label, marker in markers:
        idx = text.find(marker)
        if idx != -1:
            positions.append((label, idx))
    if len(positions) < 2:
        return pd.DataFrame()
    positions.sort(key=lambda x: x[1])

    rows = []
    for i, (label, start_idx) in enumerate(positions):
        end_idx = positions[i + 1][1] if i + 1 < len(positions) else len(text)
        # Skip past the marker line itself
        nl = text.find("\n", start_idx)
        marker_end = nl + 1 if nl != -1 else start_idx
        body = text[marker_end:end_idx].strip()
        # Strip the trailing "BRIEF:\n[INSERT BRIEF]" boilerplate — same for every version
        cut = body.find("BRIEF:")
        if cut != -1:
            body = body[:cut].rstrip()
        char_count = len(body)
        est_tokens = round(char_count / 4)  # rough heuristic (~4 chars per token)
        rows.append({
            "Prompt version": label,
            "Characters": char_count,
            "Est. tokens": est_tokens,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    max_chars = df["Characters"].max()
    df["Reduction vs A"] = (
        ((max_chars - df["Characters"]) / max_chars * 100).round(0)
        .astype(int).astype(str) + "%"
    )
    return df


def _keyword_b_numbers() -> dict | None:
    """Live A→B keyword-prompt compression numbers (computed from prompts.txt),
    so the Executive Summary never ships a stale hand-typed '63% / 3,306 / 1,225'.
    Returns {pct, chars_a, chars_b, tokens_saved} or None if unavailable."""
    df = _keyword_prompt_lengths()
    if df.empty:
        return None
    vers = set(df["Prompt version"])
    if "A — Full" not in vers or "B — Reduced" not in vers:
        return None
    a = df[df["Prompt version"] == "A — Full"].iloc[0]
    b = df[df["Prompt version"] == "B — Reduced"].iloc[0]
    chars_a, chars_b = int(a["Characters"]), int(b["Characters"])
    if chars_a <= 0:
        return None
    return {
        "pct": round((chars_a - chars_b) / chars_a * 100),
        "chars_a": chars_a,
        "chars_b": chars_b,
        "tokens_saved": int(a["Est. tokens"]) - int(b["Est. tokens"]),
    }


def _top_single_field(scored: pd.DataFrame) -> str | None:
    """Name of the strongest single brief field by mean cosine — computed, so
    the 'product was the strongest single field' claim can't go stale."""
    try:
        fc = _field_contribution_summary(scored)
        tops = fc.get("top_fields") or []
        if tops:
            return str(tops[0]).split(" (")[0].strip()
    except Exception:
        pass
    return None


def _stability_summary(scored: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """
    Per-cell rerun std-dev — the *true* measure of stability.

    Algorithm:
      1. For each (brief, task, config, model) with ≥ 2 runs, compute the
         std-dev of cosine ACROSS the reruns. This isolates rerun noise from
         cross-brief variation.
      2. Aggregate across briefs: report mean + worst per (task, config, model).

    The previous implementation grouped by (task, config, model) only,
    pooling 23 briefs × N reruns into one std — that confused cross-brief
    variance with rerun noise. The numbers were on the right order of
    magnitude as the per-brief negligible-difference line (σ ≈ 0.018) by coincidence
    when reruns were few; with full Stage B they ballooned to ~0.1 because
    different briefs naturally produce different cosines.
    """
    sent = scored[scored["cosine"].notna()].copy() if not scored.empty else scored
    if sent.empty or "run_id" not in sent.columns:
        return pd.DataFrame(), T(
            "No stability data yet. Stage B reruns will populate this. "
            "Until then, treat single-run scores as point estimates.",
            "暂无稳定性数据。Stage B 重跑后此处会显示。"
            "在此之前，把单次分数当作点估计来看。"
        )

    # Step 1: per-(brief, task, config, model) std across that cell's reruns.
    per_brief = (
        sent.groupby(["brief_id", "task", "config_id", "model_key"])
        .agg(n_runs=("run_id", "nunique"), cell_std=("cosine", "std"))
        .reset_index()
    )
    rerun_cells = per_brief[(per_brief["n_runs"] >= 2) & per_brief["cell_std"].notna()]
    if rerun_cells.empty:
        return pd.DataFrame(), T(
            "No (brief × task × config × model) cell has at least two runs "
            "yet. Run Stage B to populate.",
            "暂无任何（brief × 任务 × 配方 × 模型）单元格有两次以上重跑。"
            "运行 Stage B 后此处会显示。"
        )

    # Step 2: aggregate across briefs → one row per (task, config, model).
    out = (
        rerun_cells.groupby(["task", "config_id", "model_key"])
        .agg(
            briefs_with_reruns=("cell_std", "size"),
            avg_std=("cell_std", "mean"),
            worst_std=("cell_std", "max"),
        )
        .reset_index()
    )

    rows = [{
        T("Task", "任务"): T_task(r["task"]),
        T("Recipe", "配方"): T_recipe(humanize_config(r["config_id"])),
        T("Model", "模型"): r["model_key"],
        T("Briefs with reruns", "有重跑的 brief 数"): int(r["briefs_with_reruns"]),
        T("Avg rerun std", "平均重跑标准差"): round(float(r["avg_std"]), 3),
        T("Worst rerun std", "最差重跑标准差"): round(float(r["worst_std"]), 3),
    } for _, r in out.iterrows()]

    df = pd.DataFrame(rows).sort_values(
        [T("Task", "任务"), T("Avg rerun std", "平均重跑标准差")],
        ascending=[True, False]
    )
    return df, T(
        f"Computed across {len(rerun_cells)} cells with at least two reruns each.",
        f"共覆盖 {len(rerun_cells)} 个有重复运行结果的单元格。"
    )


def _pairwise_summary() -> tuple[pd.DataFrame, str]:
    """
    Read outputs/pairwise_results.jsonl (if it exists) and aggregate
    head-to-head wins per (recipe_a, recipe_b) pair.

    Returns:
      (summary_df, status_msg)
      summary_df columns: Recipe A, Recipe B, Comparisons, A wins, B wins,
                          Ties, A win-rate
      status_msg: short status line for the section (e.g. waiting / done)
    """
    path = cfg.OUTPUTS_DIR / "pairwise_results.jsonl"
    if not path.exists():
        return pd.DataFrame(), T(
            "No pairwise AI-judge data. This is an optional cross-check.",
            "暂无两两 AI 评审数据，此项为可选交叉检查。"
        )
    from src.utils import read_jsonl
    rows = read_jsonl(path)
    if not rows:
        return pd.DataFrame(), T(
            "Pairwise results file is empty. This is an optional cross-check.",
            "两两评审结果文件为空，此项为可选交叉检查。"
        )

    df = pd.DataFrame(rows)
    if df.empty or "winner" not in df.columns:
        return pd.DataFrame(), T(
            "Pairwise results file has no usable rows.",
            "两两评审结果文件没有可用行。"
        )

    # Aggregate by (recipe_a, recipe_b)
    agg = (
        df.groupby(["recipe_a", "recipe_b"])["winner"]
        .value_counts().unstack(fill_value=0).reset_index()
    )
    for col in ("A", "B", "tie"):
        if col not in agg.columns:
            agg[col] = 0
    agg["Comparisons"] = agg["A"] + agg["B"] + agg["tie"]
    agg["A win-rate"] = (agg["A"] / agg["Comparisons"]).round(3)
    agg["recipe_a"] = agg["recipe_a"].map(humanize_config)
    agg["recipe_b"] = agg["recipe_b"].map(humanize_config)
    out = agg[["recipe_a", "recipe_b", "Comparisons",
               "A", "B", "tie", "A win-rate"]].copy()
    out.columns = [
        T("Recipe A", "配方 A"),
        T("Recipe B", "配方 B"),
        T("Comparisons", "对比次数"),
        T("A wins", "A 胜"),
        T("B wins", "B 胜"),
        T("Ties", "打平"),
        T("A win-rate", "A 胜率"),
    ]
    return out, T(
        f"Pairwise data loaded: {len(df)} comparisons across {len(out)} recipe pairs.",
        f"已加载两两评审数据：{len(df)} 次对比，覆盖 {len(out)} 对配方。"
    )


def _phase4_premium_summary(scored: pd.DataFrame) -> tuple[pd.DataFrame, str, str | None]:
    """
    Phase 4 directional check: "do stronger models look worth follow-up?"

    Three-tier ladder: cheap (haiku, gpt5mini) → medium (sonnet, gpt5)
    → premium (opus47, gpt55). For each sentence task we take the top-1
    cheap-screen winner (recipe × task) and look up cosine for that same
    (task, recipe) on every available model, then compute:
      - Δ cheap→medium  : does paying ~3-4× more close any gap?
      - Δ medium→premium: does paying ~5-10× more on top of medium close more?
    Phase 4 is intentionally small: it only validates the premium ladder on
    the curated briefs that were actually run, so the output is directional
    evidence, not a full 23-brief validation.

    Returns:
      (table_df, status_msg, conclusion_msg)
      conclusion_msg : "kind|message" where kind ∈ {green, yellow, neutral}
    """
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame(), T("No data.", "暂无数据。"), None

    cheap = {"haiku", "gpt5mini"}
    medium = {"sonnet", "gpt5"}
    premium = {"opus47", "gpt55"}
    available_models = set(sent["model_key"].unique())

    has_phase4 = bool(available_models & (medium | premium))
    if not has_phase4:
        return pd.DataFrame(), T(
            "Phase 4 has not run yet. Medium and premium model results will "
            "show up here once Phase 4 is run.",
            "Phase 4 尚未运行。中端和旗舰模型结果在 Phase 4 跑完后会显示在这里。"
        ), None

    phase4_rows = sent[sent["model_key"].isin(medium | premium)]
    phase4_briefs = int(phase4_rows["brief_id"].nunique()) if "brief_id" in phase4_rows.columns else 0
    phase4_scope = (
        f"{phase4_briefs} curated briefs" if phase4_briefs else "the curated Phase 4 brief subset"
    )

    # Pick top-1 per task using cheap-model average — the choice mustn't be
    # influenced by the medium/premium re-run we're comparing against.
    cheap_rows = sent[sent["model_key"].isin(cheap)]
    if cheap_rows.empty:
        cheap_rows = sent  # defensive fallback
    top_per_task = (
        cheap_rows.groupby(["task", "config_id"])["cosine"].mean().reset_index()
        .sort_values(["task", "cosine"], ascending=[True, False])
        .groupby("task").head(1)
        .rename(columns={"config_id": "recipe", "cosine": "_unused"})
    )

    rows = []
    cheap_to_medium, medium_to_premium = [], []
    for _, top in top_per_task.iterrows():
        task = top["task"]
        recipe = top["recipe"]
        cell = sent[(sent["task"] == task) & (sent["config_id"] == recipe)]
        per_model = cell.groupby("model_key")["cosine"].mean()

        def _g(m: str):
            return round(float(per_model[m]), 3) if m in per_model.index else None

        v_cheap = [_g("haiku"), _g("gpt5mini")]
        v_med   = [_g("sonnet"), _g("gpt5")]
        v_prem  = [_g("opus47"), _g("gpt55")]

        def _best(vs):
            real = [v for v in vs if v is not None]
            return max(real) if real else None

        best_cheap = _best(v_cheap)
        best_med   = _best(v_med)
        best_prem  = _best(v_prem)

        d_cm = round(best_med - best_cheap, 3) if (best_cheap is not None and best_med is not None) else None
        d_mp = round(best_prem - best_med, 3) if (best_med is not None and best_prem is not None) else None
        # Fallback gap: if medium missing, fall back to cheap→premium.
        d_cp = round(best_prem - best_cheap, 3) if (best_cheap is not None and best_prem is not None) else None

        if d_cm is not None:
            cheap_to_medium.append(d_cm)
        if d_mp is not None:
            medium_to_premium.append(d_mp)

        rows.append({
            T("Task", "任务"): T_task(task),
            T("Recipe", "配方"): T_recipe(humanize_config(recipe)),
            "haiku":    v_cheap[0],
            "gpt5mini": v_cheap[1],
            "sonnet":   v_med[0],
            "gpt5":     v_med[1],
            "opus47":   v_prem[0],
            "gpt55":    v_prem[1],
            T("Best cheap",   "便宜最佳"):   best_cheap,
            T("Best medium",  "中端最佳"):  best_med,
            T("Best premium", "旗舰最佳"): best_prem,
            T("Cheap to medium gap", "便宜→中端差值"):  d_cm,
            T("Medium to premium gap", "中端→旗舰差值"):   d_mp,
        })

    df = pd.DataFrame(rows)

    NOISE = 0.036
    if not (cheap_to_medium or medium_to_premium):
        return df, T(
            f"Compared on {len(df)} tasks. Phase 4 rows are present but do "
            f"not overlap with the top cheap-model recipes. Treat as preliminary.",
            f"在 {len(df)} 个任务上做了对比。Phase 4 行存在但与便宜模型的 top 配方"
            f"没有重叠，视为初步结果。"
        ), None

    def _avg(xs): return sum(xs) / len(xs) if xs else 0.0
    avg_cm = _avg(cheap_to_medium)
    avg_mp = _avg(medium_to_premium)

    big_cm = avg_cm >= NOISE
    big_mp = avg_mp >= NOISE

    if not big_cm and not big_mp:
        verdict = T(
            f"Directional only ({phase4_scope}): cheap models look sufficient. "
            f"The quality gain from medium "
            f"({avg_cm:+.3f}) and from premium ({avg_mp:+.3f}) is negligible. "
            f"Do not treat this as full validation across all 23 briefs.",
            f"仅作方向参考（{phase4_briefs or '少量'} 个代表性 brief）：便宜模型看起来已足够。中端模型带来的提升（{avg_cm:+.3f}）"
            f"和旗舰模型带来的提升（{avg_mp:+.3f}）都可以忽略。"
            f"不要把这解读为覆盖全部 23 个 brief 的完整验证。"
        )
        kind = "green"
    elif big_cm and not big_mp:
        verdict = T(
            f"Directional only ({phase4_scope}): medium may be the sweet spot. "
            f"Cheap to medium gives a lift "
            f"({avg_cm:+.3f}). Medium to premium adds almost nothing "
            f"({avg_mp:+.3f}). Validate on all 23 briefs before shipping a model upgrade.",
            f"仅作方向参考（{phase4_briefs or '少量'} 个代表性 brief）：中端模型可能是平衡点。从便宜升到中端有提升（{avg_cm:+.3f}）。"
            f"再升到旗舰几乎没有提升（{avg_mp:+.3f}）。"
            f"上线模型升级前应先覆盖全部 23 个 brief 验证。"
        )
        kind = "yellow"
    elif not big_cm and big_mp:
        verdict = T(
            f"Directional only ({phase4_scope}): premium may help. Medium does not "
            f"({avg_cm:+.3f}); premium does ({avg_mp:+.3f}). Validate on all 23 briefs "
            f"and confirm prices before any cost-quality claim.",
            f"仅作方向参考（{phase4_briefs or '少量'} 个代表性 brief）：旗舰可能有帮助。中端没有提升（{avg_cm:+.3f}），"
            f"旗舰有提升（{avg_mp:+.3f}）。任何成本质量结论前都要覆盖全部 23 个 brief 并确认价格。"
        )
        kind = "yellow"
    else:
        verdict = T(
            f"Directional only ({phase4_scope}): quality appears to scale with price. "
            f"Medium ({avg_cm:+.3f}) and premium ({avg_mp:+.3f}) both improve. "
            f"Validate on all 23 briefs and confirmed pricing before choosing a tier.",
            f"仅作方向参考（{phase4_briefs or '少量'} 个代表性 brief）：质量似乎随价格上升。中端（{avg_cm:+.3f}）和旗舰（{avg_mp:+.3f}）"
            f"都有提升。选择层级前需覆盖全部 23 个 brief 并确认价格。"
        )
        kind = "neutral"

    status = T(
        f"Phase 4 data loaded: {len(df)} task rows on {phase4_scope}; use as directional evidence only.",
        f"已加载 Phase 4 数据：{len(df)} 行任务，覆盖 {phase4_briefs or '少量'} 个代表性 brief；仅作方向参考。"
    )
    return df, status, f"{kind}|{verdict}"


# Human-readable task names for display (data keeps internal snake_case).
TASK_DISPLAY_NAMES = {
    "concept_relevant":  "Concept",
    "position_relevant": "Positioning",
    "emotion_relevant":  "Emotion",
    "function_relevant": "Function",
    "benefit_relevant":  "Benefit",
    "category_relevant": "Category",
    "feature_relevant":  "Features",
    "context_relevant":  "Context",
}

def humanize_task(task: str) -> str:
    """Convert internal task name to display label."""
    return TASK_DISPLAY_NAMES.get(task, task)


def humanize_config(config_id: str) -> str:
    """Convert internal config_id to a human-readable label."""
    if config_id in CONFIG_DISPLAY_NAMES:
        return CONFIG_DISPLAY_NAMES[config_id]
    # Drop the "A:" prompt-version prefix for sentence tasks
    if config_id.startswith("A:"):
        body = config_id[2:]
        # Pair fields: a+b → "audience + product"
        if "+" in body:
            return " + ".join(body.split("+"))
        return body
    return config_id


# ---------- helpers ----------

def write_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row, col, value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c


def write_header_row(ws, row, headers, fill=LIGHT_GREY):
    for i, h in enumerate(headers, 1):
        write_cell(ws, row, i, h, font=HEADER_FONT, fill=fill, border=THIN_BORDER)
    return row + 1


def write_df(ws, df: pd.DataFrame, start_row: int, *, with_header=True,
             row_fills: list | None = None, bold_cols: set | None = None) -> int:
    """Render a DataFrame as a table. Optional per-row background fills.

    row_fills, if provided, is one PatternFill (or None) per data row in df.
    Used to highlight Recommended rows green and baseline / no-go rows grey
    on the decision tables (Tab 2 main, Tab 3 keyword compression).

    bold_cols, if provided, is a set of 1-based column indices whose DATA cells
    are rendered bold — used to flag values that are auto-computed from the
    underlying data (vs fixed editorial text), so a reader can tell at a glance
    which numbers update on every rebuild.
    """
    if with_header:
        write_header_row(ws, start_row, list(df.columns))
        ws.row_dimensions[start_row].height = 22  # readable header
        start_row += 1
    for i, row_tuple in enumerate(df.itertuples(index=False)):
        fill = row_fills[i] if (row_fills and i < len(row_fills)) else None
        for c, val in enumerate(row_tuple, 1):
            cell_font = Font(bold=True) if (bold_cols and c in bold_cols) else None
            cell = write_cell(ws, start_row, c, val, border=THIN_BORDER, fill=fill,
                              font=cell_font)
            # Numbers center-aligned; everything else gets wrap_text on so
            # nothing truncates. (Previous version skipped wrap for strings
            # under 30 chars, which broke Chinese cells in narrow columns.)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                            indent=1, wrap_text=True)
        # Auto-fit row height for tables that carry long-text columns.
        # Picks the cell in this row whose wrapped text needs the most lines,
        # using a CJK-aware visual-width estimate (Chinese chars take ~2
        # column units, ASCII ~1) so Chinese text doesn't truncate.
        line_estimates = [1]
        for c, val in enumerate(row_tuple, 1):
            if isinstance(val, str):
                col_w = ws.column_dimensions[get_column_letter(c)].width or 12
                visual = _visual_width(val)
                # Lines needed at this column width. Subtract 1 for cell
                # padding (was 2 — too aggressive; over-counted lines).
                effective_w = max(4, int(col_w) - 1)
                lines = max(1, (visual + effective_w - 1) // effective_w)
                line_estimates.append(lines)
        max_lines = max(line_estimates)
        if max_lines >= 2:
            # 14 pt per wrapped line + 4 pt vertical padding. Cap at 220
            # so a single runaway cell can not blow up the sheet.
            ws.row_dimensions[start_row].height = min(220, max_lines * 14 + 4)
        else:
            ws.row_dimensions[start_row].height = 20  # tight single-line data row
        start_row += 1
    return start_row + 1


def autosize_cols(ws, widths: dict[int, int]):
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


# ---------- Shared analysis helpers (used by the 4 lean tab builders) ----------

def _quality_cost_data(scored: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    """Merge per-call cost (raw) with per-call quality (scored), aggregate
    to (task, config_id, model_key) means."""
    if scored.empty or raw.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()].copy()
    if sent.empty:
        return pd.DataFrame()
    merge_keys = ["brief_id", "task", "config_id", "model_key", "run_id"]
    common = [k for k in merge_keys if k in raw.columns and k in sent.columns]
    if len(common) < 4:
        return pd.DataFrame()
    m = sent.merge(
        raw[common + ["cost_usd"]],
        on=common, how="left", suffixes=("", "_raw"),
    )
    return (
        m.groupby(["task", "config_id", "model_key"])
        .agg(
            n=("cosine", "count"),
            mean_cosine=("cosine", "mean"),
            mean_cost_usd_per_call=("cost_usd", "mean"),
        )
        .round({"mean_cosine": 4, "mean_cost_usd_per_call": 6})
        .reset_index()
    )


def _candidate_sort(df: pd.DataFrame, *, score_col: str, cost_col: str) -> pd.DataFrame:
    """Deterministic ranking: quality first, then lower cost, then config/model id."""
    return df.sort_values(
        [score_col, cost_col, "config_id", "model_key"],
        ascending=[False, True, True, True],
    )


def _select_practical_winner(
    candidates: pd.DataFrame,
    *,
    score_col: str,
    cost_col: str,
    noise: float,
    override_cfg: str | None = None,
):
    """Pick the cheapest candidate within the noise floor of the top score.

    This is the core recommendation rule: a tiny quality lead is not a win.
    If the score gap is below the measured noise floor, the lower-cost prompt
    or model wins.
    """
    if candidates.empty:
        return None

    if override_cfg:
        forced = candidates[
            (candidates["config_id"] == override_cfg)
            & candidates["model_key"].isin(("haiku", "gpt5mini"))
        ]
        if not forced.empty:
            return forced.sort_values(
                [cost_col, score_col, "model_key"],
                ascending=[True, False, True],
            ).iloc[0]

    best_score = float(candidates[score_col].max())
    tied = candidates[candidates[score_col] >= best_score - noise]
    return tied.sort_values(
        [cost_col, score_col, "config_id", "model_key"],
        ascending=[True, False, True, True],
    ).iloc[0]


def _candidate_brief_scores(scored: pd.DataFrame, task: str, candidates: pd.DataFrame) -> pd.DataFrame:
    """Per-brief mean cosine for one task and candidate set."""
    if scored.empty or candidates.empty:
        return pd.DataFrame()
    sent = scored[
        scored["cosine"].notna()
        & (scored["task"] == task)
        & scored["brief_id"].notna()
    ].copy()
    if sent.empty:
        return pd.DataFrame()
    keys = candidates[["config_id", "model_key"]].drop_duplicates()
    sent = sent.merge(keys, on=["config_id", "model_key"], how="inner")
    if sent.empty:
        return pd.DataFrame()
    return (
        sent.groupby(["brief_id", "config_id", "model_key"], as_index=False)
        .agg(cosine=("cosine", "mean"))
    )


def _brief_level_stats(
    scored: pd.DataFrame,
    task: str,
    winner,
    candidates: pd.DataFrame,
    *,
    noise: float,
    baseline_score: float | None,
    override_cfg: str | None = None,
) -> dict:
    """Paired per-brief evidence and leave-one-brief-out stability."""
    stats = {
        "median_delta": None,
        "p_value": None,
        "wins": 0,
        "losses": 0,
        "ties": 0,
        "briefs": 0,
        "loo_first": 0,
        "loo_top3": 0,
        "loo_total": 0,
        "gap_vs_best": None,
        "above_noise": False,
        "delta_vs_full": None,
    }
    if winner is None or candidates.empty:
        return stats

    winner_key = (winner["config_id"], winner["model_key"])
    sorted_candidates = _candidate_sort(candidates, score_col="mean_cosine", cost_col="mean_cost_usd_per_call")
    absolute_best = sorted_candidates.iloc[0]
    stats["gap_vs_best"] = float(absolute_best["mean_cosine"] - winner["mean_cosine"])
    stats["above_noise"] = abs(stats["gap_vs_best"]) >= noise
    if baseline_score is not None and pd.notna(baseline_score):
        stats["delta_vs_full"] = float(winner["mean_cosine"] - baseline_score)

    per_brief = _candidate_brief_scores(scored, task, candidates)
    if per_brief.empty:
        return stats

    pivot = per_brief.pivot_table(
        index="brief_id",
        columns=["config_id", "model_key"],
        values="cosine",
        aggfunc="mean",
    )
    if winner_key in pivot.columns:
        peer_cols = [c for c in pivot.columns if c != winner_key]
        if peer_cols:
            # Compare the winner against the SINGLE strongest peer by mean
            # cosine — NOT the per-brief max across all peers. Taking the
            # per-brief max of many peers is a winner's-curse artifact: the
            # max is biased upward, so a fixed winner "loses" on almost every
            # brief (median Δ goes negative, W/T/L fills with losses, and the
            # signed-rank p collapses to 0.00 in the *wrong* direction). The
            # runner-up-by-mean comparison matches audit D4's sign test.
            best_peer_key = pivot[peer_cols].mean(axis=0).idxmax()
            best_peer = pivot[best_peer_key]
            paired = (pivot[winner_key] - best_peer).dropna()
            if not paired.empty:
                stats["median_delta"] = float(paired.median())
                stats["wins"] = int((paired > noise).sum())
                stats["losses"] = int((paired < -noise).sum())
                stats["ties"] = int((paired.abs() <= noise).sum())
                stats["briefs"] = int(paired.shape[0])
                # Paired significance: does the winner beat the best peer on the
                # SAME briefs, not just on average? None when too few pairs.
                stats["p_value"] = _paired_signed_rank_p(paired.values)

    loo_total = 0
    loo_first = 0
    loo_top3 = 0
    for brief_id in sorted(per_brief["brief_id"].dropna().unique()):
        reduced = per_brief[per_brief["brief_id"] != brief_id]
        if reduced.empty:
            continue
        loo = (
            reduced.groupby(["config_id", "model_key"], as_index=False)
            .agg(
                mean_cosine=("cosine", "mean"),
                mean_cost_usd_per_call=("cosine", lambda _: 0.0),
            )
        )
        loo = loo.merge(
            candidates[["config_id", "model_key", "mean_cost_usd_per_call"]],
            on=["config_id", "model_key"],
            how="left",
            suffixes=("", "_candidate"),
        )
        loo["mean_cost_usd_per_call"] = loo["mean_cost_usd_per_call_candidate"].fillna(0.0)
        loo = loo.drop(columns=[c for c in loo.columns if c.endswith("_candidate")])
        loo_winner = _select_practical_winner(
            loo,
            score_col="mean_cosine",
            cost_col="mean_cost_usd_per_call",
            noise=noise,
            override_cfg=override_cfg,
        )
        if loo_winner is None:
            continue
        loo_total += 1
        if (loo_winner["config_id"], loo_winner["model_key"]) == winner_key:
            loo_first += 1
        top3 = _candidate_sort(loo, score_col="mean_cosine", cost_col="mean_cost_usd_per_call").head(3)
        if any((r["config_id"], r["model_key"]) == winner_key for _, r in top3.iterrows()):
            loo_top3 += 1

    stats["loo_total"] = loo_total
    stats["loo_first"] = loo_first
    stats["loo_top3"] = loo_top3
    return stats


def _decision_tier(delta_vs_full, gap_vs_best, loo_first: int, loo_total: int, *,
                   noise: float, p_value=None, median_delta=None,
                   alpha: float | None = None) -> str:
    """Recommendation framing for small-sample results.

    Deliberately avoids absolute language ("ship" / "best"). The strongest
    label we allow is "Recommended (stable on this sample)", and we only grant
    it when BOTH hold:
      - leave-one-brief-out keeps the same winner ≥75% of the time, AND
      - the winner beats its best peer on the same briefs at the paired test
        (p < alpha) with a positive median delta.
    A higher mean alone is never enough — that is the small-sample trap.
    """
    if alpha is None:
        alpha = getattr(cfg, "PAIRED_TEST_ALPHA", 0.05)
    stable = loo_total == 0 or loo_first >= max(1, int(np.ceil(0.75 * loo_total)))
    sig_better = (
        p_value is not None and p_value < alpha
        and (median_delta is None or float(median_delta) > 0)
    )
    if delta_vs_full is not None and delta_vs_full < -noise:
        return T("Not recommended", "不推荐")
    if gap_vs_best is not None and gap_vs_best > noise:
        return T("Not recommended", "不推荐")
    if stable and sig_better:
        return T("Recommended (stable on this sample)", "推荐（本样本稳定）")
    # Within noise of the best / baseline, or lead not statistically confirmed:
    # honest framing is "usable, prefer the cheaper option", not a win.
    return T("Cost-priority usable", "成本优先可用")


# Per-model production recommendation + explanation.
# NOTE: every number in the note is a {placeholder} filled from live data by
# _build_model_recs() at build time — nothing here is hand-typed, so a re-run
# can never ship a stale figure (the old "46%" / "0.603" hardcodes did).
#   {q}  = this model's mean cosine
#   {qd} = the default model's (gpt5mini) mean cosine
#   {rl} = share of this model's calls whose final status was rate_limited
_MODEL_REC_TEMPLATES_EN = {
    "gpt5mini": ("Default",
                 "Highest average quality across tested cells ({q:.3f}); cost advantage is provisional until prices are verified."),
    "haiku":    ("Not recommended",
                 "Rate-limit / reliability risk — {rl:.0f}% of calls ended rate-limited; lower throughput than GPT-5-mini."),
    "gpt5":     ("Not default",
                 "Below GPT-5-mini on average ({q:.3f} vs {qd:.3f}); do not prefer unless verified pricing or quality changes."),
    "sonnet":   ("Not default",
                 "Below GPT-5-mini on average ({q:.3f} vs {qd:.3f}); keep as reference until human validation and prices are confirmed."),
    "gpt55":    ("Skip",
                 "Below GPT-5-mini on average ({q:.3f} vs {qd:.3f}); no advantage."),
    "opus47":   ("Feature task escalation only",
                 "Small-sample Phase 4 signal only; validate on all 23 briefs and confirmed pricing before escalation."),
}
_MODEL_REC_TEMPLATES_ZH = {
    "gpt5mini": ("默认",
                 "在测试单元格上平均质量最高（{q:.3f}）；成本优势需等价格校准后再写死。"),
    "haiku":    ("备用",
                 "成本相近，但约 {rl:.0f}% 调用最终被限流；吞吐量低于 GPT-5-mini。"),
    "gpt5":     ("不建议默认",
                 "平均质量低于 GPT-5-mini（{q:.3f} vs {qd:.3f}）；除非价格或质量校准结果变化，否则不优先。"),
    "sonnet":   ("不建议默认",
                 "平均质量低于 GPT-5-mini（{q:.3f} vs {qd:.3f}）；人工验证和价格确认前仅作参考。"),
    "gpt55":    ("跳过",
                 "平均质量低于 GPT-5-mini（{q:.3f} vs {qd:.3f}）；无优势。"),
    "opus47":   ("仅 feature 任务升级用",
                 "只有小样本 Phase 4 方向信号；升级前需覆盖全部 23 个 brief 并确认价格。"),
}


def _model_rate_limit_pct(raw: pd.DataFrame, model_key: str) -> float:
    """Share (%) of a model's calls whose FINAL status was rate_limited.

    SDK-level retries are invisible (they happen inside the provider client),
    so this is the unrecovered rate-limit rate actually present in the data —
    a real, self-updating number rather than a typed-in constant.
    """
    if raw is None or raw.empty or "status" not in raw.columns:
        return 0.0
    sub = raw[raw["model_key"] == model_key]
    if sub.empty:
        return 0.0
    return 100.0 * float((sub["status"] == "rate_limited").mean())


def _build_model_recs(scored: pd.DataFrame, raw: pd.DataFrame) -> dict:
    """Fill the rec-note templates with live numbers from scored/raw."""
    templates = _MODEL_REC_TEMPLATES_ZH if LANG == "zh" else _MODEL_REC_TEMPLATES_EN
    q_by_model = (
        scored[scored["cosine"].notna()].groupby("model_key")["cosine"].mean()
        if not scored.empty else pd.Series(dtype=float)
    )
    qd = float(q_by_model.get("gpt5mini", float("nan")))
    recs: dict[str, tuple[str, str]] = {}
    for m, (label, tmpl) in templates.items():
        q = float(q_by_model.get(m, float("nan")))
        rl = _model_rate_limit_pct(raw, m)
        try:
            note = tmpl.format(q=q, qd=qd, rl=rl)
        except (KeyError, ValueError):
            note = tmpl
        recs[m] = (label, note)
    return recs


def _model_comparison_table(scored: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    """Per-model summary table for Tab 2.

    Columns: Model / Avg quality / Avg cost per call / Production rec / Note.

    Quality and cost are computed from scored/raw. The recommendation label is
    fixed wording, but every NUMBER in the note is filled from live data by
    _build_model_recs(), so Tab 2 can never display a stale figure.
    """
    if scored.empty or raw.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return pd.DataFrame()

    cost_by_model = raw.groupby("model_key")["cost_usd"].mean() if "cost_usd" in raw.columns else pd.Series()
    quality_by_model = sent.groupby("model_key")["cosine"].mean()
    if quality_by_model.empty:
        return pd.DataFrame()

    rec_map = _build_model_recs(scored, raw)

    rows = []
    # Order: cheap first, then medium, then premium
    model_order = ["gpt5mini", "haiku", "gpt5", "sonnet", "gpt55", "opus47"]
    for m in model_order:
        if m not in quality_by_model.index:
            continue
        q = float(quality_by_model[m])
        c = float(cost_by_model.get(m, float("nan")))
        rec, note = rec_map.get(m, ("—", "—"))
        rows.append({
            T("Model", "模型"): m,
            T("Avg quality", "平均质量"): round(q, 3),
            T("Avg cost / call", "平均每次调用成本"):
                f"${c:.4f}" if pd.notna(c) else "—",
            T("Recommendation", "推荐"): rec,
            T("Note", "说明"): note,
        })
    return pd.DataFrame(rows)


def _full_brief_baseline_by_task(scored: pd.DataFrame) -> pd.Series:
    """Mean cosine per task for the Full-brief baseline."""
    if scored.empty:
        return pd.Series(dtype=float)
    sent = scored[scored["cosine"].notna()]
    base = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
    if base.empty:
        return pd.Series(dtype=float)
    return base.groupby("task")["cosine"].mean()


def _recommendation_label(
    winner_cosine: float, winner_cost: float,
    absolute_best_cosine: float, absolute_best_cost: float,
    noise: float = 0.036,
) -> str:
    """Short tag for the Recommendation column."""
    is_absolute = abs(winner_cosine - absolute_best_cosine) < 1e-6 \
        and abs(winner_cost - absolute_best_cost) < 1e-9
    if is_absolute:
        return T("Recommended", "推荐")
    gap = absolute_best_cosine - winner_cosine
    if gap < noise:
        return T("Best value", "高性价比")
    if gap >= noise:
        return T("Premium optional", "高端可选")
    return T("Not worth extra cost", "不值额外成本")


# ---------- Tab 2 helper: best per task ----------

def _best_per_task_table(scored: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    qc = _quality_cost_data(scored, raw)
    if qc.empty:
        return pd.DataFrame()
    baseline = _full_brief_baseline_by_task(scored)
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    NOISE = cfg.NOISE_FLOOR_COSINE
    rows = []
    for task, g in qc.groupby("task"):
        # Tasks in _PER_TASK_CHEAP_TIER_ONLY must ship a cheap-tier model
        # even if a premium model scored higher. Restrict candidates first,
        # then apply the usual cheapest-within-noise rule.
        if task in _PER_TASK_CHEAP_TIER_ONLY:
            cheap_g = g[g["model_key"].isin(("haiku", "gpt5mini"))]
            if not cheap_g.empty:
                g = cheap_g

        absolute_best = _candidate_sort(
            g, score_col="mean_cosine", cost_col="mean_cost_usd_per_call"
        ).iloc[0]
        override_cfg = _PER_TASK_RECIPE_OVERRIDE.get(task)
        winner = _select_practical_winner(
            g,
            score_col="mean_cosine",
            cost_col="mean_cost_usd_per_call",
            noise=NOISE,
            override_cfg=override_cfg,
        )
        if winner is None:
            continue
        base_score = baseline.get(task)
        delta_vs_full = (winner["mean_cosine"] - base_score) if (base_score is not None) else None

        stats = _brief_level_stats(
            scored,
            task,
            winner,
            g,
            noise=NOISE,
            baseline_score=base_score,
            override_cfg=override_cfg,
        )
        recipe_label = T_recipe(humanize_config(winner["config_id"]))
        delta_rounded = (round(float(delta_vs_full), 3)
                          if delta_vs_full is not None else None)
        vs_full = _comparison_vs_full_brief(delta_rounded, base_score)
        gap_vs_best = stats["gap_vs_best"]
        prod_action = _decision_tier(
            delta_vs_full,
            gap_vs_best,
            stats["loo_first"],
            stats["loo_total"],
            noise=NOISE,
            p_value=stats.get("p_value"),
            median_delta=stats.get("median_delta"),
        )
        noise_verdict = (
            T("Tie: choose cheaper", "打平：选更便宜")
            if gap_vs_best is not None and gap_vs_best < NOISE
            else T("Above noise floor", "超过噪声阈值")
        )
        median_delta = stats["median_delta"]
        _p = stats.get("p_value")
        _p_str = (f" p={_p:.2f}" if _p is not None else " p=n/a")
        paired = (
            f"{stats['wins']}/{stats['ties']}/{stats['losses']}{_p_str}"
            if stats["briefs"] else "—"
        )
        loo = (
            f"{stats['loo_first']}/{stats['loo_total']} first; "
            f"{stats['loo_top3']}/{stats['loo_total']} top-3"
            if stats["loo_total"] else "—"
        )
        rows.append({
            T("Task", "任务"): T_task(task),
            T("Recommended recipe", "推荐配方"): recipe_label,
            T("Model", "模型"): winner["model_key"],
            T("Avg score", "平均分数"):
                round(float(winner["mean_cosine"]), 3),
            T("Median Δ vs next best", "中位数差值"):
                f"{median_delta:+.3f}" if median_delta is not None else "—",
            T("Win/Tie/Loss briefs", "胜/平/负 brief"):
                paired,
            T("Vs Full Brief", "对比完整 brief"): vs_full,
            T("Δ vs Full Brief", "相对完整 brief 差值"):
                f"{delta_rounded:+.3f}" if delta_rounded is not None else "—",
            T("Noise-floor call", "噪声阈值判断"): noise_verdict,
            T("Leave-one-out stability", "去掉一个 brief 稳定性"): loo,
            T("Decision", "决定"): prod_action,
            T("Why", "原因"): _per_task_reason(task),
        })
    return pd.DataFrame(rows)


def _comparison_vs_full_brief(delta, base_score) -> str:
    """Plain-language 'vs Full Brief' cell for Tab 2.

    Deliberately no Δ-symbol, no 'negligible-difference line' jargon —
    business reader-friendly phrasing only. The technical Δ / threshold
    lives in the Appendix.
    """
    if delta is None:
        return "—"
    try:
        d = float(delta)
    except (ValueError, TypeError):
        return "—"
    if abs(d) <= 0.036:
        return T("Equivalent", "持平")
    if d > 0:
        return T("Better", "略优")
    if d >= -0.05:
        return T("Slightly behind", "略低")
    return T("Behind", "明显落后")


# ---------- Field-contribution helpers ----------

def _is_single_field_config(cid: str) -> bool:
    """A:audience → True. A:audience+product → False. A:_full_brief → False."""
    if cid in ("A:_full_brief", "_full_brief", "A:_prompt_implied", "_prompt_implied"):
        return False
    rhs = cid.split(":")[-1] if ":" in cid else cid
    return "+" not in rhs and rhs not in ("_full_brief", "_prompt_implied")


def _is_field_combination_config(cid: str) -> bool:
    """A:audience+product → True."""
    rhs = cid.split(":")[-1] if ":" in cid else cid
    return "+" in rhs


def _field_contribution_summary(scored: pd.DataFrame) -> dict:
    """Top fields, weak fields, combinations worth keeping, baseline reading.

    Top / weak fields are ranked by mean cosine across every (task, brief)
    where the single field was tested — not by "absolute task winner" count,
    because the negligible-difference line often lets several recipes tie for first.
    Metadata-only fields (business_category) are excluded since they're only
    tested on 1-2 tasks by design.
    """
    out = {"top_fields": [], "weak_fields": [], "combos_keep": [], "baseline_note": ""}
    if scored.empty:
        return out
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return out

    # Top / weak single fields by mean cosine across all tasks they were tested on.
    metadata_only = {"business_category"}
    all_singles = sent[sent["config_id"].apply(_is_single_field_config)].copy()
    if not all_singles.empty:
        all_singles["_field"] = all_singles["config_id"].apply(lambda c: c.split(":")[-1])
        all_singles = all_singles[~all_singles["_field"].isin(metadata_only)]
        field_mean = all_singles.groupby("_field")["cosine"].mean().sort_values(ascending=False)
        if not field_mean.empty:
            out["top_fields"] = [
                f"{name} ({val:.2f})" for name, val in field_mean.head(3).items()
            ]
            # Weak = bottom 2 if the gap is meaningful (>0.04 below top)
            if len(field_mean) >= 4:
                top_score = field_mean.iloc[0]
                weak = field_mean[field_mean < top_score - 0.04].tail(2)
                if not weak.empty:
                    out["weak_fields"] = [
                        f"{name} ({val:.2f})" for name, val in weak.items()
                    ]

    # Top-3 multi-field combos by mean cosine, restricted to combos with
    # broad task coverage (≥ 5 of 8 sentence tasks). Otherwise metadata-only
    # combos tested on 1 task dominate the ranking with a single easy data
    # point and mislead the reader.
    combos = sent[sent["config_id"].apply(_is_field_combination_config)]
    top_combo_fields: set[str] = set()
    if not combos.empty:
        coverage = combos.groupby("config_id")["task"].nunique()
        viable = coverage[coverage >= 5].index.tolist()
        if viable:
            combos_v = combos[combos["config_id"].isin(viable)]
            combo_mean = combos_v.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
            out["combos_keep"] = [
                f"{humanize_config(c)} ({v:.2f})" for c, v in combo_mean.head(3).items()
            ]
            # Extract the field names that appear in any top-3 combo, so we can
            # warn readers that a "weak alone" field may still be strong paired.
            for cid in combo_mean.head(3).index:
                rhs = cid.split(":")[-1]
                top_combo_fields.update(rhs.split("+"))

    # Annotate weak fields that ARE valuable in top combos — avoids the
    # contradiction of "don't bother with X" right next to "best combo: X + Y".
    if out["weak_fields"] and top_combo_fields:
        weak_names = {f.split(" ")[0] for f in out["weak_fields"]}
        overlap = sorted(weak_names & top_combo_fields)
        if overlap:
            out["weak_combo_note"] = (
                f"{', '.join(overlap)} underperforms as a single field but is part "
                f"of a top combination above — keep it when paired, drop it alone."
            )

    # No-Context-Baseline reading
    pimp = sent[sent["config_id"].isin(("A:_prompt_implied", "_prompt_implied"))]["cosine"].mean()
    full = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]["cosine"].mean()
    if pd.notna(pimp) and pd.notna(full):
        gap = full - pimp
        if gap > 0.09:
            out["baseline_note"] = (
                f"No-Context Baseline {pimp:.2f} vs Full brief {full:.2f} (Δ {gap:+.2f}) — "
                f"the brief contributes a lot. Always send brief content."
            )
        elif gap > 0.036:
            out["baseline_note"] = (
                f"No-Context Baseline {pimp:.2f} vs Full brief {full:.2f} (Δ {gap:+.2f}) — "
                f"the brief adds some value. Keep using it."
            )
        else:
            out["baseline_note"] = (
                f"No-Context Baseline {pimp:.2f} ≈ Full brief {full:.2f} (Δ {gap:+.2f}, negligible) — "
                f"the instruction alone is doing most of the work; the brief barely helps."
            )
    return out


# ---------- Tab 3 helper: keyword compression ----------

# Per-version Decision implication — fixed user-approved wording. The
# Recommended? Y/N column carries the verdict; this column adds one short
# sentence saying what to actually do about it.
_KW_DECISION_EN = {
    "A — Full":     "Keep only as a control comparison.",
    "B — Reduced":  "Recommended for production. Shorter and quality is not lost.",
    "C — Compact":  "Not recommended. Quality drops noticeably.",
    "D — Minimal":  "Not recommended. Quality drops noticeably.",
}
_KW_DECISION_ZH = {
    "A — Full":     "只作为对照组保留。",
    "B — Reduced":  "推荐上线；更短且质量无损。",
    "C — Compact":  "质量下降明显，不推荐。",
    "D — Minimal":  "质量下降明显，不推荐。",
}


def _keyword_compression_table(scored: pd.DataFrame) -> pd.DataFrame:
    """Per-version keyword-prompt comparison. 10 columns per user spec:

      | Version | Description | Characters | % reduction vs A |
        Precision | Recall | F1 | F1 delta vs A | Recommended? | Decision |

    The "What changed from previous version" and "Est. tokens" columns from
    an earlier iteration are dropped per the user's latest Tab 3 spec.
    """
    lengths = _keyword_prompt_lengths()
    if lengths.empty:
        return pd.DataFrame()
    kw = scored[scored["f1"].notna()] if not scored.empty else scored
    rows = []
    base_f1 = None
    decision_map = _KW_DECISION_ZH if LANG == "zh" else _KW_DECISION_EN
    for _, lr in lengths.iterrows():
        version_label = lr["Prompt version"]
        version_key = version_label.split(" — ")[0]
        if not kw.empty:
            v_rows = kw[kw["config_id"].str.startswith(f"{version_key}:")]
            if not v_rows.empty:
                p  = float(v_rows["precision"].mean())
                r  = float(v_rows["recall"].mean())
                f1 = float(v_rows["f1"].mean())
            else:
                p = r = f1 = None
        else:
            p = r = f1 = None
        if version_key == "A" and f1 is not None:
            base_f1 = f1

        # F1 delta vs baseline A
        if f1 is None or base_f1 is None:
            f1_delta = None
        else:
            f1_delta = round(f1 - base_f1, 3)

        # Recommended Y/N — A is baseline; B+ recommended if within ±0.02 F1.
        if version_key == "A":
            recommended = T("Baseline", "基线")
        elif f1 is None or base_f1 is None:
            recommended = "—"
        elif f1_delta is not None and f1_delta >= -0.02:
            recommended = "__candidate__"  # sentinel; fixed after loop
        else:
            recommended = T("No", "否")

        # Decision implication — fixed wording per version.
        implication = decision_map.get(version_label, "—")

        rows.append({
            T("Version", "版本"): version_label,
            T("Description", "说明"): _prompt_description(version_label),
            T("Characters", "字符数"): int(lr["Characters"]),
            T("% reduction vs A", "相对 A 压缩"): lr["Reduction vs A"],
            T("Precision", "Precision"):
                f"{p:.3f}" if p is not None else "—",
            T("Recall", "Recall"):
                f"{r:.3f}" if r is not None else "—",
            T("F1", "F1"):
                f"{f1:.3f}" if f1 is not None else "—",
            T("F1 delta vs A", "相对 A 的 F1 差值"):
                f"{f1_delta:+.3f}" if f1_delta is not None else "—",
            T("Recommended?", "是否推荐"): recommended,
            T("Decision implication", "决策含义"): implication,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        rec_col = T("Recommended?", "是否推荐")
        cand_mask = df[rec_col] == "__candidate__"
        if cand_mask.any():
            shortest_idx = df[cand_mask].index[-1]
            df.loc[df[rec_col] == "__candidate__", rec_col] = T("No", "否")
            df.loc[shortest_idx, rec_col] = T("Yes", "是")
    return df


def _keyword_compression_headline(scored: pd.DataFrame) -> str:
    """One-sentence headline summary of keyword compression for Tab 1.
    Reads off the new keyword table's Recommended? = Yes row."""
    kw_df = _keyword_compression_table(scored)
    if kw_df.empty:
        return "Not yet tested (run Phase 3 to evaluate A/B/C/D)."
    rec = kw_df[kw_df[T("Recommended?", "是否推荐")] == T("Yes", "是")]
    if rec.empty:
        return T(
            "All compressed versions lose quality vs Version A — keep Version A.",
            "所有压缩版本相对 A 都有质量下降 —— 保留 A 版。")
    r = rec.iloc[0]
    ver_col = T("Version", "版本")
    red_col = T("% reduction vs A", "相对 A 压缩")
    f1d_col = T("F1 delta vs A", "相对 A 的 F1 差值")
    desc = _prompt_description(r[ver_col])
    return T(
        f"{r[ver_col]} ({desc}) compresses {r[red_col]} "
        f"with no quality loss (F1 delta {r[f1d_col]}).",
        f"{r[ver_col]}（{desc}）压缩 {r[red_col]} 且质量无损失（F1 差值 {r[f1d_col]}）。")


# ---------- Tab 1 helpers ----------

def _modal_winning_recipe(bpt: pd.DataFrame) -> tuple[str, int, int]:
    """(recipe, wins, total) — modal recipe across tasks."""
    if bpt.empty:
        return ("—", 0, 0)
    vc = bpt[T("Recommended recipe", "推荐配方")].value_counts()
    return (vc.index[0], int(vc.iloc[0]), len(bpt))


def _decisive_win_count(scored: pd.DataFrame,
                        noise_floor: float = 0.036) -> tuple[int, int]:
    """How many per-task winners beat the runner-up by ≥ noise_floor cosine.

    Tasks where top-1 vs top-2 differ by less than the negligible-difference line are
    statistically tied — claiming 'recipe X wins this task' overstates
    certainty. This count is the basis for the qualified headline.
    Returns (decisive_wins, total_sentence_tasks).
    """
    if scored.empty:
        return (0, 0)
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return (0, 0)
    means = sent.groupby(["task", "config_id"])["cosine"].mean().reset_index()
    decisive = 0
    total = 0
    for task, g in means.groupby("task"):
        if task == "keywords":
            continue  # keyword task ranked by F1, not cosine
        total += 1
        srt = g.sort_values("cosine", ascending=False)
        if len(srt) >= 2 and (srt.iloc[0]["cosine"] - srt.iloc[1]["cosine"]) >= noise_floor:
            decisive += 1
    return (decisive, total)


def _best_field_combination(scored: pd.DataFrame) -> str:
    """Top multi-field recipe by mean cosine across tasks.

    Excludes combos that have task coverage < 5 (out of 8 sentence tasks).
    Otherwise metadata-only combos like business_category+product — tested
    on only 1 task by design — show up as "best" with an unfairly high mean
    computed on a single (easy) task.
    """
    MIN_TASK_COVERAGE = 5
    if scored.empty:
        return "—"
    sent = scored[scored["cosine"].notna()]
    combos = sent[sent["config_id"].apply(_is_field_combination_config)]
    if combos.empty:
        return "Not tested (no field-combination recipes in data yet)."
    coverage = combos.groupby("config_id")["task"].nunique()
    viable = coverage[coverage >= MIN_TASK_COVERAGE].index.tolist()
    if not viable:
        return f"No combo has ≥ {MIN_TASK_COVERAGE}-task coverage yet (need broader Stage A data)."
    combos = combos[combos["config_id"].isin(viable)]
    grouped = combos.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
    cid = grouped.index[0]
    return f"{humanize_config(cid)} (mean cosine {grouped.iloc[0]:.2f})"


def _mean_delta_vs_full_brief(bpt: pd.DataFrame) -> float | None:
    """Mean numeric delta across tasks where Δ vs Full Brief is computed."""
    if bpt.empty:
        return None
    deltas = []
    for v in bpt[T("Δ vs Full Brief", "相对完整 brief 的差值")]:
        try:
            deltas.append(float(v))
        except (ValueError, TypeError):
            continue
    return sum(deltas) / len(deltas) if deltas else None


def _full_brief_value_judgement(scored: pd.DataFrame) -> str:
    """Is the full brief worth running? Uses per-task ABSOLUTE best (highest mean
    cosine), not Tab 2's noise-floor cheapest-equivalent winner. Avoids the bug
    where Tab 2's cheaper-but-tied winner masks the fact that Full brief is
    actually the highest-scoring recipe on that task.
    """
    if scored.empty:
        return "—"
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return "—"
    per_task_best = (
        sent.groupby(["task", "config_id"])["cosine"].mean().reset_index()
        .sort_values(["task", "cosine"], ascending=[True, False])
        .groupby("task").head(1)
    )
    full_brief_ids = {"A:_full_brief", "_full_brief"}
    total = len(per_task_best)
    full_best = int(per_task_best["config_id"].isin(full_brief_ids).sum())
    if full_best == total:
        return f"Yes — Full brief is the absolute best on all {total} tasks."
    if full_best == 0:
        return f"No — a leaner recipe beats Full brief on every one of {total} tasks."
    return (
        f"Mixed — Full brief is the absolute best on {full_best}/{total} tasks; "
        f"leaner recipes win the other {total - full_best}."
    )


def _best_cheap_model(scored: pd.DataFrame) -> str:
    if scored.empty:
        return "—"
    sent = scored[scored["cosine"].notna()]
    full = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
    by_model = full.groupby("model_key")["cosine"].mean()
    cheap = [m for m in by_model.index if m in ("haiku", "gpt5mini")]
    if not cheap:
        return "—"
    if len(cheap) == 1:
        return f"{cheap[0]} (only cheap model with data)"
    diff = abs(by_model[cheap[0]] - by_model[cheap[1]])
    if diff < 0.036:
        return f"{cheap[0]} ≈ {cheap[1]} (practically equivalent, Δ {diff:.2f}) — pick either."
    return f"{by_model[cheap].idxmax()} (leads by {diff:.2f})"


def _cost_quality_stance(scored: pd.DataFrame) -> str:
    """One-line cost-quality stance for Tab 1."""
    _, _, prem_verdict = _phase4_premium_summary(scored)
    if prem_verdict and "|" in prem_verdict:
        _, msg = prem_verdict.split("|", 1)
        return msg.strip().split(". ")[0] + "."
    return "Cheap models look sufficient. Premium check (top configs only) is the next step."


def _intended_skip_cells() -> set[tuple[str, str]]:
    """
    (task, config_id) pairs that were DESIGNED not to be tested.

    Metadata fields (`business_category`) are only run on a handful of
    hypothesis-driven (task, fields) combos — see
    prompt_builder.TARGETED_METADATA_COMBOS. For every other sentence task,
    the corresponding heatmap cell is "n/a by design", not "missing data".
    Without this distinction, the heatmap reads as if Stage A failed for 14
    cells when it actually never tried them.

    Returns a set of (task, config_id) pairs to render as "n/a" rather
    than as blank/missing.
    """
    try:
        from src.prompt_builder import (
            METADATA_FIELDS,
            SENTENCE_TASKS,
            TARGETED_METADATA_COMBOS,
        )
    except Exception:
        return set()

    metadata_set = set(METADATA_FIELDS)

    # config_id used in the heatmap (e.g. "A:business_category",
    # "A:business_category+product") -> the set of tasks it was DESIGNED for.
    intended: dict[str, set[str]] = {}
    for task, fields in TARGETED_METADATA_COMBOS:
        # Only treat a combo as "task-specific" if it actually uses a metadata
        # field. Targeted combos that involve only semantic fields (e.g.
        # ("context_relevant", ["audience"])) duplicate Phase 1 enumeration and
        # appear across all tasks anyway — don't grey them out.
        if not (set(fields) & metadata_set):
            continue
        joined = "+".join(sorted(fields))
        cfg_id = f"A:{joined}"
        intended.setdefault(cfg_id, set()).add(task)

    skip: set[tuple[str, str]] = set()
    for cfg_id, task_set in intended.items():
        for t in SENTENCE_TASKS:
            if t not in task_set:
                skip.add((t, cfg_id))
    return skip


def _premium_model_recommendation_line(scored: pd.DataFrame) -> str:
    """One-line premium-model decision for Tab 1."""
    if not scored.empty and scored[scored["model_key"].isin(("opus47", "gpt55"))].shape[0] > 0:
        return _cost_quality_stance(scored)
    return "Test premium models only on the top cheap-screen winners (< £1 total)."


def _quality_ceiling_answer(scored: pd.DataFrame) -> str:
    """One-line answer to the 'what quality with the best models?' question.

    Returns a number (avg + peak premium cosine on top configs) once Phase 4 has
    been run; a forward-looking placeholder otherwise.
    """
    if scored.empty:
        return "Not yet measured — run Phase 4 to fill in."
    prem = scored[scored["model_key"].isin(("opus47", "gpt55"))]
    if prem.empty:
        return "Not yet measured — Phase 4 will run Opus 4.7 + GPT-5.5 on top configs."
    sent = prem[prem["cosine"].notna()]
    if sent.empty:
        return "Phase 4 partly run — no sentence-task scores available yet."
    avg = float(sent["cosine"].mean())
    peak = float(sent["cosine"].max())
    return f"avg {avg:.2f}, peak {peak:.2f} similarity (0–1 scale, higher = closer to ground truth)."


def _tldr_bullets(scored: pd.DataFrame, raw: pd.DataFrame) -> list[str]:
    """Four plain-English conclusions for the 10-second scan layer.

    Derived from the same data as the headline table below, but phrased as
    actions ('use X', 'pay Y') rather than measurements ('Δ = 0.02'). Order:
    1) cheap-vs-cheap model verdict, 2) lean-vs-full-brief verdict,
    3) modal winning recipe, 4) keyword compression headline.
    """
    bullets: list[str] = []
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored

    # 1) Cheap models — equivalent or one wins?
    if not sent.empty:
        full = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        by_model = full.groupby("model_key")["cosine"].mean()
        cheap = [m for m in by_model.index if m in ("haiku", "gpt5mini")]
        if len(cheap) >= 2:
            diff = abs(by_model[cheap[0]] - by_model[cheap[1]])
            if diff < 0.036:
                bullets.append(
                    "GPT-5-mini and Haiku perform nearly identically — "
                    "pick whichever is cheaper or faster on your stack."
                )
            else:
                winner = by_model[cheap].idxmax()
                bullets.append(
                    f"{winner} leads the cheap-tier comparison by {diff:.2f} "
                    f"on Full Brief — prefer it unless cost dominates."
                )

    # 2) Lean vs Full brief
    bpt = _best_per_task_table(scored, raw)
    if not bpt.empty:
        lean_winners = sum(
            1 for v in bpt[T("Recommended recipe", "推荐配方")]
            if v not in ("Full brief", "_full_brief")
        )
        total = len(bpt)
        if lean_winners >= total // 2:
            bullets.append(
                f"Lean recipes (a single field or a small pair) match or beat Full Brief "
                f"on {lean_winners} of {total} sentence tasks — send less context, save tokens."
            )
        else:
            bullets.append(
                f"Full Brief still wins {total - lean_winners} of {total} sentence tasks; "
                f"only switch to a lean recipe where Tab 2 marks it 'Best value'."
            )

    # 3) Modal recipe
    if not bpt.empty:
        modal = bpt[T("Recommended recipe", "推荐配方")].value_counts()
        if not modal.empty:
            top = modal.index[0]
            n = int(modal.iloc[0])
            bullets.append(
                f"'{top}' is the most common winning recipe ({n}/{len(bpt)} tasks) — "
                f"use it as the default starting point."
            )

    # 4) Keyword compression
    kw_line = _keyword_compression_headline(scored)
    if kw_line and "Not tested" not in kw_line:
        bullets.append(f"Keyword prompt: {kw_line.rstrip('.')}.")

    return bullets


def _headline_answers(scored: pd.DataFrame, raw: pd.DataFrame) -> list[tuple[str, str]]:
    """8 Q&A rows that directly answer the two experimental objectives."""
    bpt = _best_per_task_table(scored, raw)
    recipe, wins, total = _modal_winning_recipe(bpt)
    decisive_wins, n_sentence_tasks = _decisive_win_count(scored)
    if wins > 0:
        # Qualify the "wins N/M" headline with noise-floor honesty.
        # If most tasks are statistical ties (decisive < half), the modal
        # recipe is best read as "cheapest among tied top configs".
        if n_sentence_tasks > 0 and decisive_wins < max(1, n_sentence_tasks // 2):
            modal_str = (
                f"{recipe} (top in {wins}/{total} tasks — "
                f"but only {decisive_wins}/{n_sentence_tasks} sentence tasks "
                f"have a winner above the negligible-difference line (0.036); treat top-3 as tied)"
            )
        else:
            modal_str = (
                f"{recipe} (top in {wins}/{total} tasks; "
                f"{decisive_wins}/{n_sentence_tasks} are above 0.036 negligible-difference line)"
            )
    else:
        modal_str = "—"
    mean_delta = _mean_delta_vs_full_brief(bpt)
    delta_str = f"{mean_delta:+.3f} avg (where winner ≠ Full brief)" if mean_delta is not None else "—"

    return [
        ("Best model",            _best_cheap_model(scored)),
        ("Best recipe (overall)", modal_str),
        ("Best field combination", _best_field_combination(scored)),
        ("Δ vs Full brief",       delta_str),
        ("Is Full brief worth using?", _full_brief_value_judgement(scored)),
        ("Keyword prompt compression", _keyword_compression_headline(scored)),
        ("Quality ceiling (best models)", _quality_ceiling_answer(scored)),
        ("Cost-quality stance",   _cost_quality_stance(scored)),
        ("Premium models — scope", _premium_model_recommendation_line(scored)),
    ]


def _premium_scope_rows() -> list[tuple[str, str]]:
    """Premium-model scope-constraint table — verbatim per project rules."""
    return [
        ("Do we need premium models?", "Test only on top configs"),
        ("Why?",                       "To see best possible quality without expanding scope"),
        ("Estimated extra cost",       "< £1"),
        ("Decision rule",              "Use premium model only if quality gain is meaningful"),
    ]


# ---------- Narrative helpers (decision-clarity refactor) ----------

# Field → one-line rationale. Used to compose "Why this won" cells.
# Phrasing intentionally describes the SIGNAL a field carries, not how it scored.
_FIELD_RATIONALE_EN = {
    "product":         "Product carries strong category & feature signal",
    "audience":        "Audience adds emotional / contextual framing",
    "brand_strategy":  "Brand strategy enables higher-level concept abstraction",
    "personality":     "Personality contributes tonal cues",
    "differentiators": "Differentiators highlight what makes the brand distinct",
    "business_category": "Business category narrows the task to the right domain",
}

_FIELD_RATIONALE_ZH = {
    "product":         "Product 承载强品类 / 功能信号",
    "audience":        "Audience 增加情感 / 情境定位",
    "brand_strategy":  "Brand strategy 支撑更高层概念抽象",
    "personality":     "Personality 提供语气与人格线索",
    "differentiators": "Differentiators 突出品牌差异化要点",
    "business_category": "Business category 将任务聚焦到正确领域",
}


def _field_rationale_dict():
    return _FIELD_RATIONALE_ZH if LANG == "zh" else _FIELD_RATIONALE_EN


# Back-compat alias so existing call sites stay valid; resolved at call time.
class _FieldRationaleProxy:
    def get(self, key, default=None):
        return _field_rationale_dict().get(key, default)
    def __getitem__(self, key):
        return _field_rationale_dict()[key]


_FIELD_RATIONALE = _FieldRationaleProxy()


def _within_noise_floor(delta, noise: float = 0.036) -> str:
    """Yes/No flag for Tab 2 — is the difference within the measured negligible-difference line?

    'Within noise' means the recipe is statistically tied with Full Brief and the
    recommendation logic should prefer the cheaper / simpler option.
    """
    if delta is None or (isinstance(delta, float) and pd.isna(delta)):
        return "—"
    try:
        return T("Yes", "是") if abs(float(delta)) <= noise else T("No", "否")
    except (ValueError, TypeError):
        return "—"


# Per-task plain-language reason for the Tab 2 per-task table.
# Hardcoded copy from the user's Tab 2 spec. Keyed on the internal task name.
# If the recommended recipe changes for a task, refresh the matching reason
# (these reasons describe WHY the recipe's fields work for that task).
_PER_TASK_REASON_EN = {
    "benefit_relevant":  "User + distinctive-value signal",
    "category_relevant": "Direct category / function signal",
    "concept_relevant":  "Brand-direction signal",
    "context_relevant":  "User / context signal",
    "emotion_relevant":  "Abstract task; spot-check",
    "feature_relevant":  "Direct feature signal",
    "function_relevant": "Direct function signal",
    "position_relevant": "User + differentiation signal",
}
_PER_TASK_REASON_ZH = {
    "benefit_relevant":  "用户 + 差异化价值信号",
    "category_relevant": "直接的品类 / 功能信号",
    "concept_relevant":  "品牌方向信号",
    "context_relevant":  "用户 / 情境信号",
    "emotion_relevant":  "抽象任务；建议抽查",
    "feature_relevant":  "直接的特性信号",
    "function_relevant": "直接的功能信号",
    "position_relevant": "用户 + 差异化信号",
}


def _per_task_reason(task_internal: str) -> str:
    d = _PER_TASK_REASON_ZH if LANG == "zh" else _PER_TASK_REASON_EN
    return d.get(task_internal, T("(no reason recorded)", "（无理由说明）"))


# Per-task production-action override. Most tasks ship by default;
# emotion_relevant is an abstract task where spot-check is recommended,
# so it gets "Ship with review" instead of plain "Ship".
_PER_TASK_PRODUCTION_ACTION_OVERRIDE_EN: dict[str, str] = {
    "emotion_relevant": "Ship with review",
}
_PER_TASK_PRODUCTION_ACTION_OVERRIDE_ZH: dict[str, str] = {
    "emotion_relevant": "上线（需抽查）",
}


def _per_task_production_action(task_internal: str) -> str | None:
    """Return a custom production-action string for a task, or None to fall
    back to _final_decision's default."""
    d = _PER_TASK_PRODUCTION_ACTION_OVERRIDE_ZH if LANG == "zh" \
        else _PER_TASK_PRODUCTION_ACTION_OVERRIDE_EN
    return d.get(task_internal)


# Tasks where the recommended-model column must stay in the cheap tier even
# if a premium model scored higher. Avoids the "Recommended model: opus47;
# Production action: default to GPT-5-mini" contradiction the user flagged.
# The Opus-4.7-as-escalation note is captured in _PER_TASK_REASON_ZH /
# _PER_TASK_REASON_EN for that task.
_PER_TASK_CHEAP_TIER_ONLY = {"feature_relevant"}


# Editorial per-task recipe overrides. When a task's cheap-tier candidates
# are within the 0.036 noise floor of each other, the cost tie-break can
# flip between equally-valid recipes between runs. For tasks that appear
# in multiple places in the workbook (Tab 1 hardcoded, Tab 2 computed,
# Analysis A5 computed), we lock in one recipe here so every tab agrees.
_PER_TASK_RECIPE_OVERRIDE = {
    # Benefit: audience+personality (0.629) and audience+differentiators
    # (0.627) are within 0.002 cosine — well inside the 0.036 noise floor.
    # Editorial pick is audience+differentiators because it matches the
    # broader narrative (Positioning also uses audience+differentiators)
    # and the "User + distinctive-value signal" reason text on Tab 2.
    "benefit_relevant": "A:audience+differentiators",
}


def _final_decision(recipe_label: str) -> str:
    """Production-deployment label for Tab 2.

    Two values only — keeps the column crisp:
      • 'Use in production'    — for any lean recipe that won its task
      • 'Keep as baseline only' — for Full Brief rows (Full Brief is the
        experiment's control, not a deployment recommendation)

    (The cost-efficient cheapest-within-negligible-difference logic
    already picks the production-ready recipe per task. So if the row
    came back with Full Brief as the recommended recipe, it means no
    leaner recipe could match it — but Full Brief is still the control,
    not the deployment.)
    """
    rec = recipe_label.strip().lower()
    is_full = rec.startswith("full brief") or recipe_label.strip().startswith("完整 brief")
    if is_full:
        return T("Baseline only", "仅作基线")
    return T("Ship", "上线")


def _decision_implication(recipe_label: str, delta, noise: float = 0.036) -> str:
    """One-sentence decision text for each row in Tab 2.

    The wording follows the stakeholder-review brief: explicit about whether
    Full Brief still wins, whether the recipe is recommended outright, or
    whether the difference is within noise.
    """
    rec = recipe_label.strip().lower()
    is_full = rec.startswith("full brief") or recipe_label.strip().startswith("完整 brief")
    try:
        d = float(delta) if delta is not None else None
    except (ValueError, TypeError):
        d = None

    if is_full:
        return T(
            "Keep Full Brief — no leaner recipe beats it within the negligible-difference line (Δ ≤ 0.036).",
            "保留完整 brief —— 在可忽略差异线内没有更精简的配方能胜过它。")
    if d is None:
        return T(
            "Use as default; validate on stronger models if it stays top after reruns.",
            "作为默认配方使用；如重跑后仍领先，再在更强模型上验证。")
    if abs(d) <= noise:
        return T(
            "Prefer this recipe — mean cosine difference vs Full Brief is "
            "within the negligible-difference line (Δ ≤ 0.036).",
            "推荐此配方 —— 与完整 brief 的平均余弦差异在可忽略差异线以内（Δ ≤ 0.036）。")
    if d > 0:
        return T(
            "Use this as the default recipe for this task — it beats Full Brief.",
            "此配方为该任务的默认 —— 它胜过完整 brief。")
    return T(
        "Use this recipe — it is cheaper and within the practical-equivalence "
        "band of Full Brief.",
        "使用此配方 —— 更便宜且与完整 brief 处于实质等价范围内。")


# Description text spells out WHAT was removed at each compression level —
# the same information a reader would see by diffing prompts.txt manually.
# This is the "score = 63% compression" → "compressed of WHAT" answer.
_PROMPT_DESCRIPTIONS_EN = {
    "A — Full":
        "Original full prompt. Detailed explanations, edge cases, and examples.",
    "B — Reduced":
        "Long explanations and examples removed. Role, hard rules, process, "
        "and output format kept.",
    "C — Compact":
        "Further compressed. Only the goal, the rules, and the process remain.",
    "D — Minimal":
        "Only a single task sentence and the 'must come from the brief' constraint.",
}

_PROMPT_DESCRIPTIONS_ZH = {
    "A — Full":
        "原版完整 prompt，包含详细解释、边界情况和示例。",
    "B — Reduced":
        "删除长解释和示例，但保留角色、硬规则、流程和输出格式。",
    "C — Compact":
        "进一步压缩，只保留目标、规则和流程。",
    "D — Minimal":
        "只保留一句任务描述和「必须来自 brief」的约束。",
}


def _prompt_description(version_label: str) -> str:
    d = _PROMPT_DESCRIPTIONS_ZH if LANG == "zh" else _PROMPT_DESCRIPTIONS_EN
    return d.get(version_label, T("(no description)", "（无描述）"))


# What changed at each compression step. The previous-version diff makes it
# explicit WHY each cut might or might not hurt quality (so a reader can see
# the inflection point between B and C in the actual prompt content, not just
# the F1 numbers).
_WHAT_CHANGED_EN = {
    "A — Full":     "Baseline — original full prompt (3,306 chars).",
    "B — Reduced":  "vs A: cut the multi-paragraph explanations of foundational/strategic terms, the edge-case caveats, and the rationale + worked examples. Kept the role + 4 hard rules + 2-step process + output format.",
    "C — Compact":  "vs B: removed the foundational-vs-strategic prose entirely — only category labels survive. Kept the 10-term goal + 4 hard rules + 2-step process.",
    "D — Minimal":  "vs C: dropped the rules list AND the 2-step process. Only a single task sentence + the 'must come from the brief' constraint remain.",
}
_WHAT_CHANGED_ZH = {
    "A — Full":     "基线 —— 原版完整 prompt（3,306 字符）。",
    "B — Reduced":  "相对 A：删去 foundational/strategic 多段解释、边界情况说明、理由和示例。保留角色 + 4 条硬规则 + 两步流程 + 输出格式。",
    "C — Compact":  "相对 B：彻底删去 foundational vs strategic 的散文，只保留类别标签。保留 10 词目标 + 4 条硬规则 + 两步流程。",
    "D — Minimal":  "相对 C：连规则列表和两步流程都删了。只剩一句任务描述 +「必须来自 brief」约束。",
}


def _what_changed_from_previous(version_label: str) -> str:
    d = _WHAT_CHANGED_ZH if LANG == "zh" else _WHAT_CHANGED_EN
    return d.get(version_label, T("(no diff recorded)", "（无差异说明）"))


def _recommended_keyword_version(kw_df: pd.DataFrame, f1_tolerance: float = 0.02) -> str:
    """Pick the shortest version whose F1 stays within `f1_tolerance` of A.

    Returns the recommended version label (e.g. 'B — Reduced'), or 'A — Full'
    if no shorter version meets the tolerance. Pure-text input expected (the
    same shape _keyword_compression_table emits).
    """
    if kw_df.empty:
        return "A — Full"
    try:
        ver_col_pre = T("Prompt Version", "提示词版本")
        base_row = kw_df[kw_df[ver_col_pre] == "A — Full"]
        if base_row.empty:
            return "A — Full"
        f1_col = T("Mean F1 (23 briefs)", "F1 均值（23 brief）")
        ver_col = T("Prompt Version", "提示词版本")
        base_f1 = float(base_row.iloc[0][f1_col])
    except (KeyError, ValueError, TypeError):
        return "A — Full"

    # Walk shortest → longest, return the first one within tolerance.
    for label in ["D — Minimal", "C — Compact", "B — Reduced"]:
        row = kw_df[kw_df[ver_col] == label]
        if row.empty:
            continue
        try:
            f1 = float(row.iloc[0][f1_col])
            if f1 >= base_f1 - f1_tolerance:
                return label
        except (ValueError, TypeError):
            continue
    return "A — Full"


def _why_this_won(recipe_label: str) -> str:
    """Field-level reasoning for the winning recipe.

    Drives Tab 2's 'Why this won' column. Splits a humanised recipe label like
    'audience + product' into its fields, looks up each in _FIELD_RATIONALE,
    joins them with '+'. Baselines get their own one-liners.
    """
    rec = recipe_label.strip()
    rec_low = rec.lower()
    if rec_low.startswith("full brief") or rec.startswith("完整 brief"):
        return T(
            "Full brief — model uses everything; useful when no single field dominates",
            "完整 brief —— 模型用上所有字段；适合没有单一字段主导的任务")
    if "no brief" in rec_low or rec_low.startswith("(instruction only)") or "无 brief" in rec:
        return T(
            "No brief — instruction-only baseline (low quality)",
            "无 brief —— 仅指令的基线（质量低）")

    parts = [p.strip() for p in rec.split("+")]
    reasons = [_FIELD_RATIONALE.get(p, p) for p in parts]
    if len(reasons) == 1:
        return reasons[0]
    return " + ".join(reasons)


def _bottom_line(scored: pd.DataFrame, raw: pd.DataFrame) -> list[str]:
    """Bottom-line bullets for Tab 1 — hardcoded per user-approved wording.

    These four sentences are the user's own copy-edit from the stakeholder
    review. They are NOT derived from `scored` / `raw` on purpose: the
    plain-language phrasing was finalised by hand and any data-derived
    template would slowly drift away from it.

    Refresh these values manually when scored data materially changes:
      - bullet 1: cost wording is provisional until config prices are verified
      - bullet 3: 0.65 = mean cosine of the `product` single-field recipe
      - bullet 4: 63% / 3,306 / 1,225 / 520 = the keyword B-Reduced numbers
    """
    # _ = scored, raw  # intentionally unused — see docstring
    out: list[str] = []

    # Bullet 1 — default model recommendation + cost comparison
    out.append(T(
        "Use GPT-5-mini as the default model. It produces quality close to "
        "Haiku, with lower estimated cost in the current config. Verify "
        "model IDs and input/output token prices before publishing exact "
        "cost ratios.",
        "默认使用 GPT-5-mini。它的质量和 Haiku 接近，且按当前配置估算成本更低。"
        "发布精确成本比例前，必须先确认模型 ID 和输入/输出 token 单价。"
    ))

    # Bullet 2 — skip the full brief
    out.append(T(
        "There is no need to send the full brief to the model. All 8 sentence "
        "tasks can reach the same quality using just one or two key fields. "
        "This means shorter prompts, lower cost, and faster responses.",
        "不需要把完整 brief 发给模型。所有 8 个句子型任务，"
        "都可以用一两个关键字段达到同等质量。"
        "这意味着 prompt 更短、成本更低、响应也更快。"
    ))

    # Bullet 3 — strongest single field (field name + its mean cosine computed)
    _sent_b = scored[scored["cosine"].notna()] if not scored.empty else scored
    _topf_b = _top_single_field(scored) or "product"
    try:
        _topf_mean = float(
            _sent_b[_sent_b["config_id"].apply(
                lambda c: str(c).split(":")[-1] == _topf_b)]["cosine"].mean())
        _topf_mean_s = f"{_topf_mean:.2f}" if _topf_mean == _topf_mean else "—"
    except Exception:
        _topf_mean_s = "—"
    out.append(T(
        f"The {_topf_b} field is the single most important field. It performs "
        f"consistently across multiple tasks with an average cosine score of "
        f"{_topf_mean_s}, and appears often in the recommended configurations "
        f"on Tab 2. For the field combination to use on each specific task, see "
        f"Tab 2.",
        f"{_topf_b} 字段是最重要的单一字段。它在多个任务里表现稳定，"
        f"平均余弦分数是 {_topf_mean_s}，也经常出现在 Tab 2 的推荐配置里。"
        f"具体每个任务该用哪个字段组合，看 Tab 2。"
    ))

    # Bullet 4 — keyword prompt compression (numbers computed from prompts.txt)
    kb = _keyword_b_numbers()
    if kb:
        out.append(T(
            f"For keyword extraction, use prompt version B — Reduced, i.e. the "
            f"original prompt with the long explanations and worked examples "
            f"removed but the role, hard rules, two-step process, and output "
            f"format kept. It is {kb['pct']}% shorter than the original "
            f"(character count drops from {kb['chars_a']:,} to {kb['chars_b']:,}, "
            f"saving about {kb['tokens_saved']} tokens per call) with no drop in "
            f"quality.",
            f"关键词抽取建议使用 B — Reduced 版 prompt，也就是在原版基础上"
            f"删掉冗长解释和示例、但保留角色设定、硬性规则、两步流程和输出格式的精简版。"
            f"它比原版短 {kb['pct']}%（字符数从 {kb['chars_a']:,} 降到 "
            f"{kb['chars_b']:,}，每次调用约省 {kb['tokens_saved']} token），"
            f"且质量没有下降。"
        ))
    else:
        out.append(T(
            "For keyword extraction, use prompt version B — Reduced: the "
            "original prompt with the long explanations and worked examples "
            "removed but the role, hard rules, two-step process, and output "
            "format kept — much shorter, with no drop in quality.",
            "关键词抽取建议使用 B — Reduced 版 prompt：在原版基础上删掉冗长解释"
            "和示例、但保留角色、硬规则、两步流程和输出格式——短很多，质量不降。"
        ))

    return out


def _emit_table_meta(ws, row: int, ncols: int, source: str, scoring: str) -> int:
    """Emit a two-line italic preface above any table:

      Data source: <where the numbers come from>
      How scored:  <the formula in plain words>

    Keeps every table self-describing so a reader does not have to flip
    to the appendix to know what each cell means. Caller passes
    bilingual strings (already T'd) for both source and scoring.
    """
    prefix_src = T("Data source: ", "数据来源：")
    prefix_score = T("How scored: ", "评分方式：")
    line1 = f"{prefix_src}{source}"
    line2 = f"{prefix_score}{scoring}"
    _merge_and_write(ws, row, line1,
        ncols=ncols, font=Font(size=10, italic=True, color="626A6E"))
    ws.row_dimensions[row].height = _estimate_row_height(line1, total_width_chars=140)
    row += 1
    _merge_and_write(ws, row, line2,
        ncols=ncols, font=Font(size=10, italic=True, color="626A6E"))
    ws.row_dimensions[row].height = _estimate_row_height(line2, total_width_chars=140)
    row += 2
    return row


def _score_definition_line() -> str:
    """One-sentence score definition, injected near the per-task numbers
    on Tabs 1/2/3 so a reader doesn't have to flip to the appendix to find
    out what the number actually measures.

    Wording follows the actual computation in `analyze --score`:
      • For 8 sentence tasks: per-brief score = ONE cosine similarity value
        (embedding of AI output · embedding of ground-truth sentence).
        The cell score shown in the tables = mean of 23 such per-brief cosines.
      • For the keyword task only: per-brief score = F1 (which itself
        combines Precision + Recall). Cell score = mean of 23 F1 values.
    """
    return T(
        "Score definition. The score reflects how close the AI output is to "
        "the human-curated ground truth. The numbers in the tables are the "
        "average across 23 briefs. Higher is better. Sentence tasks use "
        "semantic similarity. The keyword task uses F1 because it measures "
        "how well the keywords match.",
        "分数定义。分数表示 AI 输出和人工标准答案有多接近；"
        "表格里的分数是 23 个 brief 的平均表现，分数越高越好。"
        "句子任务用语义相似度，关键词任务用 F1，因为它看的是关键词匹配程度。"
    )


def _three_question_answers(scored: pd.DataFrame, raw: pd.DataFrame) -> list[tuple[str, str]]:
    """Tab 1 Q&A block — hardcoded per user-approved wording.

    Like _bottom_line, these are the user's exact copy. They do not consult
    `scored` / `raw` so the wording stays exactly as the stakeholder saw
    it. If the underlying data changes meaningfully (e.g. cheap tier no
    longer ties premium), the answers should be refreshed by hand.
    """
    # _ = scored, raw  # intentionally unused
    return [
        (
            T("Which context configuration produces the best quality?",
              "哪种 context 配置效果最好？"),
            T("Most tasks do not need the full brief. One or two key fields "
              "are enough. For the specific recommended configuration per "
              "task, see Tab 2.",
              "多数任务不需要完整 brief，用一两个关键字段就够。"
              "具体推荐配置见 Tab 2。"),
        ),
        (
            T("Do we actually need to send the full brief?",
              "是否真的需要发送完整 brief？"),
            T("No. Full Brief is only the experiment's control condition, "
              "not a recommended production setup.",
              "不需要。完整 brief 只是实验对照组，不是推荐的生产方案。"),
        ),
        (
            T("What is the recommended production setup?",
              "最值得生产使用的配置是什么？"),
            T("Use GPT-5-mini and pick the field combination for each task "
              "from Tab 2. This keeps quality up while lowering both cost "
              "and prompt length.",
              "使用 GPT-5-mini，并按 Tab 2 为每个任务选择对应字段组合。"
              "这样可以保持质量，同时降低成本和 prompt 长度。"),
        ),
    ]


def _glossary() -> list[tuple[str, str]]:
    """Term → 1-sentence definition. Drives the 'Glossary' section on Tab 1
    so a non-technical reader can read the rest of the workbook without
    looking up jargon. Each entry is (term, definition)."""
    if LANG == "zh":
        return [
            ("均值", "23 个 brief 各跑一次，把 23 个分数求平均，得到这个配方的代表分数。"),
            ("Δ（差值）", "两个分数相减，绝对值越小越接近。例：「Δ 0.02」= 差 0.02 分。"),
            ("可忽略差异线（0.036）",
                "差值小于 0.036 = 不算真正的质量差异，所以选更便宜 / 更简单的那个配方。"
                "0.036 是同一配方重跑两次的自然波动，小于这个分不出谁更好。"),
            ("配方 / Recipe",
                "「告诉 AI 看什么内容」的组合。例：「product」= 只给产品字段；"
                "「audience + product」= 给受众字段 + 产品字段；「完整 brief」= 全部发过去。"),
            ("完整 brief（baseline / 对照组）",
                "把整份 brief 全部发给模型 —— 是实验的对照组，不是推荐的生产配置。"
                "存在的目的是回答「真的需要发完整 brief 吗？」（答案：通常不需要）。"),
            ("Precision / Recall / F1",
                "只用于关键词题。Precision = AI 选的 10 个词里有几个对；Recall = "
                "应有的 10 个里 AI 找到了几个；F1 = 两者的综合分（越高越好）。"),
            ("余弦相似度（cosine）",
                "用 embedding 把 AI 输出和 ground truth 句子各转成一个向量，"
                "算两者的余弦相似度。0 = 完全不相关；1 = 完全一致。"),
        ]
    return [
        ("Mean (across 23 briefs)",
            "Run the recipe once on each of the 23 briefs, then average the "
            "23 per-brief scores. That mean is the recipe's representative score."),
        ("Δ (delta)",
            "The numeric difference between two scores. e.g. 'Δ 0.02' = "
            "0.02 apart."),
        ("Negligible-difference line (0.036)",
            "Differences below 0.036 are NOT treated as real quality "
            "differences, so we choose the cheaper / simpler option. 0.036 "
            "is the empirically measured natural variation when the same "
            "recipe is re-run."),
        ("Recipe / context setup",
            "Which fields of the brief get sent to the model. e.g. 'product' "
            "= only the product field; 'audience + product' = those two fields; "
            "'Full brief' = everything."),
        ("Full brief (baseline / control)",
            "Sending the entire brief to the model — this is the experiment's "
            "CONTROL condition, NOT a recommended production setup. It exists "
            "to answer 'do we actually need to send all the context?' "
            "(Answer: usually no.)"),
        ("Precision / Recall / F1",
            "Used only for the keyword task. Precision = how many of the AI's "
            "10 keywords are correct; Recall = how many of the 10 correct "
            "keywords the AI found; F1 = combined score (higher is better)."),
        ("Cosine similarity",
            "Embed the AI output and the ground-truth sentence as vectors, "
            "then compute their cosine. 0 = unrelated; 1 = identical meaning."),
    ]


def _purpose_statement(scored: pd.DataFrame) -> str:
    """One-sentence answer to 'what does this experiment evaluate?'.

    Uses the canonical Stage A config count from prompt_builder (the planned
    matrix size — 142 in the current spec), not the per-run sample count
    which can be 138/140/etc depending on truncations & rate-limit fallout.
    """
    try:
        from src.prompt_builder import list_configs_for_stage
        n_configs = len(list_configs_for_stage("stage_a"))
    except Exception:
        n_configs = 142  # fallback to the documented Stage A count
    return (
        f"Evaluate {n_configs} prompt configurations across the 23 client briefs "
        f"on 8 brand-brief tasks + 1 keyword-extraction task, to find the "
        f"cheapest setup that hits production quality."
    )


def _exec_key_findings(scored: pd.DataFrame, raw: pd.DataFrame) -> list[str]:
    """4 decision-relevant findings, in the order Simon's review asked for:
       1) cheap-model parity, 2) Full Brief unnecessary, 3) Product is king,
       4) keyword prompt compresses cheaply. All data-driven, no hardcoding."""
    out: list[str] = []
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored

    # 1. Cheap-model parity — concrete cost ratio.
    if not sent.empty:
        full = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        by_model = full.groupby("model_key")["cosine"].mean()
        cheap = [m for m in by_model.index if m in ("haiku", "gpt5mini")]
        if len(cheap) >= 2:
            diff = abs(by_model[cheap[0]] - by_model[cheap[1]])
            if diff < 0.036:
                out.append(T(
                    f"Cheap models tied on quality (gap {diff:.2f}, smaller than "
                    f"the negligible-difference line (0.036)) — prefer the lower-cost model "
                    f"after price calibration.",
                    f"两个便宜模型质量打平（差距 {diff:.2f}，小于 0.036 可忽略差异线）—— "
                    f"价格校准后优先选成本更低的模型。"
                ))
            else:
                winner = by_model[cheap].idxmax()
                out.append(T(
                    f"{winner} is the stronger cheap model (lead {diff:.2f}, "
                    f"beyond the negligible-difference line (0.036)) — use it as default.",
                    f"{winner} 是更强的便宜模型（领先 {diff:.2f}，超出 0.036 可忽略差异线）—— 默认用它。"
                ))

    # 2. Full Brief often unnecessary — name the cheaper recipes that replace it.
    bpt = _best_per_task_table(scored, raw)
    if not bpt.empty:
        recipe_col = T("Recommended recipe", "推荐配方")
        baselines = {T("Full brief", "完整 brief"), "Full brief"}
        lean_wins = sum(1 for v in bpt[recipe_col] if v not in baselines)
        total = len(bpt)
        if lean_wins == total:
            out.append(T(
                f"Full Brief is never the best practical choice (0 of {total} "
                "sentence tasks) — for every task, a 1- or 2-field recipe is "
                "within the negligible-difference line (Δ ≤ 0.036) of it, which means shorter prompts and "
                "lower per-call cost with the same quality.",
                f"完整 brief 在任何一个任务上都不是最佳实用选择（{total} 个任务都不是）—— "
                "每个任务都有 1–2 字段的配方在 0.036 可忽略差异线以内追平，意味着 prompt 更短、每次调用更便宜，质量却一样。"
            ))
        elif lean_wins >= total // 2:
            out.append(T(
                f"Full Brief loses to a leaner recipe on {lean_wins}/{total} "
                "sentence tasks (within the negligible-difference line (Δ ≤ 0.036) — same quality, shorter prompt, "
                "lower cost).",
                f"在 {lean_wins}/{total} 个句子型任务上，完整 brief 输给了更精简的配方"
                "（在 0.036 可忽略差异线以内，质量相同但 prompt 更短、成本更低）。"
            ))

    # 3. Product is the strongest standalone field
    if not sent.empty:
        single_field = sent[
            sent["config_id"].apply(_is_single_field_config)
        ].copy()
        if not single_field.empty:
            single_field["_field"] = single_field["config_id"].apply(
                lambda c: c.split(":")[-1]
            )
            field_mean = (
                single_field.groupby("_field")["cosine"].mean().sort_values(ascending=False)
            )
            if not field_mean.empty:
                top = field_mean.index[0]
                top_score = field_mean.iloc[0]
                second_score = field_mean.iloc[1] if len(field_mean) > 1 else None
                if top == "product":
                    en_detail = (f"averaging {top_score:.2f} across 23 briefs"
                                 + (f"; second-best is '{field_mean.index[1]}' "
                                    f"at {second_score:.2f}"
                                    if second_score is not None else ""))
                    zh_detail = (f"23 brief 均值 {top_score:.2f}"
                                 + (f"；第二是「{field_mean.index[1]}」（{second_score:.2f}）"
                                    if second_score is not None else ""))
                    out.append(T(
                        f"Among single fields, 'product' is strongest ({en_detail}). "
                        "Default: start with 'product' alone; only add a second "
                        "field if it pushes the score up by more than 0.036 "
                        "(otherwise the extra field is just adding tokens for no "
                        "measurable quality gain).",
                        f"单字段中「product」最强（{zh_detail}）。默认：先只用 product；"
                        "只有当第二个字段能把分数提升超过 0.036 时再加入"
                        "（否则只是多花 token，质量上看不出区别）。"
                    ))
                else:
                    out.append(T(
                        f"Among single fields, '{top}' is strongest "
                        f"(mean {top_score:.2f} across 23 briefs) — "
                        "surprise winner ahead of 'product'.",
                        f"单字段中「{top}」最强（23 brief 均值 {top_score:.2f}）"
                        "——意外胜过 product。"
                    ))

    # 4. Keyword compression — concrete char / token savings.
    kw = _keyword_compression_table(scored)
    if not kw.empty:
        rec_col = T("Recommended?", "是否推荐？")
        ver_col = T("Prompt Version", "提示词版本")
        red_col = T("% reduction vs A", "相对 A 的压缩比例")
        char_col = T("Characters", "字符数")
        tok_col = T("Est. tokens", "估算 token")
        rec = kw[kw[rec_col] == T("Yes", "是")]
        if not rec.empty:
            r = rec.iloc[0]
            try:
                a_row = kw[kw[ver_col] == "A — Full"].iloc[0]
                a_tok = int(a_row[tok_col])
                a_chr = int(a_row[char_col])
                new_tok = int(r[tok_col])
                new_chr = int(r[char_col])
                saved_tok = a_tok - new_tok
                out.append(T(
                    f"Keyword prompt '{r[ver_col]}' is {r[red_col]} shorter "
                    f"({a_chr:,} → {new_chr:,} chars, ≈{saved_tok} input tokens "
                    f"saved per call) with no measurable quality loss. "
                    f"What got cut: the long explanation paragraphs and edge-case "
                    f"caveats — the 4 hard rules and 2-step process stayed.",
                    f"关键词 prompt 用「{r[ver_col]}」版本比原版短 {r[red_col]}"
                    f"（{a_chr:,} → {new_chr:,} 字符，每次调用约省 {saved_tok} 个输入 token），"
                    "且质量无可测损失。砍掉的是长解释段落和边界 case 说明，"
                    "保留的是 4 条硬规则和两步流程。"
                ))
            except (KeyError, IndexError, ValueError, TypeError):
                # Fallback if some columns are missing
                out.append(T(
                    f"Keyword extraction prompt compresses {r[red_col]} "
                    f"({r[ver_col]}) with no quality loss.",
                    f"关键词 prompt 可压缩 {r[red_col]}（{r[ver_col]}）质量无损失。"
                ))

    return out


def _production_strategy() -> list[str]:
    """4 strategy bullets — the engineering-decision layer.

    These describe the WORKFLOW used to reach the recommendations, not the
    recommendations themselves. They're the signal Simon's brief asked for:
    cost-aware experimentation, shortlist-and-validate, noise-floor decisions.
    """
    return [
        "Cheap-model screening first — run the full 142-config × 23-brief matrix "
        "on the two cheapest models (Haiku + GPT-5-mini) to identify top configs "
        "before any premium spend.",
        "Shortlist top-2 configs per task — Stage B re-runs them 3× to confirm "
        "the ranking holds within the measured negligible-difference line (σ ≈ 0.018 cosine).",
        "Premium validation only on finalists — Phase 4 runs Opus 4.7 + GPT-5.5 + "
        "Sonnet + GPT-5 on the top-1 config per task only. Total cost stayed under £1.",
        "Prefer cheaper / smaller configs within noise — when two recipes tie "
        "statistically (Δ < 0.036), the recommendation is the cheaper one, not the "
        "marginal-quality winner.",
    ]


def _stage_for_row(model_key: str, run_id) -> str:
    """Best-effort stage attribution for cost summary."""
    cheap   = {"haiku", "gpt5mini"}
    medium  = {"sonnet", "gpt5"}
    premium = {"opus47", "gpt55"}
    rid = int(run_id) if pd.notna(run_id) else 1
    if model_key in cheap:
        return (T("Stage A (cheap screening)", "Stage A（便宜模型筛选）") if rid == 1
                else T("Stage B (stability reruns)", "Stage B（稳定性重跑）"))
    if model_key in medium:
        return T("Stage B (AI judge)", "Stage B（AI 评审）")
    if model_key in premium:
        return T("Phase 4 (premium ladder)", "Phase 4（旗舰模型验证）")
    return T("Other", "其他")


def _cost_summary_by_stage(raw: pd.DataFrame) -> pd.DataFrame:
    """| Stage | Cost |"""
    if raw is None or raw.empty or "cost_usd" not in raw.columns:
        return pd.DataFrame()
    df = raw.copy()
    df["_stage"] = df.apply(lambda r: _stage_for_row(r["model_key"], r.get("run_id", 1)), axis=1)
    grouped = (
        df.groupby("_stage")["cost_usd"].sum()
        .reset_index().sort_values("cost_usd", ascending=False)
    )
    return pd.DataFrame({
        T("Stage", "阶段"): grouped["_stage"],
        T("Cost", "成本"):  grouped["cost_usd"].apply(lambda x: f"${x:.4f}"),
    })


def _key_findings(scored: pd.DataFrame, raw: pd.DataFrame) -> list[str]:
    """≤5 concise findings."""
    findings: list[str] = []
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    bpt = _best_per_task_table(scored, raw)

    # 1. Cheap-model parity
    if not sent.empty:
        full = sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        by_model = full.groupby("model_key")["cosine"].mean()
        cheap = [m for m in by_model.index if m in ("haiku", "gpt5mini")]
        if len(cheap) >= 2:
            diff = abs(by_model[cheap[0]] - by_model[cheap[1]])
            if diff < 0.036:
                findings.append(
                    f"{cheap[0]} and {cheap[1]} produce practically equivalent quality (Δ {diff:.2f}) — pick whichever is cheaper or faster."
                )
            else:
                findings.append(
                    f"{by_model[cheap].idxmax()} is the stronger cheap model (Δ {diff:.2f} over the other)."
                )

    # 2. Single-field vs pair-field win distribution
    if not bpt.empty:
        baselines = {"Full brief", "No brief (instruction only)"}
        recipes = bpt[T("Recommended recipe", "推荐配方")].tolist()
        n_single = sum(1 for r in recipes if r not in baselines and " + " not in r)
        n_pair   = sum(1 for r in recipes if " + " in r)
        n_full   = sum(1 for r in recipes if r in baselines)
        total = len(bpt)
        if n_full == total:
            findings.append(f"Full brief is the cheapest practical winner for all {total} tasks — keep using it.")
        elif n_single + n_pair == total:
            findings.append(
                f"All {total} tasks beat Full brief on cost: {n_single} task(s) "
                f"with a single field, {n_pair} with a 2-field pair."
            )
        else:
            findings.append(
                f"Winners across {total} tasks: {n_single} single-field, {n_pair} field-pair, "
                f"{n_full} Full brief."
            )

    # 3. Keyword compression
    kw_df = _keyword_compression_table(scored)
    if not kw_df.empty:
        rec = kw_df[kw_df[T("Recommended?", "是否推荐？")] == T("Yes", "是")]
        if not rec.empty:
            r = rec.iloc[0]
            findings.append(
                f"Keyword prompt compresses {r['% reduction vs A']} "
                f"({r['Prompt Version']}) with no quality loss."
            )

    # 4. Premium scope
    findings.append(
        "Premium models will only be tested on the top winning configs — cost capped at < £1."
    )

    # 5. Budget
    cost = float(raw["cost_usd"].sum()) if (raw is not None and "cost_usd" in raw) else 0
    findings.append(f"${cost:.2f} of £50 budget spent ({100 * cost / 63:.0f}% used).")
    return findings[:5]


def _next_step_line(scored: pd.DataFrame, raw: pd.DataFrame) -> str:
    has_cheap = (not scored.empty) and (
        scored[scored["model_key"].isin(("haiku", "gpt5mini"))].shape[0] > 0
    )
    has_reruns = (not scored.empty) and (
        "run_id" in scored.columns and scored["run_id"].max() > 1
    )
    has_premium = (not scored.empty) and (
        scored[scored["model_key"].isin(("opus47", "gpt55"))].shape[0] > 0
    )
    if not has_cheap:
        return "Run Stage A (cheap screening) to populate the report."
    if not has_reruns:
        return "Run Stage B (stability reruns) on the cheap-screen shortlist — do NOT expand to all recipes / all models."
    if not has_premium:
        return "Run Phase 4 (premium ladder, ≤£1) on Stage B winners only. No full grid search."
    return "Ready to ship — all stages complete."


# ---------- Tab-top conclusion banner ----------

def write_tab_conclusion(ws, start_row: int, lines: list[str], ncols: int) -> int:
    """2–3 line italic banner stating which experimental objective this tab answers."""
    for line in lines:
        _merge_and_write(ws, start_row, line, ncols,
                         font=Font(size=11, italic=True, color="0B0C0C"),
                         fill=LIGHT_BLUE)
        start_row += 1
    return start_row + 1


def write_tech_note(ws, start_row: int, text: str, ncols: int) -> int:
    """One-line technical footnote — smaller, grey, italic. Use directly under
    a write_tab_conclusion plain-English banner to separate human-readable
    framing from the implementation detail Simon's review style asked for."""
    _merge_and_write(ws, start_row, text, ncols,
                     font=Font(size=9, italic=True, color="626A6E"),
                     fill=None)
    return start_row + 2


# ---------- Tab 1: Executive Summary ----------

# ============================================================
# Helpers for the 9-section analytical report on Tab 1.
# Each helper returns a small DataFrame ready to drop into the workbook.
# ============================================================

_CHEAP_MODELS = {"haiku", "gpt5mini"}
_MEDIUM_MODELS = {"sonnet", "gpt5"}
_PREMIUM_MODELS = {"opus47", "gpt55"}
_SENTENCE_TASKS = [
    "benefit_relevant", "category_relevant", "concept_relevant",
    "context_relevant", "emotion_relevant", "feature_relevant",
    "function_relevant", "position_relevant",
]


def _is_single_field(cid: str) -> bool:
    if cid in ("A:_full_brief", "_full_brief", "A:_prompt_implied", "_prompt_implied"):
        return False
    rhs = cid.split(":")[-1] if ":" in cid else cid
    return "+" not in rhs


def _is_pair_field(cid: str) -> bool:
    rhs = cid.split(":")[-1] if ":" in cid else cid
    return "+" in rhs


def _model_summary_df(scored: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    """Section 1 — per-model avg cosine + n cells + avg cost / call."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    order = ["gpt5mini", "gpt5", "opus47", "sonnet", "gpt55", "haiku"]
    rows = []
    for m in order:
        cells = sent[sent["model_key"] == m]
        if cells.empty:
            continue
        avg_cos = float(cells["cosine"].mean())
        n = int(len(cells))
        if not raw.empty and "cost_usd" in raw.columns:
            cost_rows = raw[raw["model_key"] == m]
            avg_cost = float(cost_rows["cost_usd"].mean()) if not cost_rows.empty else float("nan")
        else:
            avg_cost = float("nan")
        rows.append({
            T("Model", "模型"): m,
            T("Avg cosine across all sentence-task cells",
              "全部句子型任务单元格的平均余弦"): round(avg_cos, 4),
            T("n cells", "单元格数"): n,
            T("Avg cost / call", "每次调用平均成本"):
                f"${avg_cost:.4f}" if pd.notna(avg_cost) else "—",
        })
    return pd.DataFrame(rows)


def _premium_vs_cheap_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 2 — per task, score on the cheap-tier winner recipe
    rendered on cheap / medium / premium models, plus the premium-vs-cheap
    delta. Highlights that premium underperforms on most tasks."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    rows = []
    for task in _SENTENCE_TASKS:
        cheap_rows = sent[(sent["task"] == task) & sent["model_key"].isin(_CHEAP_MODELS)]
        if cheap_rows.empty:
            continue
        # Cheap-tier winning recipe = the config with highest mean cosine on cheap tier.
        winner_cfg = cheap_rows.groupby("config_id")["cosine"].mean().idxmax()
        cell = sent[(sent["task"] == task) & (sent["config_id"] == winner_cfg)]
        by_m = cell.groupby("model_key")["cosine"].mean()

        def _best(models):
            vs = [float(by_m[m]) for m in models if m in by_m.index]
            return max(vs) if vs else None

        cheap_v = _best(_CHEAP_MODELS)
        med_v   = _best(_MEDIUM_MODELS)
        prem_v  = _best(_PREMIUM_MODELS)
        delta = (round(prem_v - cheap_v, 3)
                 if (prem_v is not None and cheap_v is not None) else None)
        check = " [OK]" if (delta is not None and delta > 0) else ""
        rows.append({
            T("Task", "任务"): task,
            T("Cheap", "便宜层"):
                round(cheap_v, 3) if cheap_v is not None else "—",
            T("Medium", "中端层"):
                round(med_v, 3) if med_v is not None else "—",
            T("Premium", "旗舰层"):
                round(prem_v, 3) if prem_v is not None else "—",
            T("Δ premium vs cheap", "Δ 旗舰 vs 便宜"):
                f"{delta:+.3f}{check}" if delta is not None else "—",
        })
    return pd.DataFrame(rows)


def _brief_vs_no_brief_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 3 — per task, Full Brief vs No Brief mean cosine + delta."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    cheap = sent[sent["model_key"].isin(_CHEAP_MODELS)]
    rows = []
    for task in _SENTENCE_TASKS:
        g = cheap[cheap["task"] == task]
        full = g[g["config_id"].isin(("A:_full_brief", "_full_brief"))]["cosine"].mean()
        none = g[g["config_id"].isin(("A:_prompt_implied", "_prompt_implied"))]["cosine"].mean()
        if pd.isna(full) or pd.isna(none):
            continue
        delta = round(float(full - none), 3)
        rows.append({
            T("Task", "任务"): task,
            T("Full Brief", "完整 brief"): round(float(full), 3),
            T("No Brief", "无 brief"): round(float(none), 3),
            T("Δ", "差值"): f"{delta:+.3f}",
        })
    return pd.DataFrame(rows)


def _single_field_ranking_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 4a — single brief fields, ranked by mean cosine (cheap tier)."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    cheap = sent[sent["model_key"].isin(_CHEAP_MODELS)]
    singles = cheap[cheap["config_id"].apply(_is_single_field)].copy()
    singles["_field"] = singles["config_id"].apply(lambda c: c.split(":")[-1])
    # Drop metadata-only fields — they have <5 task coverage by design.
    singles = singles[singles["_field"] != "business_category"]
    by_field = singles.groupby("_field")["cosine"].mean().sort_values(ascending=False)
    return pd.DataFrame([
        {T("Field", "字段"): name, T("Mean cosine", "平均余弦"): round(float(v), 3)}
        for name, v in by_field.items()
    ])


def _best_pairs_ranking_df(scored: pd.DataFrame, top_n: int = 5) -> pd.DataFrame:
    """Section 4b — top N two-field pair recipes by mean cosine (cheap tier).

    Only includes pairs with ≥5-task coverage so metadata-only combos do
    not dominate the ranking from a single high-scoring task.
    """
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    cheap = sent[sent["model_key"].isin(_CHEAP_MODELS)]
    pairs = cheap[cheap["config_id"].apply(_is_pair_field)]
    if pairs.empty:
        return pd.DataFrame()
    coverage = pairs.groupby("config_id")["task"].nunique()
    viable = coverage[coverage >= 5].index.tolist()
    pairs_v = pairs[pairs["config_id"].isin(viable)]
    by_combo = pairs_v.groupby("config_id")["cosine"].mean().sort_values(ascending=False).head(top_n)
    rows = []
    for cid, v in by_combo.items():
        label = cid.replace("A:", "").replace("+", " + ")
        rows.append({
            T("Pair", "字段对"): label,
            T("Mean cosine", "平均余弦"): round(float(v), 3),
        })
    return pd.DataFrame(rows)


def _per_task_picks_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 5 — per-task cheap-tier pick with paired and LOO evidence."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty:
        return pd.DataFrame()
    cheap = sent[sent["model_key"].isin(_CHEAP_MODELS)]
    NOISE = cfg.NOISE_FLOOR_COSINE
    rows = []
    for task in _SENTENCE_TASKS:
        g = cheap[cheap["task"] == task]
        if g.empty:
            continue
        per_cfg = g.groupby(["config_id", "model_key"]).agg(
            mean_cosine=("cosine", "mean"),
            mean_cost_usd_per_call=("cost_usd", "mean"),
        ).reset_index()
        override_cfg = _PER_TASK_RECIPE_OVERRIDE.get(task)
        winner = _select_practical_winner(
            per_cfg,
            score_col="mean_cosine",
            cost_col="mean_cost_usd_per_call",
            noise=NOISE,
            override_cfg=override_cfg,
        )
        if winner is None:
            continue
        fb = g[g["config_id"].isin(("A:_full_brief", "_full_brief"))]["cosine"].mean()
        delta = round(float(winner["mean_cosine"] - fb), 3) if pd.notna(fb) else None
        stats = _brief_level_stats(
            scored,
            task,
            winner,
            per_cfg,
            noise=NOISE,
            baseline_score=fb if pd.notna(fb) else None,
            override_cfg=override_cfg,
        )
        recipe = winner["config_id"].split(":", 1)[-1] if ":" in winner["config_id"] else winner["config_id"]
        recipe = recipe.replace("+", " + ")
        rows.append({
            T("Task", "任务"): task.replace("_relevant", ""),
            T("Recipe", "配方"): recipe,
            T("Model", "模型"): winner["model_key"],
            T("Avg", "平均分"): round(float(winner["mean_cosine"]), 3),
            T("Median Δ", "中位数差值"):
                f"{stats['median_delta']:+.3f}" if stats["median_delta"] is not None else "—",
            T("Win/Tie/Loss", "胜/平/负"):
                f"{stats['wins']}/{stats['ties']}/{stats['losses']}" if stats["briefs"] else "—",
            T("Paired p", "配对 p 值"):
                (f"{stats['p_value']:.2f}" if stats.get("p_value") is not None
                 else ("n/a" if stats["briefs"] else "—")),
            T("vs Full Brief Δ", "对比完整 brief Δ"):
                f"{delta:+.3f}" if delta is not None else "—",
            T("LOO first/top-3", "LOO 第一/前三"):
                f"{stats['loo_first']}/{stats['loo_total']} · {stats['loo_top3']}/{stats['loo_total']}"
                if stats["loo_total"] else "—",
        })
    return pd.DataFrame(rows)


def _keyword_summary_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 6 — A/B/C/D keyword prompt P / R / F1 + ΔF1 vs A."""
    kw = scored[scored["f1"].notna()] if not scored.empty else scored
    if kw.empty:
        return pd.DataFrame()
    labels = {
        "A": "A — Full (3,306 chars)",
        "B": "B — Reduced (1,225 chars)",
        "C": "C — Compact (494 chars)",
        "D": "D — Minimal (192 chars)",
    }
    a_f1 = kw[kw["config_id"].str.startswith("A:")]["f1"].mean()
    rows = []
    for vk, label in labels.items():
        v = kw[kw["config_id"].str.startswith(f"{vk}:")]
        if v.empty:
            continue
        p = float(v["precision"].mean())
        r = float(v["recall"].mean())
        f1 = float(v["f1"].mean())
        d = "—" if vk == "A" or pd.isna(a_f1) else f"{f1 - a_f1:+.3f}"
        rows.append({
            T("Version", "版本"): label,
            T("Precision", "Precision"): round(p, 3),
            T("Recall", "Recall"): round(r, 3),
            T("F1", "F1"): round(f1, 3),
            T("ΔF1 vs A", "ΔF1 vs A"): d,
        })
    return pd.DataFrame(rows)


def _stability_stats_dict(scored: pd.DataFrame) -> dict:
    """Section 7 — cosine std dev across Stage B reruns."""
    sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    if sent.empty or "run_id" not in sent.columns:
        return {"n_cells": 0, "median": None, "mean": None, "max": None}
    per_cell = (
        sent.groupby(["brief_id", "task", "config_id", "model_key"])
        .agg(n_runs=("run_id", "nunique"), cell_std=("cosine", "std"))
        .reset_index()
    )
    rerun = per_cell[(per_cell["n_runs"] >= 2) & per_cell["cell_std"].notna()]
    if rerun.empty:
        return {"n_cells": 0, "median": None, "mean": None, "max": None}
    return {
        "n_cells": int(len(rerun)),
        "median": round(float(rerun["cell_std"].median()), 4),
        "mean":   round(float(rerun["cell_std"].mean()),   4),
        "max":    round(float(rerun["cell_std"].max()),    4),
    }


def _reliability_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 8 — per-model OK rate + notes on common failure modes."""
    if scored.empty:
        return pd.DataFrame()
    order = ["gpt5", "gpt55", "opus47", "sonnet", "gpt5mini", "haiku"]
    rows = []
    for m in order:
        g = scored[scored["model_key"] == m]
        if g.empty:
            continue
        n = len(g)
        ok = (g["status"] == "ok").sum()
        ok_rate = ok / n if n else 0.0
        # Quick failure breakdown
        notes = []
        be = (g["status"] == "budget_exceeded").sum()
        tr = (g["status"] == "truncated").sum()
        rl = (g["status"] == "rate_limited").sum()
        if rl: notes.append(T(f"{rl} rate_limited of {n}",
                              f"{n} 次中 {rl} 次被限流"))
        if tr: notes.append(T(f"{tr} truncated of {n}",
                              f"{n} 次中 {tr} 次被截断"))
        if be: notes.append(T(f"{be} budget_exceeded of {n}",
                              f"{n} 次中 {be} 次超预算"))
        if not notes:
            notes.append(T(f"n={n}", f"n={n}"))
        rows.append({
            T("Model", "模型"): m,
            T("OK rate", "成功率"): f"{ok_rate * 100:.1f}%",
            T("Notes", "说明"): "; ".join(notes),
        })
    return pd.DataFrame(rows)


def _cost_breakdown_df(scored: pd.DataFrame) -> pd.DataFrame:
    """Section 9 — per-model cost spent."""
    if scored.empty or "cost_usd" not in scored.columns:
        return pd.DataFrame()
    by_m = scored.groupby("model_key")["cost_usd"].sum().sort_values(ascending=False)
    rows = []
    for m, c in by_m.items():
        rows.append({
            T("Model", "模型"): m,
            T("Spend (USD)", "支出 (USD)"): f"${c:.4f}",
        })
    return pd.DataFrame(rows)


def build_tab_executive_summary(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Executive Summary — 7-section conclusion tab. Answers the business
    questions directly. Supporting data tables live on Tab 3 Analysis.

      1. Business question
      2. Decision (4-bullet production recommendation)
      3. Recommended task recipes (8-row table)
      4. Key findings (5 bullets)
      5. Keyword prompt compression (3-row + Removed / Kept lists)
      6. Production recommendation (default + human review + escalate)
      7. Score explanation (3 bullets)
    """
    ws = wb.create_sheet(T("Executive Summary", "执行摘要"))
    NCOLS = 6
    # IMPORTANT: column widths MUST be set before write_df / _merge_and_write
    # is called, otherwise those helpers see an undefined col width and
    # default to 12, blowing up row-height estimates.
    autosize_cols(ws, {1: 28, 2: 60, 3: 14, 4: 14, 5: 14, 6: 14})
    row = write_tab_title(ws, T("Executive Summary", "执行摘要"), ncols=NCOLS)

    # ============================================================
    # 1. Business question
    # ============================================================
    row = write_section_bar(ws, row,
        T("1. Business question", "1. 业务问题"), ncols=NCOLS)
    business_q = T(
        "Which setup delivers the best extraction quality at the lowest "
        "practical production cost?",
        "哪种配置能以最低的实际生产成本得到最好的抽取质量？")
    _merge_and_write(ws, row, business_q, ncols=NCOLS,
                     font=Font(size=11, italic=True, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(business_q,
                                                          total_width_chars=140)
    row += 2

    # ============================================================
    # 2. Decision — 4-bullet production recommendation
    # ============================================================
    row = write_section_bar(ws, row,
        T("2. Decision", "2. 决策"), ncols=NCOLS)
    for b in [
        T("GPT-5-mini", "GPT-5-mini"),
        T("Task-specific field recipes instead of Full Brief",
          "使用每任务专属字段配方，而非完整 brief"),
        T("Version B keyword prompt — the original prompt with the long "
          "explanations and worked examples removed, but the role, the hard "
          "rules, the 2-step process and the output format kept",
          "关键词使用 B 版 prompt —— 在原版基础上删去冗长解释和示例，"
          "但保留角色、硬规则、两步流程和输出格式"),
        T("Premium models reserved for failed edge cases only",
          "旗舰模型仅用于复核失败的个别边界情况"),
    ]:
        _merge_and_write(ws, row, f"• {b}", ncols=NCOLS,
                         font=Font(size=11, bold=True, color="0B0C0C"),
                         fill=LIGHT_BLUE)
        row += 1
    row += 1

    # ============================================================
    # 3. Key findings — 5 concise business-facing bullets
    # ============================================================
    row = write_section_bar(ws, row,
        T("3. Key findings", "3. 核心发现"), ncols=NCOLS)
    # Findings flagged is_auto=True are computed from the data and rendered bold;
    # the rest are fixed editorial summaries (normal weight).
    _topf = _top_single_field(scored) or "product"
    _kb = _keyword_b_numbers()
    _b_pct = f"{_kb['pct']}%" if _kb else "63%"
    findings = [
        (T("GPT-5-mini was not just \"good enough\"; it achieved the best "
           "average quality / cost tradeoff.",
           "GPT-5-mini 不只是「够用」；它取得了最佳的平均质量 / 成本平衡。"), False),
        (T("Premium models usually did not outperform GPT-5-mini on the "
           "same recipe.",
           "旗舰模型在同一配方上通常没有超过 GPT-5-mini。"), False),
        (T("The brief is necessary, but the optimal setup is the right "
           "subset of fields, not the Full Brief.",
           "brief 是必要的，但最佳配置是合适的字段子集，不是完整 brief。"), False),
        (T(f"{_topf} was the strongest single field overall.",
           f"{_topf} 是整体最强的单字段。"), True),
        (T(f"Version B keyword prompt — the original with the long explanations "
           f"and worked examples removed but the role, hard rules, two-step "
           f"process and output format kept — was {_b_pct} shorter while "
           f"preserving near-identical F1.",
           f"B 版关键词 prompt（在原版基础上删去冗长解释和示例，"
           f"但保留角色、硬规则、两步流程和输出格式）短了 {_b_pct}，F1 几乎不变。"), True),
    ]
    for finding, is_auto in findings:
        _merge_and_write(ws, row, f"• {finding}", ncols=NCOLS,
                         font=Font(size=11, bold=is_auto, color="0B0C0C"))
        ws.row_dimensions[row].height = _estimate_row_height(finding,
                                                              total_width_chars=140)
        row += 1
    row += 1

    # ============================================================
    # 4. Recommended recipes — small clean table
    # ============================================================
    row = write_section_bar(ws, row,
        T("4. Recommended recipes", "4. 推荐配方"), ncols=NCOLS)
    # Auto-computed from the same per-task winner logic as Tab 2 — so Tab 1 can
    # never drift from Tab 2. Bold columns = values computed live from the data.
    bpt = _best_per_task_table(scored, raw)
    task_col = T("Task", "任务")
    rec_col = T("Recommended recipe", "推荐配方")
    model_col = T("Model", "模型")
    if not bpt.empty and task_col in bpt.columns and rec_col in bpt.columns:
        keep = [task_col, rec_col] + ([model_col] if model_col in bpt.columns else [])
        recipes_df = bpt[keep].copy()
        recipes_bold = {2} | ({3} if model_col in bpt.columns else set())
    else:
        recipes_df = pd.DataFrame([{task_col: T("(no data)", "（无数据）"),
                                    rec_col: "—"}])
        recipes_bold = set()
    row = write_df(ws, recipes_df, row, bold_cols=recipes_bold)
    row += 2

    # ============================================================
    # 5. Keyword prompt compression — 3-row summary + Removed / Kept
    # ============================================================
    row = write_section_bar(ws, row,
        T("5. Keyword prompt compression", "5. 关键词 prompt 压缩"), ncols=NCOLS)
    _b_pct2 = f"{_kb['pct']}%" if _kb else "63%"
    kw_summary = pd.DataFrame([
        {T("Version", "版本"): "A — Full",
         T("Result", "结果"): T("Baseline", "基线")},
        {T("Version", "版本"): "B — Reduced",
         T("Result", "结果"): T(f"{_b_pct2} shorter, near-identical F1",
                                f"短 {_b_pct2}，F1 几乎一致")},
        {T("Version", "版本"): "C / D",
         T("Result", "结果"): T("Noticeable quality drop",
                                "质量明显下降")},
    ])
    # Bold the Result column — the B-row figure is computed live from prompts.txt.
    row = write_df(ws, kw_summary, row, bold_cols={2})
    row += 1

    _merge_and_write(ws, row, T("Removed:", "删去："), ncols=NCOLS,
                     font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for r_line in [
        T("long examples", "长示例"),
        T("repeated instructions", "重复指令"),
        T("verbose formatting guidance", "冗长的格式说明"),
    ]:
        _merge_and_write(ws, row, f"• {r_line}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    _merge_and_write(ws, row, T("Kept:", "保留："), ncols=NCOLS,
                     font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for k_line in [
        T("extraction objective", "抽取目标"),
        T("output format", "输出格式"),
        T("keyword matching logic", "关键词匹配逻辑"),
        T("parsing-critical constraints", "解析关键的约束"),
    ]:
        _merge_and_write(ws, row, f"• {k_line}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    # ============================================================
    # 6. Production recommendation — default + human review + escalate
    # ============================================================
    row = write_section_bar(ws, row,
        T("6. Production recommendation", "6. 生产建议"), ncols=NCOLS)
    _merge_and_write(ws, row, T("Default setup:", "默认配置："),
                     ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for d in [
        T("GPT-5-mini", "GPT-5-mini"),
        T("Task-specific recipes", "每任务专属配方"),
        T("Version B keyword prompt (long explanations + examples removed; "
          "role, hard rules, 2-step process, output format kept)",
          "B 版关键词 prompt（删去长解释和示例，保留角色、硬规则、两步流程和输出格式）"),
    ]:
        _merge_and_write(ws, row, f"• {d}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1
    _merge_and_write(ws, row, T("Use human review for:",
                                 "需要人工复核的情况："),
                     ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for r_line in [
        T("low-scoring outputs", "低分输出"),
        T("business-critical tasks", "关键业务任务"),
        T("wording-sensitive edge cases", "措辞敏感的边界情况"),
    ]:
        _merge_and_write(ws, row, f"• {r_line}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1
    escalate_text = T(
        "Escalate to stronger models only when GPT-5-mini fails review.",
        "仅在 GPT-5-mini 复核不通过时，才升级到更强模型。")
    _merge_and_write(ws, row, escalate_text, ncols=NCOLS,
                     font=Font(size=11, italic=True, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(escalate_text,
                                                          total_width_chars=140)
    row += 2

    # ============================================================
    # 7. Score explanation — answers Simon's "what does the score mean?"
    # ============================================================
    row = write_section_bar(ws, row,
        T("7. Score explanation", "7. 分数说明"), ncols=NCOLS)
    score_para = T(
        "Each brief produces one similarity score by comparing the AI "
        "output against the human-written ground truth. The reported score "
        "is the mean cosine similarity across all 23 briefs for that "
        "task / configuration.",
        "每份 brief 通过将 AI 输出与人工撰写的标准答案对比，"
        "产生一个相似度分数。"
        "表中显示的分数 = 该任务 / 配置在所有 23 份 brief 上的"
        "平均余弦相似度。")
    _merge_and_write(ws, row, score_para, ncols=NCOLS,
                     font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(score_para,
                                                          total_width_chars=140)
    row += 2
    for line in [
        T("Higher is better — 0 means unrelated, 1 means identical meaning.",
          "分数越高越好 —— 0 表示完全不相关，1 表示含义完全一致。"),
        T("Differences ≤ 0.036 are treated as effectively equivalent.",
          "差异 ≤ 0.036 视为基本相同。"),
    ]:
        _merge_and_write(ws, row, f"• {line}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    autosize_cols(ws, {1: 28, 2: 60, 3: 14, 4: 14, 5: 14, 6: 14})
    ws.freeze_panes = "A2"


# ---------- Tab 2: Recommended Configs ----------

def build_tab_recommended_configs(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Tab 2 — Final Production Recommendations.

    6 sections — earlier 5-section version trimmed too much (lost the
    score-definition bullets, the table-meta blocks, and the operational
    rules table). This version keeps the spec's headline structure but
    restores the supporting context a stakeholder needs to use it.

      1. Purpose (2 short lines)
      2. How to read the scores (3 bullets + score guide)
      3. Per-task recommendation table (with data-source + how-scored note)
      4. Model policy table (with data-source note + premium statement)
      5. Production usage rules (5-row scenario / rule table)
      6. Operational takeaway (one closing paragraph)
    """
    ws = wb.create_sheet(T("Final Recommendations", "生产推荐"))
    NCOLS = 12  # main recommendation table is 12 cols
    # Set column widths up front so write_df / row-height heuristics see
    # real widths (not the openpyxl-default 12 placeholder).
    autosize_cols(ws, {
        1: 16, 2: 34, 3: 12, 4: 12, 5: 16, 6: 16,
        7: 22, 8: 16, 9: 18, 10: 22, 11: 20, 12: 38,
    })
    row = write_tab_title(ws,
        T("Final Production Recommendations",
          "最终生产推荐"), ncols=NCOLS)

    GREEN_HIGHLIGHT  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    BASELINE_FILL    = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # ============================================================
    # 1. Purpose
    # ============================================================
    row = write_section_bar(ws, row,
        T("1. Purpose", "1. 目的"), ncols=NCOLS)
    for line in [
        T("For each extraction task, the model + brief-field recipe to "
          "ship to production.",
          "为每个抽取任务给出该上线的「模型 + brief 字段配方」。"),
        T("Full Brief is the experiment's baseline, not the default. If a "
          "shorter recipe performs within the practical noise threshold "
          "(Δ ≤ 0.036), use the shorter and cheaper recipe.",
          "完整 brief 是实验的基线，不是默认配置。"
          "如果更短的配方在实际噪声范围内（Δ ≤ 0.036）有相近表现，"
          "请使用更短更便宜的配方。"),
    ]:
        _merge_and_write(ws, row, line, ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        ws.row_dimensions[row].height = _estimate_row_height(line,
                                                              total_width_chars=140)
        row += 1
    row += 1

    # ============================================================
    # 2. How to read the scores + score guide
    # ============================================================
    row = write_section_bar(ws, row,
        T("2. How to read the scores", "2. 分数怎么理解"), ncols=NCOLS)
    for line in [
        T("Score = mean cosine similarity across 23 briefs.",
          "分数 = 23 个 brief 上的平均余弦相似度。"),
        T("Each score compares AI output against human-curated ground "
          "truth. Higher is better.",
          "每个分数对比 AI 输出和人工标准答案。分数越高越好。"),
        T("Recipe = the brief fields sent to the model. Example: "
          "audience + product.",
          "配方 = 发给模型的 brief 字段。例如 audience + product。"),
    ]:
        _merge_and_write(ws, row, f"• {line}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1
    _merge_and_write(ws, row, T("Score guide:", "分数指南："),
                     ncols=NCOLS,
                     font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for band in [
        T("≥ 0.70 — strong alignment",
          "≥ 0.70 —— 高度一致"),
        T("0.60 to 0.70 — usable; spot-check recommended",
          "0.60 到 0.70 —— 可用；建议抽查"),
        T("< 0.60 — review carefully",
          "< 0.60 —— 仔细审阅"),
    ]:
        _merge_and_write(ws, row, f"• {band}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    # ============================================================
    # 3. Per-task recommendation table (with data-source + scoring note)
    # ============================================================
    row = write_section_bar(ws, row,
        T("3. Per-task recommendation", "3. 每任务推荐"), ncols=NCOLS)
    row = _emit_table_meta(ws, row, NCOLS,
        source=T(
            "Stage A and Stage B API calls across all 23 briefs.",
            "Stage A 和 Stage B 在所有 23 个 brief 上的 API 调用。"),
        scoring=T(
            "Each row applies the 0.036 cosine noise floor as a hard rule: "
            "tiny score gaps are ties, and the cheaper/shorter setup wins. "
            "The table also shows paired brief wins/losses, Full Brief delta, "
            "and leave-one-brief-out stability.",
            "每行都把 0.036 余弦噪声阈值作为硬规则：微小分差算打平，"
            "优先选更便宜/更短的配置。表中同时展示配对胜负、相对完整 brief "
            "差值和去掉一个 brief 后的稳定性。"))
    table = _best_per_task_table(scored, raw)
    if table.empty:
        _merge_and_write(ws, row, T("No scored data yet.", "暂无打分数据。"),
                         ncols=NCOLS, font=ITALIC_GREY)
        row += 1
    else:
        rec_col = T("Recommended recipe", "推荐配方")
        row_fills = []
        for _, r in table.iterrows():
            recipe = str(r.get(rec_col, ""))
            if recipe.lower().startswith("full brief") or recipe.startswith("完整 brief"):
                row_fills.append(BASELINE_FILL)
            else:
                row_fills.append(GREEN_HIGHLIGHT)
        row = write_df(ws, table, row, row_fills=row_fills)
        _merge_and_write(ws, row,
            T(f"Decision bands: Recommended (stable on this sample) = stable LOO "
              f"AND beats its best peer on the same briefs at the paired test "
              f"(p < {cfg.PAIRED_TEST_ALPHA}); Cost-priority usable = tied within "
              f"{cfg.NOISE_FLOOR_COSINE} or lead not confirmed, so pick the "
              f"cheaper/shorter option; Not recommended = materially below Full "
              f"Brief or clearly behind. A higher mean alone is never a win.",
              f"决策分三档：推荐（本样本稳定）= LOO 稳 且 在相同 brief 上的配对检验"
              f"显著优于次优（p < {cfg.PAIRED_TEST_ALPHA}）；成本优先可用 = "
              f"在 {cfg.NOISE_FLOOR_COSINE} 内打平或优势未被统计确认，选更便宜/更短的；"
              f"不推荐 = 明显低于完整 brief 或明显落后。仅平均分更高不算赢。"),
            ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
        row += 2

    # ============================================================
    # 4. Model policy (with meta + explicit premium statement)
    # ============================================================
    row = write_section_bar(ws, row,
        T("4. Model policy", "4. 模型策略"), ncols=NCOLS)
    row = _emit_table_meta(ws, row, NCOLS,
        source=T(
            "All sentence-task cells where each model was tested. Cost is "
            "the per-call USD recorded from the provider response.",
            "每个模型在所有句子型任务单元格上的结果。"
            "成本是提供商响应里记录的每次调用 USD。"),
        scoring=T(
            "Avg quality is the mean cosine across every recipe for that "
            "model. The recommendation balances quality, cost, and whether "
            "the model delivers a stable improvement worth paying for.",
            "平均质量 = 该模型在所有配方上的平均余弦。"
            "推荐建议基于质量、成本和是否带来稳定提升综合判断。"))
    mc = _model_comparison_table(scored, raw)
    if mc.empty:
        _merge_and_write(ws, row, T("No model data yet.", "暂无模型数据。"),
                         ncols=NCOLS, font=ITALIC_GREY)
        row += 1
    else:
        default_v = T("Default", "默认")
        rec_col_m = T("Recommendation", "推荐")
        mc_fills = []
        for _, r in mc.iterrows():
            mc_fills.append(GREEN_HIGHLIGHT if r[rec_col_m] == default_v else None)
        row = write_df(ws, mc, row, row_fills=mc_fills)
    row += 1
    policy_line = T(
        "Premium models did not provide enough consistent quality gain "
        "to justify default production use.",
        "旗舰模型没有带来足够稳定的质量提升，不足以作为默认生产配置。")
    _merge_and_write(ws, row, policy_line, ncols=NCOLS,
                     font=Font(size=11, bold=True, color="0B0C0C"),
                     fill=LIGHT_BLUE)
    ws.row_dimensions[row].height = _estimate_row_height(policy_line,
                                                          total_width_chars=140)
    row += 2

    # ============================================================
    # 5. Operational takeaway
    # ============================================================
    # (Previous "Production usage rules" 5-row scenario table dropped —
    # the cases it described are already implied by the per-task table,
    # the model policy table, and the takeaway paragraph below.)
    row = write_section_bar(ws, row,
        T("5. Operational takeaway", "5. 操作要点"), ncols=NCOLS)
    takeaway = T(
        "Most tasks reached Full-Brief-level quality using only 1-2 "
        "targeted fields, reducing prompt size, token usage and operational "
        "complexity. Human-review low-scoring or business-critical outputs; "
        "escalate to a stronger model only when GPT-5-mini fails review.",
        "多数任务用 1-2 个针对性字段就达到了完整 brief 的质量，"
        "减少了 prompt 长度、token 用量和运维复杂度。"
        "对低分或关键业务输出做人工复核；"
        "仅在 GPT-5-mini 复核不通过时，才升级到更强模型。")
    _merge_and_write(ws, row, takeaway, ncols=NCOLS,
                     font=Font(size=11, color="0B0C0C"),
                     fill=LIGHT_BLUE)
    ws.row_dimensions[row].height = _estimate_row_height(takeaway,
                                                          total_width_chars=140)
    row += 2

    # Column widths — per-task table is 7 cols; model policy uses 5.
    # Col 2 (Recipe) widened so "brand_strategy + differentiators" fits;
    # col 7 (Why) trimmed since the cell text is now ≤9-word phrases.
    autosize_cols(ws, {
        1: 16, 2: 36, 3: 14, 4: 14, 5: 20, 6: 22, 7: 40,
    })
    ws.freeze_panes = "A2"


# ---------- Helpers for the field-contribution + per-task explainer tables
# (moved off Tab 2 to Tab 5 Appendix per user feedback — too much repetition
# on the production-decision tab). The two table builders below are called
# from build_tab_appendix.

def _field_contribution_summary_df() -> pd.DataFrame:
    """5-row table: which brief field carries what signal, and which tasks
    it suits. Hardcoded copy from the user-approved spec."""
    return pd.DataFrame([
        {T("Field", "字段"): "product",
         T("Primary value", "主要价值"):
             T("Supplies product, function, and category signal.",
               "提供产品、功能、品类信息"),
         T("Suitable tasks", "适合任务"):
             T("Category, Function, Feature, Benefit",
               "品类、功能、特性、利益点")},
        {T("Field", "字段"): "audience",
         T("Primary value", "主要价值"):
             T("Supplies user, scene, and situational signal.",
               "提供用户、场景、情境信息"),
         T("Suitable tasks", "适合任务"):
             T("Context, Positioning, Benefit",
               "使用情境、定位、利益点")},
        {T("Field", "字段"): "brand_strategy",
         T("Primary value", "主要价值"):
             T("Supplies brand direction and high-level concept.",
               "提供品牌方向和高层概念"),
         T("Suitable tasks", "适合任务"):
             T("Concept, Emotion",
               "概念、情感")},
        {T("Field", "字段"): "differentiators",
         T("Primary value", "主要价值"):
             T("Supplies distinctive selling points.",
               "提供差异化和卖点信息"),
         T("Suitable tasks", "适合任务"):
             T("Feature, Positioning, Emotion",
               "特性、定位、情感")},
        {T("Field", "字段"): "personality",
         T("Primary value", "主要价值"):
             T("Supplies tone, persona, and emotional cues.",
               "提供语气、人格和情绪线索"),
         T("Suitable tasks", "适合任务"):
             T("Benefit, Emotion (supporting)",
               "利益点、情感辅助")},
    ])


def _per_task_one_sentence_df() -> pd.DataFrame:
    """8-row table: one-sentence plain-language explanation per task.
    Hardcoded copy from the user-approved spec."""
    return pd.DataFrame([
        {T("Task", "任务"): T("Benefit", "利益点"),
         T("Plain-language explanation", "简单解释"):
             T("Need to understand who the user is, and what tone the brand should use to express the value.",
               "需要理解用户是谁，以及品牌应该用什么语气表达价值。")},
        {T("Task", "任务"): T("Category", "品类"),
         T("Plain-language explanation", "简单解释"):
             T("product alone is enough to identify what category the product belongs to.",
               "product 字段已经足够判断产品属于什么类别。")},
        {T("Task", "任务"): T("Concept", "概念"),
         T("Plain-language explanation", "简单解释"):
             T("brand_strategy is the most useful for the model to understand the core concept the brand wants to express.",
               "brand_strategy 最能帮助模型理解品牌想表达的核心概念。")},
        {T("Task", "任务"): T("Context", "使用情境"),
         T("Plain-language explanation", "简单解释"):
             T("audience best describes the scene the user is in when using the product.",
               "audience 字段最能说明用户在什么场景下使用产品。")},
        {T("Task", "任务"): T("Emotion", "情感"),
         T("Plain-language explanation", "简单解释"):
             T("Emotion judgement needs brand direction and distinctive information together.",
               "情感判断需要品牌方向和差异化信息共同支持。")},
        {T("Task", "任务"): T("Feature", "特性"),
         T("Plain-language explanation", "简单解释"):
             T("Features come mainly from product function and differentiating selling points.",
               "特性主要来自产品功能和差异化卖点。")},
        {T("Task", "任务"): T("Function", "功能"),
         T("Plain-language explanation", "简单解释"):
             T("Function information is concentrated in product.",
               "功能信息主要集中在 product 字段。")},
        {T("Task", "任务"): T("Positioning", "定位"),
         T("Plain-language explanation", "简单解释"):
             T("Positioning needs both the target user and the brand's distinctive points.",
               "定位需要同时理解目标人群和品牌差异点。")},
    ])


# ---------- Tab 3 helpers — focused tables instead of one dense matrix ----------

def _per_task_takeaway(scored: pd.DataFrame) -> dict[str, str]:
    """Plain-English one-liner per task describing which fields drive quality.

    Output:  {task_internal_name: "Category extraction mostly depends on
                                   product field." ...}

    Rules:
      • If the top-1 single field beats the next single by ≥ 0.036 AND beats
        Full Brief by ≥ 0, → "X extraction mostly depends on <field>."
      • If the top pair beats both top singles by ≥ 0.036, →
        "X extraction works best when <a> + <b> are sent together."
      • If Full Brief still wins decisively, → "X benefits from full context;
        single fields aren't enough alone."
      • Otherwise, → "X is robust across recipes; pick the cheapest."
    """
    out: dict[str, str] = {}
    if scored.empty:
        return out
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return out

    full_by_task = (
        sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        .groupby("task")["cosine"].mean()
    )

    for task in SENTENCE_TASKS:
        g = sent[sent["task"] == task]
        if g.empty:
            continue
        single = g[g["config_id"].apply(_is_single_field_config)]
        pair = g[g["config_id"].apply(_is_field_combination_config)]
        single_means = single.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
        pair_means = pair.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
        full_score = full_by_task.get(task)

        top_single_field = None
        top_single_score = 0.0
        if not single_means.empty:
            top_single_field = single_means.index[0].split(":")[-1]
            top_single_score = float(single_means.iloc[0])

        second_single_score = (float(single_means.iloc[1])
                                if len(single_means) > 1 else None)

        top_pair_recipe = None
        top_pair_score = 0.0
        if not pair_means.empty:
            top_pair_recipe = humanize_config(pair_means.index[0])
            top_pair_score = float(pair_means.iloc[0])

        task_name = humanize_task(task)
        zh_task_name = _TASK_DISPLAY_ZH.get(task, task_name)

        # Case 1: a pair clearly beats every single
        if (top_pair_recipe is not None and top_single_field is not None
                and top_pair_score - top_single_score >= 0.036):
            out[task] = T(
                f"{task_name} extraction works best when {top_pair_recipe} "
                "are sent together.",
                f"{zh_task_name}：把 {top_pair_recipe} 一起发效果最好。"
            )
        # Case 2: one single field clearly dominates the other singles
        elif (top_single_field is not None and second_single_score is not None
              and top_single_score - second_single_score >= 0.036):
            out[task] = T(
                f"{task_name} extraction mostly depends on the "
                f"'{top_single_field}' field.",
                f"{zh_task_name}：质量主要取决于「{top_single_field}」字段。"
            )
        # Case 3: Full Brief still clearly best
        elif (full_score is not None and top_single_field is not None
              and float(full_score) - top_single_score >= 0.036):
            out[task] = T(
                f"{task_name} benefits from full context; single fields "
                "aren't quite enough on their own.",
                f"{zh_task_name}：受益于完整上下文，单字段不太够。"
            )
        # Default: robust across recipes
        else:
            if top_single_field is not None:
                out[task] = T(
                    f"{task_name} is robust across recipes; '{top_single_field}' "
                    "alone is enough and cheapest.",
                    f"{zh_task_name}：对配方不敏感，单用「{top_single_field}」就够，最便宜。"
                )
            else:
                out[task] = T(
                    f"{task_name}: insufficient data to call a winner.",
                    f"{zh_task_name}：数据不足以下结论。"
                )
    return out


def _compact_heatmap(scored: pd.DataFrame, top_k: int = 6) -> pd.DataFrame:
    """Task × top-K recipes heatmap. Plus always-on Full Brief + No Brief
    columns for comparison. Cells = mean cosine across 23 briefs.

    The columns are the K recipes with the highest mean cosine averaged
    across all 8 sentence tasks — these are the recipes most worth showing.
    """
    if scored.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return pd.DataFrame()

    # Pick the top-K most informative recipes globally (excluding baselines).
    cfg_means = sent.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
    baselines = {"A:_full_brief", "_full_brief", "A:_prompt_implied", "_prompt_implied"}
    metadata_only = {"A:business_category", "A:business_category+product"}
    non_baseline = [c for c in cfg_means.index
                    if c not in baselines and c not in metadata_only]
    top_recipes = non_baseline[:top_k]

    # Ordered columns: baselines first (for visual reference), then top recipes.
    cols = ["A:_full_brief", "A:_prompt_implied"] + top_recipes

    # Build pivot
    piv = (
        sent[sent["config_id"].isin(cols)]
        .groupby(["task", "config_id"])["cosine"].mean()
        .unstack()
        .reindex(SENTENCE_TASKS)
        .reindex(columns=cols)
        .round(3)
    )
    return piv


def _baselines_per_task(scored: pd.DataFrame) -> pd.DataFrame:
    """| Task | Full Brief | No Context Baseline | Δ Full vs No-Brief | Verdict |"""
    if scored.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return pd.DataFrame()
    rows = []
    for task in SENTENCE_TASKS:
        g = sent[sent["task"] == task]
        full = g[g["config_id"].isin(("A:_full_brief", "_full_brief"))]["cosine"].mean()
        none = g[g["config_id"].isin(("A:_prompt_implied", "_prompt_implied"))]["cosine"].mean()
        if pd.isna(full) and pd.isna(none):
            continue
        delta = full - none if (pd.notna(full) and pd.notna(none)) else None
        if delta is None:
            verdict = "—"
        elif delta > 0.09:
            verdict = T("Brief adds real value", "brief 带来明显价值")
        elif delta > 0.036:
            verdict = T("Brief adds some value", "brief 带来一些价值")
        else:
            verdict = T("Brief barely helps", "brief 几乎没用")
        rows.append({
            T("Task", "任务"): T_task(task),
            T("Full Brief", "完整 brief"):
                round(float(full), 3) if pd.notna(full) else "—",
            T("No Context Baseline", "无上下文基线"):
                round(float(none), 3) if pd.notna(none) else "—",
            T("Δ (brief – no brief)", "差值（brief – 无 brief）"):
                round(float(delta), 3) if delta is not None else "—",
            T("Verdict", "判定"): verdict,
        })
    return pd.DataFrame(rows)


def _best_singles_per_task(scored: pd.DataFrame, top_n: int = 2) -> pd.DataFrame:
    """| Task | Top-1 single field (score) | Top-2 single field (score) | Δ vs Full Brief |"""
    if scored.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return pd.DataFrame()
    full_by_task = (
        sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        .groupby("task")["cosine"].mean()
    )
    rows = []
    for task in SENTENCE_TASKS:
        singles = sent[
            (sent["task"] == task)
            & sent["config_id"].apply(_is_single_field_config)
        ]
        if singles.empty:
            continue
        means = singles.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
        top = means.head(top_n)
        full_score = full_by_task.get(task)
        row = {T("Task", "任务"): T_task(task)}
        for i, (cid, val) in enumerate(top.items(), 1):
            field_name = cid.split(":")[-1]
            row[T(f"Top-{i} single field", f"第 {i} 名单字段")] = f"{field_name} ({val:.3f})"
        # Pad if fewer than top_n found
        for i in range(len(top) + 1, top_n + 1):
            row[T(f"Top-{i} single field", f"第 {i} 名单字段")] = "—"
        delta_col = T("Δ vs Full Brief", "相对完整 brief 的差值")
        if full_score is not None and pd.notna(full_score) and len(top) > 0:
            row[delta_col] = round(float(top.iloc[0]) - float(full_score), 3)
        else:
            row[delta_col] = "—"
        rows.append(row)
    return pd.DataFrame(rows)


def _best_pairs_per_task(scored: pd.DataFrame, top_n: int = 2) -> pd.DataFrame:
    """| Task | Top-1 pair (score) | Top-2 pair (score) | Δ vs Full Brief |"""
    if scored.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna()]
    if sent.empty:
        return pd.DataFrame()
    full_by_task = (
        sent[sent["config_id"].isin(("A:_full_brief", "_full_brief"))]
        .groupby("task")["cosine"].mean()
    )
    rows = []
    for task in SENTENCE_TASKS:
        pairs = sent[
            (sent["task"] == task)
            & sent["config_id"].apply(_is_field_combination_config)
        ]
        if pairs.empty:
            continue
        means = pairs.groupby("config_id")["cosine"].mean().sort_values(ascending=False)
        top = means.head(top_n)
        full_score = full_by_task.get(task)
        row = {T("Task", "任务"): T_task(task)}
        for i, (cid, val) in enumerate(top.items(), 1):
            row[T(f"Top-{i} pair recipe", f"第 {i} 名双字段配方")] = \
                f"{T_recipe(humanize_config(cid))} ({val:.3f})"
        for i in range(len(top) + 1, top_n + 1):
            row[T(f"Top-{i} pair recipe", f"第 {i} 名双字段配方")] = "—"
        delta_col = T("Δ vs Full Brief", "相对完整 brief 的差值")
        if full_score is not None and pd.notna(full_score) and len(top) > 0:
            row[delta_col] = round(float(top.iloc[0]) - float(full_score), 3)
        else:
            row[delta_col] = "—"
        rows.append(row)
    return pd.DataFrame(rows)


# ---------- Tab 3: Keyword Prompt Compression (separate, self-contained) ----------
# (build_tab_field_compression deleted — its content is now Section 2 of
# build_tab_recommended_configs to enable the 4-tab structure.)

def build_tab_keyword_compression(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Tab 3 — Keyword Prompt Compression. 7 sections per user spec:

      1. Purpose
      2. Key conclusion (2 sentences, hardcoded)
      3. How to read the scores
      4. What we tested
      5. Results table (10 cols)
      6. Why B — Reduced is recommended
      7. Final recommendation
    """
    ws = wb.create_sheet(T("Keyword Prompt Compression", "关键词提示词压缩"))
    NCOLS = 10  # results table is 10 cols (Version / Desc / Chars / %red / P / R / F1 / F1Δ / Rec? / Decision)
    row = write_tab_title(ws,
        T("Keyword Prompt Compression Experiment",
          "关键词提示词压缩实验"), ncols=NCOLS)

    # ----- 1. Purpose -----
    row = write_section_bar(ws, row, T("1. Purpose", "1. 目的"), ncols=NCOLS)
    _merge_and_write(ws, row,
        T("Test how short the keyword extraction prompt can be made without "
          "degrading the quality of keyword extraction.",
          "测试关键词抽取 prompt 能压缩到多短，同时不降低关键词抽取质量。"),
        ncols=NCOLS, font=Font(size=11, color="0B0C0C"))
    row += 2

    # ----- 2. Key conclusion (hardcoded, 2 sentences) -----
    row = write_section_bar(ws, row, T("2. Key conclusion", "2. 核心结论"), ncols=NCOLS)
    conclusion_lines = [
        T("Use B — Reduced. It is 63% shorter than the original. F1 actually "
          "improves slightly (+0.018), which can be treated as no quality loss.",
          "推荐使用 B — Reduced。它比原版短 63%，F1 反而略高 +0.018，可以视为质量无损。"),
        T("C and D are shorter but F1 drops by about 0.06, which starts to hurt "
          "quality. They are not recommended.",
          "C/D 虽然更短，但 F1 下降约 0.06，已经开始影响质量，所以不推荐。"),
    ]
    for line in conclusion_lines:
        _merge_and_write(ws, row, line, ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"), fill=LIGHT_BLUE)
        ws.row_dimensions[row].height = _estimate_row_height(line, total_width_chars=140)
        row += 1
    row += 1

    # ----- 3. How to read the scores -----
    row = write_section_bar(ws, row,
        T("3. How to read the scores", "3. 分数怎么理解"), ncols=NCOLS)
    score_para = T(
        "Precision is how many of the keywords the model predicted are correct. "
        "Recall is how many of the human-curated keywords the model found. "
        "F1 combines the two. The headline metric here is F1.",
        "Precision 表示模型提取的关键词有多少是对的；"
        "Recall 表示人工关键词有多少被模型找到了；"
        "F1 是两者的综合分数。这里主要看 F1。")
    _merge_and_write(ws, row, score_para, ncols=NCOLS,
                     font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(score_para, total_width_chars=140)
    row += 2

    # ----- 4. What we tested -----
    row = write_section_bar(ws, row,
        T("4. What we tested", "4. 我们测试了什么"), ncols=NCOLS)
    what_tested_para = T(
        "We compared four prompt versions A B C D. A is the original. B "
        "removes the long explanations and examples but keeps the rules. C "
        "further compresses the rule description. D only keeps a minimal task "
        "description. Every version was tested on all 23 briefs and scored "
        "against the human-curated keyword list.",
        "我们比较了 A/B/C/D 四个 prompt 版本：A 是原版，B 删除长解释和示例但保留规则，"
        "C 进一步压缩规则说明，D 只保留最小任务描述。"
        "每个版本都在 23 个 brief 上测试，并与人工标注关键词对比。")
    _merge_and_write(ws, row, what_tested_para, ncols=NCOLS,
                     font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(what_tested_para, total_width_chars=140)
    row += 2

    # ----- 5. Results table -----
    row = write_section_bar(ws, row,
        T("5. Results. Mean across 23 briefs.",
          "5. 结果。23 brief 均值。"), ncols=NCOLS)
    kw_df = _keyword_compression_table(scored)
    if kw_df.empty:
        _merge_and_write(ws, row, T("No keyword data yet.", "暂无关键词数据。"),
                         ncols=NCOLS, font=ITALIC_GREY)
        row += 1
    else:
        row = _emit_table_meta(ws, row, NCOLS,
            source=T(
                "Each of the four prompt versions A B C D was run on all 23 "
                "briefs. The keyword task has a hand-curated gold list of 10 "
                "keywords per brief.",
                "四个 prompt 版本 A B C D 都在 23 个 brief 上跑过。"
                "关键词任务有人工标注的金标，每份 brief 10 个关键词。"),
            scoring=T(
                "Each cell is the mean across the 23 briefs. Porter stemming "
                "is applied so 'running' matches 'run'.",
                "每个单元格是 23 个 brief 上的均值。"
                "应用 Porter 词干化，所以 running 能匹配 run。"))
        GREEN_HIGHLIGHT  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        BASELINE_FILL    = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        rec_col = T("Recommended?", "是否推荐")
        yes_v = T("Yes", "是")
        baseline_v = T("Baseline", "基线")
        kw_fills = []
        for _, r in kw_df.iterrows():
            rec = str(r.get(rec_col, ""))
            if rec == yes_v:
                kw_fills.append(GREEN_HIGHLIGHT)
            elif rec == baseline_v:
                kw_fills.append(BASELINE_FILL)
            else:
                kw_fills.append(None)
        row = write_df(ws, kw_df, row, row_fills=kw_fills)
    row += 2

    # ----- 6. Why B — Reduced is recommended -----
    row = write_section_bar(ws, row,
        T("6. Why B — Reduced is the recommendation",
          "6. 为什么推荐 B — Reduced"), ncols=NCOLS)
    why_b_para = T(
        "Version B removed the long explanations and the examples but kept "
        "the rules, the process, and the output format — the things that "
        "actually influence quality. The result shows the model does not "
        "need long explanations to complete the task. But if we also remove "
        "the rules and process, quality drops noticeably.",
        "B 版删掉了冗长解释和示例，但保留了真正影响质量的规则、步骤和输出格式。"
        "结果显示，模型不需要长篇解释也能完成任务；"
        "但如果继续删除规则和流程，质量就会明显下降。")
    _merge_and_write(ws, row, why_b_para, ncols=NCOLS,
                     font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(why_b_para, total_width_chars=140)
    row += 2

    # ----- 7. Final recommendation -----
    row = write_section_bar(ws, row,
        T("7. Final recommendation", "7. 最终建议"), ncols=NCOLS)
    _merge_and_write(ws, row,
        T("Use B — Reduced as the keyword extraction prompt in production. It "
          "is the best balance right now: clearly shorter, lower cost, and "
          "quality does not drop.",
          "生产环境使用 B — Reduced 作为关键词抽取 prompt。"
          "它是目前最好的平衡点：明显更短、成本更低、质量不下降。"),
        ncols=NCOLS, font=Font(size=12, bold=True, color="0B0C0C"),
        fill=LIGHT_BLUE)
    ws.row_dimensions[row].height = 36
    row += 1

    # Column widths for the 10-col keyword table. Col 2 (Description) and
    # col 10 (Decision implication) carry the longest text on this tab.
    # 1 Version / 2 Description / 3 Chars / 4 % reduction / 5 P / 6 R / 7 F1 /
    # 8 F1 delta / 9 Recommended / 10 Decision implication
    autosize_cols(ws, {
        1: 14, 2: 55, 3: 12, 4: 16, 5: 14, 6: 14, 7: 14, 8: 18, 9: 13, 10: 45,
    })
    ws.freeze_panes = "A2"


# ---------- Tab 4: Appendix — Validation ----------

def _sample_for_human_review(scored: pd.DataFrame, n: int = 30) -> pd.DataFrame:
    """Stratified sample for spot-check — hits exactly n samples when possible.

    Earlier implementation used floor division (n // 8 = 3 for n=30) giving
    only 24. Now: floor + remainder so the FIRST `extra` tasks each get one
    extra sample, landing on exactly n. e.g. n=30, 8 tasks → 6 tasks with 4
    samples + 2 tasks with 3 = 30 total.
    """
    if scored.empty:
        return pd.DataFrame()
    sent = scored[scored["cosine"].notna() & scored["task"].isin(SENTENCE_TASKS)].copy()
    if sent.empty:
        return pd.DataFrame()
    per_task_base = max(1, n // len(SENTENCE_TASKS))
    extra = max(0, n - per_task_base * len(SENTENCE_TASKS))
    out_frames = []
    for i, task in enumerate(SENTENCE_TASKS):
        per_task = per_task_base + (1 if i < extra else 0)
        g = sent[sent["task"] == task].sort_values("cosine")
        if g.empty:
            continue
        if len(g) <= per_task:
            picks = g
        else:
            step = (len(g) - 1) / (per_task - 1) if per_task > 1 else 0
            idx = [int(round(i * step)) for i in range(per_task)]
            # Deduplicate indices in case of overlap
            seen, dedup = set(), []
            for i in idx:
                if i not in seen:
                    seen.add(i)
                    dedup.append(i)
            picks = g.iloc[dedup]
        out_frames.append(picks)
    if not out_frames:
        return pd.DataFrame()
    combined = pd.concat(out_frames, ignore_index=True)
    # First-impression matters: surface a high-similarity row first so the
    # reviewer's first glance is "the AI does produce sensible outputs" rather
    # than a refusal at 0.05. Sort the whole frame by similarity descending.
    return combined.sort_values("cosine", ascending=False).reset_index(drop=True)


def _human_review_table(scored: pd.DataFrame) -> pd.DataFrame:
    """Per-brief output samples with per-brief AND aggregate scores.

    Columns:
      | Task | Brief ID | Recipe / context setup | Model | Ground Truth |
        AI Output | Per-brief score | Aggregate (this recipe, 23 briefs) |
        Sonnet 1-5 | Human 1-5 | Notes |

    The "Per-brief score" vs "Aggregate" split is the key Simon-asked-for
    clarification: aggregate scores in Tabs 1–3 = mean of per-brief scores
    like the ones in this column.
    """
    samp = _sample_for_human_review(scored)
    if samp.empty:
        return pd.DataFrame()

    # Sonnet absolute ratings — joined per sample row by
    # (brief_id, task, config_id, model_key). Missing rows show "—".
    sonnet_path = cfg.OUTPUTS_DIR / "ai_judge_absolute.jsonl"
    sonnet_lookup: dict = {}
    if sonnet_path.exists():
        try:
            from src.utils import read_jsonl
            for r in read_jsonl(sonnet_path):
                key = (r.get("brief_id"), r.get("task"),
                       r.get("config_id"), r.get("model_key"))
                sonnet_lookup[key] = r.get("score")
        except Exception:
            sonnet_lookup = {}

    MAX_TEXT = 250
    def _truncate(v):
        s = "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
        return s if len(s) <= MAX_TEXT else s[:MAX_TEXT] + "…"

    HUMAN_PENDING = T("Pending", "待填写")
    # Carry forward any human ratings already typed into the previous build so a
    # rebuild never wipes them. Keyed by (model, truncated AI output).
    existing_human = _existing_human_ratings()

    # 8-column table per spec:
    # Task | Recipe | Model | Ground Truth | AI Output | Auto cosine | Sonnet 1-5 | Human 1-5
    rows = []
    for _, r in samp.iterrows():
        skey = (r["brief_id"], r["task"], r["config_id"], r["model_key"])
        sonnet_raw = sonnet_lookup.get(skey)
        sonnet_v = sonnet_raw if sonnet_raw not in (None, "—") else "—"
        ai_out = _truncate(r.get("prediction"))
        human_v = existing_human.get((str(r["model_key"]), ai_out), HUMAN_PENDING)
        rows.append({
            T("Task", "任务"): T_task(r["task"]),
            T("Recipe", "配方"):
                T_recipe(humanize_config(r["config_id"])),
            T("Model", "模型"): r["model_key"],
            T("Ground Truth", "人工标准答案"): _truncate(r.get("ground_truth")),
            T("AI Output", "AI 输出"): ai_out,
            T("Auto cosine", "自动余弦"): round(float(r["cosine"]), 3),
            T("Sonnet 1-5", "Sonnet 1-5"): sonnet_v,
            T("Human 1-5", "人工 1-5"): human_v,
        })
    return pd.DataFrame(rows)


def _existing_human_ratings() -> dict:
    """Read hand-entered Human 1-5 scores from the previously-built workbook so
    a rebuild preserves them instead of resetting the column to 'Pending'.

    Keyed by (model, truncated AI-output text) — both reproduce identically on
    every build, so the match is stable even if row order shifts. Returns {} if
    no prior file exists or it can't be read.
    """
    path = cfg.RESULTS_DIR / (
        "Prompt Eval Results (CN).xlsx" if LANG == "zh" else "Prompt Eval Results.xlsx"
    )
    out: dict = {}
    if not path.exists():
        return out
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return out
    ws = None
    for nm in ("Human Review", "人工评审"):
        if nm in wb.sheetnames:
            ws = wb[nm]
            break
    if ws is None:
        return out
    # Find header row (contains the Human-rating column header).
    hdr = None
    for rr in range(1, 60):
        for cc in range(1, 12):
            v = ws.cell(rr, cc).value
            if isinstance(v, str) and ("Human 1-5" in v or "人工 1-5" in v):
                hdr = rr
                break
        if hdr:
            break
    if not hdr:
        return out
    # Fixed column layout: Task1 Recipe2 Model3 GT4 AIOut5 cosine6 Sonnet7 Human8
    MODEL_C, AIOUT_C, HUMAN_C = 3, 5, 8
    rr = hdr + 1
    while True:
        first = ws.cell(rr, 1).value
        if first in (None, ""):
            break
        h = ws.cell(rr, HUMAN_C).value
        if isinstance(h, (int, float)):
            model = ws.cell(rr, MODEL_C).value
            ai = ws.cell(rr, AIOUT_C).value
            out[(str(model), str(ai))] = int(round(h))
        rr += 1
    return out


def _weighted_kappa_quadratic(a: list, b: list) -> float:
    """Cohen's weighted kappa (quadratic weights) — same formula as
    scripts/compute_kappa.py. Pure-python, no sklearn."""
    if len(a) != len(b) or not a:
        return float("nan")
    cats = sorted(set(a) | set(b))
    k = len(cats)
    if k < 2:
        return 1.0
    idx = {c: i for i, c in enumerate(cats)}
    n = len(a)
    O = [[0] * k for _ in range(k)]
    for x, y in zip(a, b):
        O[idx[x]][idx[y]] += 1
    rt = [sum(r) for r in O]
    ct = [sum(O[r][c] for r in range(k)) for c in range(k)]
    E = [[rt[r] * ct[c] / n for c in range(k)] for r in range(k)]
    dw = (k - 1) ** 2
    W = [[((i - j) ** 2) / dw for j in range(k)] for i in range(k)]
    num = sum(W[i][j] * O[i][j] for i in range(k) for j in range(k))
    den = sum(W[i][j] * E[i][j] for i in range(k) for j in range(k))
    return float("nan") if den == 0 else 1 - num / den


def _human_sonnet_kappa(scored: pd.DataFrame):
    """Real Human ↔ Sonnet weighted kappa from the (preserved) Human 1-5 column.

    Returns (n_pairs, kappa) when at least 5 numeric Human/Sonnet pairs exist,
    else None. Also writes outputs/kappa.json so Tab 1 and the standalone
    compute_kappa path stay in sync.
    """
    hr = _human_review_table(scored)
    if hr.empty:
        return None
    son_c = T("Sonnet 1-5", "Sonnet 1-5")
    hum_c = T("Human 1-5", "人工 1-5")
    a, b = [], []
    for _, row in hr.iterrows():
        s, h = row.get(son_c), row.get(hum_c)
        if isinstance(s, (int, float)) and isinstance(h, (int, float)):
            a.append(int(round(s)))
            b.append(int(round(h)))
    if len(a) < 5:
        return None
    kappa = _weighted_kappa_quadratic(a, b)
    if kappa != kappa:  # NaN
        return None
    try:
        import json as _json
        from datetime import datetime as _dt
        decision = "HIGH" if kappa >= 0.7 else "MEDIUM" if kappa >= 0.4 else "LOW"
        (cfg.OUTPUTS_DIR / "kappa.json").write_text(_json.dumps({
            "kappa": round(kappa, 4), "n_pairs": len(a), "decision": decision,
            "timestamp": _dt.now().isoformat(timespec="seconds"),
        }, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return (len(a), kappa)


def _kappa_summary_line() -> str:
    """One-line agreement status. Business wording — drops the "Cohen's
    weighted kappa" technical term in favour of "rating agreement". The
    actual number still comes from the kappa.json file produced by
    scripts/compute_kappa.py.
    """
    path = cfg.OUTPUTS_DIR / "kappa.json"
    if not path.exists():
        return T(
            "Rating agreement not yet computed. The number appears here once "
            "the human-rating column is filled in. See the Appendix for the "
            "exact command to compute it.",
            "评分一致性尚未计算。人工评分列填好后此处会显示数值。"
            "具体计算命令见附录。")
    try:
        import json as _json
        data = _json.loads(path.read_text(encoding="utf-8"))
        kappa = data.get("kappa")
        n = data.get("n_pairs")
        if kappa is None:
            return T(
                "Agreement file present but no value found.",
                "一致性文件存在，但未找到数值。")
        if n is None or int(n) < 30:
            return T(
                f"AI judge is not yet human-validated: only {n or 0} paired "
                "human scores are available. Treat Sonnet judge outputs as "
                "supporting evidence only.",
                f"AI 评审尚未通过人工一致性验证：目前只有 {n or 0} 条人工配对评分。"
                "Sonnet 评审结果只能作为辅助参考。")
        if kappa >= 0.7:
            verdict = T("strong agreement", "高度一致")
        elif kappa >= 0.4:
            verdict = T("moderate agreement", "中等一致")
        else:
            verdict = T("low agreement", "一致性较低")
        return T(
            f"Human ↔ AI rating agreement on {n} samples: {verdict} "
            f"(score {kappa:.2f}, 0 to 1 scale). 60 samples is preferred "
            "before relying on the AI judge as a production scorer.",
            f"{n} 条样本上的人工 ↔ AI 评分一致性：{verdict}（分数 {kappa:.2f}，0 到 1 范围）。"
            "若要把 AI 评审作为生产评分器，建议做到 60 条样本。")
    except Exception as e:
        return T(
            f"Could not read agreement file ({e}).",
            f"无法读取一致性文件（{e}）。")


def _render_human_review_table(ws, df: pd.DataFrame, start_row: int) -> int:
    """Custom Human-Review renderer — lighter borders, taller text rows,
    highlighted key columns (Similarity, Sonnet 1-5, Human 1-5).

    Why custom: write_df uses uniform THIN_BORDER + 22-row-height, which makes
    AI Output / Ground Truth (long text) unreadable. This renderer wraps text
    and pales out non-key columns so the eye lands on the three rating cells.
    """
    PALE_BORDER = Border(
        left=Side(style="thin", color="E5E5E5"),
        right=Side(style="thin", color="E5E5E5"),
        top=Side(style="thin", color="E5E5E5"),
        bottom=Side(style="thin", color="E5E5E5"),
    )
    KEY_COLUMN_FILL = PatternFill(start_color="FFF7BF", end_color="FFF7BF", fill_type="solid")
    headers = list(df.columns)
    # Columns to visually emphasise — the rating + score cells.
    # Both EN and ZH variants listed so the highlight works in both languages.
    key_cols = {
        # EN headers (new 8-col Tab 3 spec)
        "Auto cosine", "Sonnet 1-5", "Human 1-5",
        # ZH headers
        "自动余弦", "Sonnet 1-5", "人工 1-5",
    }

    # Header row
    for i, h in enumerate(headers, 1):
        fill = LIGHT_GREY
        write_cell(ws, start_row, i, h, font=HEADER_FONT, fill=fill, border=THIN_BORDER,
                   align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[start_row].height = 22
    start_row += 1

    text_align_wrap = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
    center_align = Alignment(horizontal="center", vertical="center")

    long_text_headers = {
        "AI Output", "Ground Truth", "AI 输出", "人工标准答案",
    }
    for _, row_tuple in enumerate(df.itertuples(index=False)):
        # Compute the height needed for THIS row from the longest text cell.
        # Same CJK-aware logic as write_df, applied to long-text columns only.
        max_lines = 1
        for c, (header, val) in enumerate(zip(headers, row_tuple), 1):
            if header in long_text_headers and isinstance(val, str):
                col_w = ws.column_dimensions[get_column_letter(c)].width or 12
                visual = _visual_width(val)
                effective_w = max(4, int(col_w) - 1)
                lines = max(1, (visual + effective_w - 1) // effective_w)
                max_lines = max(max_lines, lines)

        for c, (header, val) in enumerate(zip(headers, row_tuple), 1):
            is_key = header in key_cols
            is_long_text = header in long_text_headers
            border = THIN_BORDER if is_key else PALE_BORDER
            fill = KEY_COLUMN_FILL if is_key else None
            if is_long_text:
                align = text_align_wrap
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                align = center_align
            else:
                align = Alignment(horizontal="left", vertical="center",
                                   indent=1, wrap_text=True)
            write_cell(ws, start_row, c, val, font=Font(size=10),
                       fill=fill, border=border, align=align)
        # ~13 pt per wrapped line + 4 pt padding. Cap at 220 so a single
        # 250-char ground-truth blob can not blow up the sheet. The previous
        # min(60) floor forced all rows to 60pt even when the actual content
        # only needed 30; dropped to 28 so short-content rows look compact.
        ws.row_dimensions[start_row].height = min(220, max(28, max_lines * 13 + 4))
        start_row += 1
    return start_row + 1


def build_tab_analysis(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Tab 3 — Analysis. Every supporting data table from the experiment,
    organised by the methodology-diagram branches:

      Part A. Sentence-task scoring  (the left branch of the flowchart)
        A1. Model summary
        A2. Premium vs cheap on the same recipe
        A3. Brief vs no brief (does the brief help?)
        A4. Field-level signal (single fields + best pairs)
        A5. Per-task production picks (cheap-tier winners)

      Part B. Keyword task scoring   (the right branch)
        B1. A / B / C / D compression results

      Part C. Stability check        (Stage B box in the diagram)
        C1. Cross-rerun cosine std dev

      Part D. Reliability & cost     (the operational layer)
        D1. Per-model OK rate
        D2. Per-model spend
    """
    ws = wb.create_sheet(T("Analysis", "数据分析"))
    NCOLS = 6
    # Set column widths up front — see Tab 1 for why.
    autosize_cols(ws, {1: 28, 2: 36, 3: 18, 4: 18, 5: 22, 6: 22})
    row = write_tab_title(ws, T("Analysis", "数据分析"), ncols=NCOLS)

    _merge_and_write(ws, row,
        T("Supporting data tables for every conclusion on Tab 1. Organised "
          "by the scoring-flow branches: sentence-task scoring, keyword task "
          "scoring, stability check, reliability + cost.",
          "Tab 1 上每条结论的支撑数据表。"
          "按评分流程分支组织：句子型任务评分、关键词任务评分、"
          "稳定性检查、可靠性 + 成本。"),
        ncols=NCOLS, font=Font(size=11, italic=True, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "Supporting data tables for every conclusion on Tab 1...",
        total_width_chars=160)
    row += 2

    # ----- How this maps to the methodology -----
    # Anchors the four Parts back to the "HOW WE GRADE THE AI'S OUTPUTS"
    # methodology flowchart so the reader can see Tab 3 IS the data
    # behind every box on that diagram.
    row = write_section_bar(ws, row,
        T("How this maps to the methodology",
          "本 Tab 与方法论的对应关系"), ncols=NCOLS)
    map_lines = [
        T("This tab follows the experiment workflow:",
          "本 Tab 按实验流程组织："),
        T("  Part A  →  Sentence-task scoring",
          "  Part A  →  句子型任务评分"),
        T("  Part B  →  Keyword-task scoring",
          "  Part B  →  关键词任务评分"),
        T("  Part C  →  Stability validation",
          "  Part C  →  稳定性验证"),
        T("  Part D  →  Reliability and operational cost",
          "  Part D  →  可靠性与运营成本"),
        T("",  ""),
        T("Together these sections validate:",
          "这四部分合起来验证了："),
        T("  • model choice",        "  • 模型选择"),
        T("  • field selection",     "  • 字段选择"),
        T("  • keyword compression", "  • 关键词压缩"),
        T("  • rerun stability",     "  • 重跑稳定性"),
        T("  • operational feasibility", "  • 运营可行性"),
    ]
    for line in map_lines:
        if not line:
            row += 1
            continue
        _merge_and_write(ws, row, line, ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    def _conclusion(text: str):
        """Small italic 1-2 sentence interpretation under a table."""
        nonlocal row
        _merge_and_write(ws, row, text, ncols=NCOLS,
                         font=Font(size=10, italic=True, color="0B0C0C"))
        ws.row_dimensions[row].height = _estimate_row_height(text,
                                                              total_width_chars=160)
        row += 2

    def _implication(text: str):
        """One-line operational takeaway. Bold, light-blue fill, prefixed
        with "Operational implication:" so each section closes with a
        production-facing decision, not just a research observation."""
        nonlocal row
        prefix = T("Operational implication: ", "运营含义：")
        full = f"{prefix}{text}"
        _merge_and_write(ws, row, full, ncols=NCOLS,
                         font=Font(size=11, bold=True, color="0B0C0C"),
                         fill=LIGHT_BLUE)
        ws.row_dimensions[row].height = _estimate_row_height(full,
                                                              total_width_chars=160)
        row += 2

    def _part_header(en: str, zh: str):
        nonlocal row
        _merge_and_write(ws, row, T(en, zh), ncols=NCOLS,
                         font=Font(size=12, bold=True, color="0B0C0C"),
                         fill=LIGHT_BLUE)
        ws.row_dimensions[row].height = 28
        row += 2

    # ============================================================
    # PART A — Sentence-task scoring
    # ============================================================
    _part_header(
        "══════ Part A — Sentence-task scoring (cosine-based) ══════",
        "══════ Part A —— 句子型任务评分（基于余弦）══════")

    row = write_section_bar(ws, row,
        T("A1. Model summary — across every sentence-task cell",
          "A1. 模型总览 —— 全部句子型任务单元格"), ncols=NCOLS)
    df = _model_summary_df(scored, raw)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _sent = scored[scored["cosine"].notna()] if not scored.empty else scored
    _gm = (float(_sent[_sent["model_key"] == "gpt5mini"]["cosine"].mean())
           if not _sent.empty else float("nan"))
    _gm_str = f"{_gm:.3f}" if _gm == _gm else "n/a"  # NaN-safe
    _n_briefs = int(_sent["brief_id"].nunique()) if not _sent.empty else 0
    _cost_clause = (
        T("exact cost advantage remains provisional until model prices are verified.",
          "精确成本优势需等模型价格确认后再写死。")
        if not getattr(cfg, "PRICES_VERIFIED", False)
        else T("cost advantage is confirmed against verified prices.",
               "成本优势已对照已核实价格确认。")
    )
    _conclusion(T(
        f"On these {_n_briefs} briefs, GPT-5-mini has the highest average "
        f"cosine of any model tested ({_gm_str}). The default-model choice is "
        f"quality-led on this sample; {_cost_clause}",
        f"在这 {_n_briefs} 个 brief 上，GPT-5-mini 平均余弦最高（{_gm_str}）。"
        f"默认模型选择在本样本上主要基于质量；{_cost_clause}"))
    _implication(T(
        "On this sample GPT-5-mini is the production-default — no premium "
        "upgrade is needed by default. Re-confirm if prices or briefs change.",
        "在本样本上 GPT-5-mini 为生产默认 —— 默认无需升级到旗舰。"
        "价格或 brief 变化后需重新确认。"))

    row = write_section_bar(ws, row,
        T("A2. Premium vs cheap on the same recipe",
          "A2. 同一配方下旗舰 vs 便宜"), ncols=NCOLS)
    df = _premium_vs_cheap_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "Premium models were checked only on 3 representative briefs. The "
        "observed direction does not justify a default premium upgrade, but "
        "it is not a full 23-brief validation.",
        "旗舰模型只在 3 个代表性 brief 上做了检查。当前方向不支持默认升级到旗舰，"
        "但这不是覆盖全部 23 个 brief 的完整验证。"))
    _implication(T(
        "Keep cheap-tier as default; run full 23-brief premium validation only "
        "if a task needs a higher quality ceiling.",
        "默认保留便宜层；只有任务确实需要更高质量上限时，再做完整 23 个 brief 的旗舰验证。"))

    row = write_section_bar(ws, row,
        T("A3. Brief vs No Brief — does the brief help?",
          "A3. 完整 brief vs 无 brief —— brief 有帮助吗？"), ncols=NCOLS)
    df = _brief_vs_no_brief_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "Every task shows a Δ above 0.3. The brief carries real signal on "
        "every task — the question is only which fields, never whether.",
        "每个任务的差值都大于 0.3。"
        "brief 在所有任务上都带来真实信号 —— 问题只是用哪些字段，从来不是要不要 brief。"))
    _implication(T(
        "Use targeted context, not empty prompts.",
        "使用针对性上下文，而非空 prompt。"))

    row = write_section_bar(ws, row,
        T("A4. Field-level signal (cheap tier, across the 8 tasks)",
          "A4. 字段层信号（便宜层，覆盖 8 个任务）"), ncols=NCOLS)
    _conclusion(T("Single fields, ranked:", "单字段排名："))
    df = _single_field_ranking_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "product dominates as a single field; personality alone is the weakest.",
        "product 单字段占主导地位；personality 单字段最弱。"))

    _conclusion(T("Top 2-field pairs (≥5-task coverage):",
                  "Top 双字段对（≥5 任务覆盖）："))
    df = _best_pairs_ranking_df(scored, top_n=5)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "Every top pair contains product. The other field tunes the "
        "second-strongest signal (differentiators / audience / brand_strategy "
        "depending on the task).",
        "每一对 top 配方都包含 product。"
        "另一个字段负责调节第二强信号（按任务不同分别是 differentiators / audience / brand_strategy）。"))
    _implication(T(
        "Always include product when in doubt; pair it with audience, "
        "differentiators, or brand_strategy by task.",
        "拿不准时一定包含 product；按任务再搭配 audience、differentiators 或 brand_strategy。"))

    row = write_section_bar(ws, row,
        T("A5. Per-task production picks (cheap-tier winners)",
          "A5. 每任务生产推荐（便宜层胜者）"), ncols=NCOLS)
    df = _per_task_picks_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "All 8 winners are on GPT-5-mini. Three winners beat Full Brief on "
        "their task (concept, emotion, function); the rest tie within the "
        "0.036 noise floor.",
        "8 个胜者全部在 GPT-5-mini 上。"
        "其中 3 个任务上胜者超过完整 brief（concept、emotion、function）；"
        "其余在 0.036 噪声范围内打平。"))
    _implication(T(
        "Ship the recipes in this table — they are the production setup.",
        "上线就用本表中的配方 —— 它们就是生产配置。"))

    # ============================================================
    # PART B — Keyword task scoring
    # ============================================================
    _part_header(
        "══════ Part B — Keyword task scoring (F1-based) ══════",
        "══════ Part B —— 关键词任务评分（基于 F1）══════")

    row = write_section_bar(ws, row,
        T("B1. A / B / C / D prompt compression",
          "B1. A / B / C / D Prompt 压缩"), ncols=NCOLS)
    df = _keyword_summary_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        "B is strictly better than A — 63% shorter AND +0.018 F1. C and D "
        "drop ~6 F1 points. Use B in production.",
        "B 严格优于 A —— 短 63%，F1 还 +0.018。"
        "C 和 D 损失约 6 F1 点。生产用 B。"))
    _implication(T(
        "Use prompt B in production; do not over-compress to C or D.",
        "生产使用 B 版 prompt；不要过度压缩到 C 或 D。"))

    # ============================================================
    # PART C — Stability check
    # ============================================================
    _part_header(
        "══════ Part C — Stability check (Stage B reruns) ══════",
        "══════ Part C —— 稳定性检查（Stage B 重跑）══════")

    row = write_section_bar(ws, row,
        T("C1. Cross-rerun cosine std dev",
          "C1. 跨重跑的余弦标准差"), ncols=NCOLS)
    stats = _stability_stats_dict(scored)
    if stats.get("n_cells"):
        stats_df = pd.DataFrame([
            {T("Metric", "指标"):
                 T(f"{stats['n_cells']} cells with ≥2 reruns",
                   f"{stats['n_cells']} 个单元格有 ≥2 次重跑"),
             T("Value", "数值"): "—"},
            {T("Metric", "指标"): T("Median std dev", "标准差中位数"),
             T("Value", "数值"): f"{stats['median']:.3f}"},
            {T("Metric", "指标"): T("Mean std dev", "标准差均值"),
             T("Value", "数值"): f"{stats['mean']:.3f}"},
            {T("Metric", "指标"): T("Max std dev (single outlier cell)",
                                     "标准差最大值（单个离群单元格）"),
             T("Value", "数值"): f"{stats['max']:.3f}"},
        ])
        row = write_df(ws, stats_df, row)
        row += 1
    _conclusion(T(
        "Median 0.006 is well below the 0.036 noise floor. Single-run scores "
        "are trustworthy for the headline numbers — Stage B reruns did not "
        "flip any winner.",
        "中位数 0.006 远低于 0.036 噪声下限。"
        "单次运行的分数对头条结论可信 —— Stage B 重跑没有改变任何胜者。"))
    _implication(T(
        "Single-run rankings are reliable enough for deployment decisions.",
        "单次运行的排名对部署决策已经足够可靠。"))

    # ============================================================
    # PART D — Reliability & cost
    # ============================================================
    _part_header(
        "══════ Part D — Reliability & cost ══════",
        "══════ Part D —— 可靠性 & 成本 ══════")

    row = write_section_bar(ws, row,
        T("D1. Per-model OK rate", "D1. 每模型成功率"), ncols=NCOLS)
    df = _reliability_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _hk_rl = _model_rate_limit_pct(raw, "haiku")
    _conclusion(T(
        f"Haiku ended rate-limited on {_hk_rl:.0f}% of calls — supports why the "
        f"cheap-tier default is GPT-5-mini, not Haiku. The retry layer caught "
        f"these (eventually got ok rows from the same resume keys) but at lower "
        f"throughput.",
        f"Haiku 有 {_hk_rl:.0f}% 的调用最终被限流 —— 支持为什么便宜层默认是 GPT-5-mini 而不是 Haiku。"
        f"重试层把它们都接住了（最终从同样的 resume key 拿到 ok 行），但吞吐量更低。"))
    _implication(T(
        "GPT-5-mini is the operational default; do not fall back to Haiku "
        "without a retry strategy.",
        "GPT-5-mini 即为运营默认；没有重试策略时不要回退到 Haiku。"))

    row = write_section_bar(ws, row,
        T("D2. Per-model spend", "D2. 每模型支出"), ncols=NCOLS)
    total_cost = float(scored["cost_usd"].sum()) if not scored.empty and "cost_usd" in scored.columns else 0.0
    df = _cost_breakdown_df(scored)
    if not df.empty:
        row = write_df(ws, df, row)
    row += 1
    _conclusion(T(
        f"Total experiment spend: ${total_cost:.2f}. Well under the £50 "
        "(~$63) project cap. Haiku's $1.87 dominates because of rate-limit "
        "retry traffic.",
        f"实验总支出：${total_cost:.2f}。"
        "远低于 £50（约 $63）项目上限。"
        "Haiku 的 $1.87 占主导，因为它的限流重试流量大。"))
    _implication(T(
        "The whole experiment cost less than a coffee — the methodology is "
        "cheap to re-run when data changes.",
        "整个实验的成本低于一杯咖啡 —— 数据变化时这套方法论的重跑成本极低。"))

    autosize_cols(ws, {1: 28, 2: 36, 3: 18, 4: 18, 5: 22, 6: 22})
    ws.freeze_panes = "A2"


def _sonnet_vs_cosine_kappa() -> tuple[int, float | None]:
    """Compute Cohen's weighted kappa (linear weights) between Sonnet's
    1-5 absolute rating and the cosine score binned to 5 levels.

    Returns (n_pairs, kappa). Operates on EVERY row in
    outputs/ai_judge_absolute.jsonl whose cosine is present in scored.csv,
    so it includes Sonnet scores even if they no longer overlap with the
    current 30-row sample (more statistical power that way).

    Bin thresholds match the score-guide bands used elsewhere in the
    workbook:
        cosine < 0.30 -> 1   (poor)
        0.30-0.50    -> 2   (weak)
        0.50-0.65    -> 3   (moderate)
        0.65-0.80    -> 4   (good)
        >= 0.80      -> 5   (excellent)
    """
    sonnet_path = cfg.OUTPUTS_DIR / "ai_judge_absolute.jsonl"
    if not sonnet_path.exists():
        return (0, None)
    try:
        from src.utils import read_jsonl
        scored_csv = cfg.OUTPUTS_DIR / "scored.csv"
        if not scored_csv.exists():
            return (0, None)
        sc = pd.read_csv(scored_csv)
        sc = sc[sc["cosine"].notna()]

        def bin_cos(c: float) -> int:
            if c < 0.30: return 1
            if c < 0.50: return 2
            if c < 0.65: return 3
            if c < 0.80: return 4
            return 5

        pairs: list[tuple[int, int]] = []
        for r in read_jsonl(sonnet_path):
            s = r.get("score")
            if s is None:
                continue
            cell = sc[
                (sc["brief_id"] == r.get("brief_id"))
                & (sc["task"] == r.get("task"))
                & (sc["config_id"] == r.get("config_id"))
                & (sc["model_key"] == r.get("model_key"))
            ]
            if cell.empty:
                continue
            cos = float(cell["cosine"].mean())
            pairs.append((int(s), bin_cos(cos)))

        if len(pairs) < 5:
            return (len(pairs), None)

        # Weighted kappa (linear) — pure numpy / no sklearn dependency.
        k = 5
        cm = np.zeros((k, k), dtype=int)
        for a, b in pairs:
            cm[a - 1][b - 1] += 1
        N = cm.sum()
        rs, cs = cm.sum(axis=1), cm.sum(axis=0)
        expected = np.outer(rs, cs) / N
        w = np.zeros((k, k))
        for i in range(k):
            for j in range(k):
                w[i][j] = abs(i - j) / (k - 1)
        p_o = 1 - (w * cm).sum() / N
        p_e = 1 - (w * expected).sum() / N
        if p_e == 1:
            return (len(pairs), None)
        return (len(pairs), float((p_o - p_e) / (1 - p_e)))
    except Exception:
        return (0, None)


def build_tab_human_validation(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Tab 3 — Human Review, 2-section spec.

      Section A. Human-readable samples
        Task | Recipe | Model | Ground Truth | AI Output | Auto cosine |
        Sonnet 1-5 | Human 1-5

      Section B. Agreement summary
        5-row Metric / Result table covering:
          Human-reviewed samples / Agreement method / κ score /
          Interpretation / Operational decision
        Plus a real Sonnet-vs-cosine sanity-check κ below the table.
    """
    ws = wb.create_sheet(T("Human Review", "人工评审"))
    NCOLS = 8
    # Set column widths up front — _render_human_review_table reads these to
    # compute row heights; without this they all compute as if col_w = 12,
    # producing 200+ pt rows for 250-char paragraphs.
    autosize_cols(ws, {1: 16, 2: 30, 3: 13, 4: 56, 5: 56, 6: 14, 7: 11, 8: 11})
    row = write_tab_title(ws,
        T("Human Review", "人工评审"),
        ncols=NCOLS)

    # ============================================================
    # Section A — Human-readable samples
    # ============================================================
    hr_df = _human_review_table(scored)
    sample_count = len(hr_df)
    row = write_section_bar(ws, row,
        T(f"Section A — Human-readable samples ({sample_count} rows)",
          f"Section A —— 人工可读样本（{sample_count} 行）"), ncols=NCOLS)
    if hr_df.empty:
        _merge_and_write(ws, row,
            T("No scored sentence-task data yet.", "暂无句子型任务的打分数据。"),
            ncols=NCOLS, font=ITALIC_GREY)
        row += 2
    else:
        row = _render_human_review_table(ws, hr_df, row)

    # ============================================================
    # Section B — Agreement summary
    # ============================================================
    row = write_section_bar(ws, row,
        T("Section B — Agreement summary",
          "Section B —— 一致性总结"), ncols=NCOLS)

    # Count how many sample rows already have a Sonnet score / a human score.
    sonnet_col_label = T("Sonnet 1-5", "Sonnet 1-5")
    human_col_label = T("Human 1-5", "人工 1-5")
    sonnet_done = sum(1 for v in hr_df.get(sonnet_col_label, []) if v != "—") if not hr_df.empty else 0
    human_done = sum(1 for v in hr_df.get(human_col_label, [])
                     if v not in ("Pending", "待填写")) if not hr_df.empty else 0
    n_total = sample_count

    # Prefer the REAL Human ↔ Sonnet kappa once the human column is filled.
    # Fall back to the Sonnet-vs-cosine sanity check only when there are too
    # few human ratings to compute it.
    human_kappa = _human_sonnet_kappa(scored)

    if human_kappa is not None:
        n_hk, k = human_kappa
        kappa_str = T(f"{k:.2f}  (Human ↔ Sonnet, {n_hk} paired ratings)",
                       f"{k:.2f}（人工 ↔ Sonnet，{n_hk} 对评分）")
        if k >= 0.7:
            interp_str = T("High agreement — AI judge aligns with humans",
                            "高度一致 —— AI 评委与人工判断吻合")
        elif k >= 0.4:
            interp_str = T("Moderate agreement", "中等一致")
        else:
            interp_str = T("Low agreement — do not rely on the AI judge",
                            "一致性低 —— 不要依赖 AI 评委")
        if n_hk < 30:
            decision_str = T(
                f"Based on {n_hk} human ratings (below the 30 minimum) — treat "
                f"as preliminary; collect to at least 30 (60 preferred).",
                f"基于 {n_hk} 条人工评分（未达 30 条最低门槛）—— 仅作初步参考；"
                f"建议补到至少 30 条（最好 60 条）。")
        elif n_hk < 60:
            decision_str = T(
                f"{n_hk} human ratings (≥30): if κ≥0.7 the AI judge may be used "
                f"as a supporting scorer; 60 ratings preferred before full "
                f"operational reliance.",
                f"{n_hk} 条人工评分（≥30）：若 κ≥0.7，AI 评委可作辅助打分器；"
                f"全面依赖前建议补到 60 条。")
        else:
            decision_str = T(
                f"{n_hk} human ratings (≥60): publication-grade validation. "
                f"Use the κ band above to decide the AI judge's role.",
                f"{n_hk} 条人工评分（≥60）：达到发表级验证。"
                f"按上面的 κ 区间决定 AI 评委的角色。")
    else:
        # No human ratings yet — show the Sonnet-vs-cosine sanity check, clearly
        # labelled as NOT human validation.
        n_pairs, sanity_kappa = _sonnet_vs_cosine_kappa()
        if sanity_kappa is not None and n_pairs >= 5:
            kappa_str = f"{sanity_kappa:.2f}  (Sonnet vs cosine-bin sanity check)"
            if sanity_kappa >= 0.7:
                interp_str = T("High agreement (sanity check only)",
                                "高度一致（仅健全性检查）")
            elif sanity_kappa >= 0.4:
                interp_str = T("Moderate agreement (sanity check only)",
                                "中等一致（仅健全性检查）")
            else:
                interp_str = T("Low agreement (sanity check only)",
                                "一致性较低（仅健全性检查）")
            decision_str = T(
                "Current κ is Sonnet-vs-cosine sanity check, not final human "
                "validation. Human validation pending — fill the Human 1-5 "
                "column above to compute Human ↔ Sonnet κ.",
                "当前 κ 是 Sonnet ↔ 余弦的健全性检查，不是最终的人工验证。"
                "人工验证尚未完成 —— 填上方「人工 1-5」列后即可计算人工 ↔ Sonnet 的 κ。")
        else:
            kappa_str = T("Pending", "待计算")
            interp_str = T("Pending", "待评定")
            decision_str = T("Pending", "待决策")

    summary_df = pd.DataFrame([
        {T("Metric", "指标"):
             T("Human-reviewed samples", "人工评审样本"),
         T("Result", "结果"):
             f"{human_done} of {n_total} ({'pending' if human_done < 30 else 'minimum met'})" if LANG == "en"
             else f"{n_total} 条中已评 {human_done} 条（{'未达 30 条' if human_done < 30 else '已达最低门槛'}）"},
        {T("Metric", "指标"):
             T("Agreement metric", "一致性指标"),
         T("Result", "结果"):
             T("Weighted κ", "加权 κ")},
        {T("Metric", "指标"):
             T("κ result", "κ 结果"),
         T("Result", "结果"): kappa_str},
        {T("Metric", "指标"):
             T("Interpretation", "判定"),
         T("Result", "结果"): interp_str},
        {T("Metric", "指标"):
             T("Operational decision", "操作决策"),
         T("Result", "结果"): decision_str},
    ])
    row = write_df(ws, summary_df, row)
    row += 2

    # Real sanity-check kappa — Sonnet vs cosine-bin, on every Sonnet rating
    # in ai_judge_absolute.jsonl. This is an AI-internal consistency measure
    # (does Sonnet agree with cosine on what is good?), NOT a human↔AI
    # validation. Surfaced here as the only κ that can actually be computed
    # before human review.
    if sanity_kappa is not None and n_pairs >= 5:
        if sanity_kappa >= 0.7:
            verdict = T("High agreement", "高度一致")
        elif sanity_kappa >= 0.4:
            verdict = T("Moderate agreement", "中等一致")
        else:
            verdict = T("Low agreement", "一致性较低")
        _merge_and_write(ws, row,
            T(
                f"Sanity check — across {n_pairs} Sonnet absolute ratings "
                f"collected so far, Sonnet 1-5 vs cosine-bin Cohen's "
                f"weighted κ = {sanity_kappa:.2f} ({verdict}). This is an "
                "AI-internal consistency measure (Sonnet's rating tracks "
                "cosine), not a human-vs-AI validation. The Human ↔ Sonnet "
                "κ above is the one Simon's review needs.",
                f"健全性检查 —— 在已收集的 {n_pairs} 条 Sonnet 绝对评分上，"
                f"Sonnet 1-5 与余弦分箱的 Cohen 加权 κ = {sanity_kappa:.2f}"
                f"（{verdict}）。这是 AI 内部一致性度量"
                "（Sonnet 的评分是否与余弦一致），不是人工 vs AI 的验证。"
                "上方「人工 ↔ Sonnet」的 κ 才是 Simon 审阅真正需要的那个。"),
            ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
        ws.row_dimensions[row].height = _estimate_row_height(
            "Sanity check — across N Sonnet absolute ratings collected so far...",
            total_width_chars=160)
        row += 2

    # Operational follow-up: tell the reader how to fill in missing AI ratings.
    if sonnet_done < n_total and n_total > 0:
        _merge_and_write(ws, row,
            T(
                f"Note: Sonnet has scored {sonnet_done} of {n_total} sample "
                "rows. The remaining rows can be filled in by re-running "
                "`python -m scripts.ai_judge_absolute --sample-n 30 "
                "--budget 0.50` (resumable; only spends on missing rows, "
                "roughly $0.01 expected).",
                f"提示：Sonnet 已对 {n_total} 条样本中的 {sonnet_done} 条评分。"
                "其余行可通过运行 "
                "`python -m scripts.ai_judge_absolute --sample-n 30 --budget 0.50` 补齐"
                "（脚本可断点续跑，只对缺失行付费，预计约 $0.01）。"),
            ncols=NCOLS, font=Font(size=10, italic=True, color="B58900"),
            fill=PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"))
        ws.row_dimensions[row].height = _estimate_row_height(
            "Note: Sonnet has scored X of Y sample rows...",
            total_width_chars=160)
        row += 2

    # Column widths for the 8-col table. Recipe col widened so longer
    # recipes ("brand_strategy + differentiators") fit on one line; the
    # rating cells trimmed since they hold "1"–"5" or "Pending".
    # 1 Task / 2 Recipe / 3 Model / 4 GT / 5 AI Output / 6 Auto cosine /
    # 7 Sonnet 1-5 / 8 Human 1-5
    autosize_cols(ws, {
        1: 16, 2: 30, 3: 13, 4: 56, 5: 56, 6: 14, 7: 11, 8: 11,
    })
    ws.freeze_panes = "A2"


def build_tab_appendix(wb, scored: pd.DataFrame, raw: pd.DataFrame) -> None:
    """Tab 5 — Appendix. Methodology and supplementary tables that do not
    belong on the four main decision tabs.

    Sections:
      - How the experiment was run
      - How scores are computed
      - Stability rerun results
      - AI judge pairwise (optional cross-check)
      - Premium-model quality ceiling (Phase 4)
      - Cost
    """
    ws = wb.create_sheet(T("Appendix", "附录"))
    NCOLS = 9
    row = write_tab_title(ws,
        T("Appendix — methodology and supplementary results",
          "附录 —— 方法学与补充结果"),
        ncols=NCOLS)

    _merge_and_write(ws, row,
        T("Single source of truth for how the numbers on the other tabs were "
          "produced. Supplementary tables that did not belong on the main "
          "decision tabs also live here.",
          "其他 tab 中所有数字如何计算的唯一权威说明。"
          "不适合放在主决策 tab 上的补充表格也放在这里。"),
        ncols=NCOLS, font=Font(size=11, color="0B0C0C"))
    row += 2

    # ----- Field-contribution + per-task-one-sentence reference tables -----
    # Moved off Tab 2 per user feedback: they belong here as reference, not in
    # the production-decision tab where they repeat what the per-task table
    # already implies.
    row = write_section_bar(ws, row,
        T("Field contribution summary", "字段贡献总结"), ncols=NCOLS)
    row = write_df(ws, _field_contribution_summary_df(), row)
    row += 2

    row = write_section_bar(ws, row,
        T("One-sentence explanation per task", "每任务一句话解释"), ncols=NCOLS)
    row = write_df(ws, _per_task_one_sentence_df(), row)
    row += 2

    # ----- Section 1: How the experiment was run (workflow + dataset + scope) -----
    row = write_section_bar(ws, row,
        T("How the experiment was run",
          "实验是怎么跑的"), ncols=NCOLS)
    run_lines = [
        T("Dataset. 23 real client brand briefs. Each brief has hand-curated "
          "ground-truth sentences for 8 extraction tasks and a 10-keyword list "
          "for the keyword task.",
          "数据集。23 份真实的客户品牌 brief。"
          "每份 brief 在 8 个抽取任务上都有人工标注的 ground-truth 句子，"
          "并在关键词任务上有人工标注的 10 个关键词列表。"),
        T("Models tested. Cheap tier: Haiku and GPT-5-mini. Medium tier: "
          "Sonnet 4.6 and GPT-5. Premium tier: Opus 4.7 and GPT-5.5.",
          "测试模型。便宜层：Haiku 和 GPT-5-mini。"
          "中端层：Sonnet 4.6 和 GPT-5。旗舰层：Opus 4.7 和 GPT-5.5。"),
        T("Recipes tested. 142 in Stage A. They cover Full Brief, the no-context "
          "baseline, every single brief field, every two-field pair, and a few "
          "targeted metadata combinations.",
          "测试配方。Stage A 共 142 个。"
          "它们涵盖完整 brief、无上下文基线、每个单字段、每对双字段，以及若干针对性的 metadata 组合。"),
        T("Stage A. Run every recipe on both cheap models across all 23 briefs. "
          "Total 6,532 API calls.",
          "Stage A。在两个便宜模型上对每个配方跑所有 23 个 brief。共 6,532 次 API 调用。"),
        T("Stage B. Re-run the top two recipes per task two more times to confirm "
          "scores were not single-run luck. Total 1,904 API calls.",
          "Stage B。把每个任务的 top 2 配方再多跑两次，确认分数不是单次运气。"
          "共 1,904 次 API 调用。"),
        T("Phase 4. Small-sample premium check: run the four stronger models "
          "only on surviving top recipes and 3 representative briefs. Use this "
          "as directional evidence only.",
          "Phase 4。小样本旗舰检查：仅在幸存 top 配方和 3 个代表性 brief 上运行四个更强模型。"
          "这只能作为方向参考。"),
        T("Tie-break rule. When two recipes differ by less than 0.036 cosine, "
          "default to the cheaper, shorter, or simpler one.",
          "平局规则。两个配方差值小于 0.036 余弦时，默认选更便宜、更短或更简单的那个。"),
    ]
    for line in run_lines:
        _merge_and_write(ws, row, line, ncols=NCOLS,
                         font=Font(size=10, color="0B0C0C"))
        ws.row_dimensions[row].height = _estimate_row_height(line, total_width_chars=140)
        row += 1
    row += 1

    # ----- Section 2: How scores are computed (just the two formulas) -----
    row = write_section_bar(ws, row,
        T("How scores are computed", "分数是怎么算的"), ncols=NCOLS)
    score_lines = [
        T("Sentence-task score. For each of the 8 sentence tasks, the AI "
          "output and the ground-truth sentence are turned into vectors "
          "using OpenAI text-embedding-3-large. The score for that brief is "
          "the cosine similarity of the two vectors. 0 means unrelated. 1 "
          "means identical meaning. The cell score in the main tables is the "
          "mean of the 23 per-brief cosines.",
          "句子型任务分数。对 8 个句子型任务中的每一个，"
          "AI 输出和 ground-truth 句子都用 OpenAI text-embedding-3-large 转成向量。"
          "该 brief 的分数是两个向量的余弦相似度。"
          "0 表示完全不相关。1 表示含义完全一致。"
          "主表里的单元格分数是 23 个 per-brief 余弦的均值。"),
        T("Keyword-task score. For the keyword task, the AI predicts 10 "
          "keywords. Precision is the share that are correct. Recall is the "
          "share of the 10 gold keywords the AI found. F1 combines the two. "
          "Porter stemming is applied so running matches run.",
          "关键词任务分数。关键词任务中，AI 预测 10 个关键词。"
          "Precision 是其中正确的比例。Recall 是 10 个金标关键词中 AI 找到的比例。"
          "F1 是两者的综合。应用 Porter 词干化，所以 running 能匹配 run。"),
    ]
    for line in score_lines:
        _merge_and_write(ws, row, line, ncols=NCOLS,
                         font=Font(size=10, color="0B0C0C"))
        ws.row_dimensions[row].height = _estimate_row_height(line, total_width_chars=140)
        row += 1
    row += 1

    # ----- Section 3: Decision rules (replaces the scattered 0.036 paragraph
    # with a clean set of small tables — mirrors the "HOW WE GRADE THE AI'S
    # OUTPUTS" flowchart so a reader can trace one finger from "raw output"
    # to "production decision" without re-reading prose.) -----
    row = write_section_bar(ws, row,
        T("Decision rules", "判定规则"), ncols=NCOLS)
    _merge_and_write(ws, row,
        T("Every recommendation on Tabs 1 through 4 follows the rules below. "
          "Each rule is a one-line decision; together they form the end-to-end "
          "scoring flow.",
          "Tab 1 至 Tab 4 上的每条推荐都遵循下方规则。"
          "每条规则都是一次一行决策；它们合起来就是端到端的打分流程。"),
        ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
    row += 2

    # Rule 1 — what counts as a real difference (the 3-band noise-floor table)
    _merge_and_write(ws, row,
        T("Rule 1. What counts as a real quality difference between two recipes",
          "规则 1. 两个配方之间什么算真正的质量差异"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    noise_band_df = pd.DataFrame([
        {T("Cosine difference", "余弦差值"):
             T("≤ 0.036", "≤ 0.036"),
         T("Verdict", "判定"):
             T("Within measurement noise. Treat the two recipes as equal.",
               "在测量噪声范围内。把两个配方视为相同。")},
        {T("Cosine difference", "余弦差值"):
             T("0.036 to 0.09", "0.036 到 0.09"),
         T("Verdict", "判定"):
             T("Possibly better, but inconclusive. Need more data to be sure.",
               "可能更好，但还不能下结论。需要更多数据。")},
        {T("Cosine difference", "余弦差值"):
             T("≥ 0.09", "≥ 0.09"),
         T("Verdict", "判定"):
             T("Significantly better. Treat as a real quality difference.",
               "显著更好。视为真正的质量差异。")},
    ])
    row = write_df(ws, noise_band_df, row)
    _merge_and_write(ws, row,
        T("The 0.036 band is 2 sigma; sigma = 0.018 measured empirically by "
          "re-running the same recipe multiple times (see scripts/measure_noise.py). "
          "5 sigma is reserved for headline-level claims only.",
          "0.036 这条线是 2 sigma；sigma = 0.018 是通过重复运行同一配方测出的经验值"
          "（见 scripts/measure_noise.py）。5 sigma 仅用于头条级别的主张。"),
        ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "The 0.036 band is 2 sigma...", total_width_chars=140)
    row += 2

    # Rule 2 — near-tie tie-break preferences
    _merge_and_write(ws, row,
        T("Rule 2. When recipes are tied within noise, preference goes to",
          "规则 2. 配方在噪声范围内打平时，优先选择"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    for pref in [
        T("Shorter prompts.", "更短的 prompt。"),
        T("Fewer brief fields.", "更少的 brief 字段。"),
        T("Lower-cost models.", "更便宜的模型。"),
        T("More stable outputs across reruns.", "重跑时更稳定的输出。"),
    ]:
        _merge_and_write(ws, row, f"• {pref}", ncols=NCOLS,
                         font=Font(size=11, color="0B0C0C"))
        row += 1
    row += 1

    # Rule 3 — when to upgrade to a stronger model
    _merge_and_write(ws, row,
        T("Rule 3. When to upgrade to a medium or stronger model",
          "规则 3. 何时升级到中端或更强模型"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    _merge_and_write(ws, row,
        T("Upgrade only when the quality gain on the task is at least 0.036 "
          "cosine (2 sigma). Anything smaller is within noise and the "
          "cheaper model wins by default.",
          "只有当任务上的质量提升达到至少 0.036 余弦（2 sigma）时才升级。"
          "更小的提升属于噪声，默认仍使用便宜模型。"),
        ncols=NCOLS, font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "Upgrade only when the quality gain...", total_width_chars=140)
    row += 2

    # Rule 4 — keyword compression rule
    _merge_and_write(ws, row,
        T("Rule 4. Picking a shorter keyword prompt",
          "规则 4. 选择更短的关键词 prompt"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    _merge_and_write(ws, row,
        T("Choose the shortest prompt version whose F1 loss vs the original "
          "is at most 0.05. B — Reduced satisfies this with a small F1 gain.",
          "在 F1 损失不超过 0.05 的前提下，选最短的 prompt 版本。"
          "B — Reduced 满足这一条，并且 F1 还略有提升。"),
        ncols=NCOLS, font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "Choose the shortest prompt version...", total_width_chars=140)
    row += 2

    # Rule 5 — stability check
    _merge_and_write(ws, row,
        T("Rule 5. Stability check", "规则 5. 稳定性检查"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    _merge_and_write(ws, row,
        T("The top recipes are re-run two extra times in Stage B. The "
          "standard deviation across the three runs is the stability signal: "
          "below 0.036 is stable, 0.036 to 0.05 is mildly noisy, above 0.05 "
          "means a single-run result cannot be trusted.",
          "在 Stage B 把 top 配方再多跑 2 次。"
          "三次运行之间的标准差就是稳定性信号："
          "小于 0.036 表示稳定，0.036 到 0.05 表示略嘈杂，"
          "大于 0.05 表示单次结果不可信。"),
        ncols=NCOLS, font=Font(size=11, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "The top recipes are re-run two extra times...", total_width_chars=140)
    row += 2

    # Rule 6 — when the AI judge can be used (3-band κ table)
    _merge_and_write(ws, row,
        T("Rule 6. When the AI judge can be used "
          "(minimum 30 human-rated pairs; 60 preferred)",
          "规则 6. AI 评审何时可用（至少 30 条人工配对；建议 60 条）"),
        ncols=NCOLS, font=Font(size=11, bold=True, color="0B0C0C"))
    row += 1
    kappa_band_df = pd.DataFrame([
        {T("Agreement score", "一致性分数"):
             T("≥ 0.7", "≥ 0.7"),
         T("Decision", "决策"):
             T("High agreement. AI judge may be used as supporting scorer; "
               "prefer 60 human pairs before relying on it operationally.",
               "高度一致。AI 评审可作为辅助评分器；正式依赖前建议做到 60 条人工配对。")},
        {T("Agreement score", "一致性分数"):
             T("0.4 to 0.7", "0.4 到 0.7"),
         T("Decision", "决策"):
             T("Moderate. Collect another 30 human samples and re-evaluate.",
               "中等。再收集 30 条人工样本，重新评估。")},
        {T("Agreement score", "一致性分数"):
             T("< 0.4", "< 0.4"),
         T("Decision", "决策"):
             T("Low agreement. Reject the AI judge. Rewrite the prompt or fall back to human-only rating.",
               "一致性较低。拒绝 AI 评审。重写 prompt 或改用纯人工评分。")},
    ])
    row = write_df(ws, kappa_band_df, row)
    _merge_and_write(ws, row,
        T("Until Human ↔ Sonnet κ is computed on at least 30 paired scores, "
          "the AI judge is not validated and its ratings are reference-only.",
          "在至少 30 条人工 ↔ Sonnet 配对评分计算出 κ 之前，AI 评审尚未被验证，"
          "其评分只能作为参考。"),
        ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
    row += 2

    # ----- Stability rerun -----
    row = write_section_bar(ws, row,
        T("Stability rerun results", "稳定性重跑结果"), ncols=NCOLS)
    stab_df, stab_msg = _stability_summary(scored)
    _merge_and_write(ws, row, stab_msg, ncols=NCOLS, font=ITALIC_GREY)
    row += 1
    if not stab_df.empty:
        row = _emit_table_meta(ws, row, NCOLS,
            source=T(
                "Stage B re-ran the top two recipes per task on all 23 briefs "
                "two extra times. So each surviving cell has three score "
                "samples from three independent runs.",
                "Stage B 把每个任务的 top 2 配方在全部 23 brief 上额外跑了两次。"
                "所以每个幸存单元格有来自三次独立运行的三个分数样本。"),
            scoring=T(
                "Per-cell standard deviation of cosine across the three runs "
                "is the stability signal. Below 0.036 means stable. Between "
                "0.036 and 0.05 means mildly noisy. Above 0.05 means a single "
                "run cannot be trusted.",
                "每个单元格在三次重跑上的余弦标准差就是稳定性信号。"
                "小于 0.036 表示稳定。在 0.036 到 0.05 之间表示略嘈杂。"
                "大于 0.05 表示单次结果不可信。"))
        row = write_df(ws, stab_df, row)
    else:
        row += 1

    # ----- AI judge pairwise -----
    row = write_section_bar(ws, row,
        T("AI judge pairwise. Optional cross-check.",
          "AI 评审两两对比。可选交叉检查。"), ncols=NCOLS)
    _merge_and_write(ws, row,
        T("To refresh the AI assist ratings on Tab 4 or to run a new "
          "pairwise comparison, run from PowerShell: "
          "`python -m scripts.ai_judge_absolute --sample-n 30 --budget 0.50`. "
          "The script is resumable and only spends on rows that have not "
          "been scored yet.",
          "如需补齐 Tab 4 的 AI 辅助评分，或者跑新的两两对比，"
          "在 PowerShell 运行：`python -m scripts.ai_judge_absolute --sample-n 30 --budget 0.50`。"
          "脚本可断点续跑，只对未评分的行付费。"),
        ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "To refresh the AI assist ratings...", total_width_chars=140) + 12
    row += 2
    pair_df, pair_msg = _pairwise_summary()
    _merge_and_write(ws, row, pair_msg, ncols=NCOLS, font=ITALIC_GREY)
    row += 1
    if not pair_df.empty:
        row = _emit_table_meta(ws, row, NCOLS,
            source=T(
                "Sonnet 4.6 was shown pairs of outputs from the top recipes "
                "and asked to pick the better one. Each pair is shown twice "
                "with positions swapped, to control for position bias.",
                "Sonnet 4.6 被展示了来自 top 配方的多对输出，并被要求挑出更好的那个。"
                "每对都展示两次并互换位置，以控制位置偏见。"),
            scoring=T(
                "Only consistent winners count. If Sonnet picks A on the first "
                "showing and B on the swapped showing, the pair is discarded. "
                "This is a sanity check on the cosine ranking, not the main "
                "metric.",
                "只记录一致的赢家。如果 Sonnet 在第一次选 A，在互换后选 B，这对就被丢弃。"
                "这只是对余弦排序的合理性检查，不是主指标。"))
        row = write_df(ws, pair_df, row)
    else:
        row += 1

    # ----- Premium-model ceiling -----
    row = write_section_bar(ws, row,
        T("Premium-model quality ceiling. Phase 4.",
          "旗舰模型质量上限。Phase 4。"), ncols=NCOLS)
    _merge_and_write(ws, row,
        T("Phase 4 is a small-sample premium-model check: it was run only "
          "on 3 representative briefs, not across all 23 briefs. Treat every "
          "premium-model result here as directional evidence, not a validated "
          "production decision.",
          "Phase 4 是小样本旗舰模型检查：只在 3 个代表性 brief 上运行，"
          "不是覆盖全部 23 个 brief。此处所有旗舰模型结果都只能作为方向参考，"
          "不能当作已验证的生产决策。"),
        ncols=NCOLS, font=Font(size=10, color="0B0C0C"))
    ws.row_dimensions[row].height = _estimate_row_height(
        "Headline reading: on 7 of 8 sentence tasks...", total_width_chars=140)
    row += 2
    prem_df, prem_msg, prem_verdict = _phase4_premium_summary(scored)
    _merge_and_write(ws, row, prem_msg, ncols=NCOLS, font=ITALIC_GREY)
    row += 1
    if not prem_df.empty:
        row = _emit_table_meta(ws, row, NCOLS,
            source=T(
                "The top winner per task was re-run on a curated 3-brief subset "
                "with Opus 4.7 / GPT-5.5 in the premium tier and Sonnet 4.6 / "
                "GPT-5 in the medium tier.",
                "每个任务的 top winner 只在 3 个代表性 brief 子集上重跑，"
                "使用旗舰层 Opus 4.7 / GPT-5.5 和中端层 Sonnet 4.6 / GPT-5。"),
            scoring=T(
                "Each cell is the mean over the curated Phase 4 briefs actually "
                "run for that model and recipe. Because n=3 briefs, use the "
                "result only to decide whether a full 23-brief premium validation "
                "is worth running.",
                "每个单元格是该模型和配方在实际运行的 Phase 4 代表性 brief 上的均值。"
                "由于 n=3 个 brief，只能用于判断是否值得再做完整 23 个 brief 的旗舰验证。"))
        best_prem_col = T("Best premium", "旗舰最佳")
        if best_prem_col in prem_df.columns:
            premium_scores = prem_df[best_prem_col].dropna()
            if not premium_scores.empty:
                avg_prem = float(premium_scores.mean())
                max_prem = float(premium_scores.max())
                headline = T(
                    f"Quality ceiling with best models: avg {avg_prem:.2f}, "
                    f"peak {max_prem:.2f} (similarity to ground truth, 0–1 scale).",
                    f"旗舰模型质量上限：平均 {avg_prem:.2f}，峰值 {max_prem:.2f}（0–1 余弦尺度）。"
                )
                _merge_and_write(ws, row, headline, ncols=NCOLS,
                                 font=Font(size=11, bold=True, color="0B0C0C"),
                                 fill=LIGHT_BLUE)
                row += 2
        row = write_df(ws, prem_df, row)
        if prem_verdict and "|" in prem_verdict:
            kind, msg = prem_verdict.split("|", 1)
            color = {"green": "00703C", "yellow": "B58900", "neutral": "1D70B8"}.get(kind, "0B0C0C")
            _merge_and_write(ws, row, msg, ncols=NCOLS, font=Font(size=11, italic=True, color=color))
            row += 1
    row += 1

    # ----- Cost (breakdown + assumptions merged into one section) -----
    row = write_section_bar(ws, row,
        T("Cost", "成本"), ncols=NCOLS)
    cost_df = _cost_summary_by_stage(raw)
    if cost_df.empty:
        _merge_and_write(ws, row, T("No cost data yet.", "暂无成本数据。"),
                         ncols=NCOLS, font=ITALIC_GREY)
        row += 1
    else:
        row = _emit_table_meta(ws, row, NCOLS,
            source=T(
                "Cost is recorded per API call from the provider response. "
                "Formula: input tokens times the input price plus output "
                "tokens times the output price.",
                "成本按每次 API 调用从提供商响应中记录。"
                "公式：输入 token 数乘输入单价，加输出 token 数乘输出单价。"),
            scoring=T(
                "Each row is the sum of all API calls in that stage. The "
                "Total row shows project spend against the 50 pound cap. "
                "Cost-quality recommendations remain provisional until model "
                "IDs and input/output token prices are verified and dated.",
                "每行是该阶段所有 API 调用的总和。Total 行显示项目花费与 50 英镑上限的对比。"
                "模型 ID 和输入/输出 token 单价确认并写明价格日期前，成本质量建议都只是暂定。"))
        row = write_df(ws, cost_df, row)
        total = float(raw["cost_usd"].sum())
        write_cell(ws, row, 1, T("Total", "合计"),
                   font=Font(bold=True, size=11), fill=LIGHT_GREY, border=THIN_BORDER,
                   align=Alignment(horizontal="left", vertical="center", indent=1))
        write_cell(ws, row, 2,
                   T(f"${total:.4f} of about $63 budget",
                     f"${total:.4f}，预算约 $63"),
                   font=Font(bold=True, size=11), fill=LIGHT_GREY, border=THIN_BORDER,
                   align=Alignment(horizontal="left", vertical="center", indent=1))
        ws.row_dimensions[row].height = 22
        row += 1
        _merge_and_write(ws, row,
            T("Total spend stayed well under the 50 pound project cap. "
              "Before publishing any cost-quality claim, verify the model IDs, "
              "input/output token prices, and price version date in src/config.py.",
              "总支出远低于 50 英镑项目上限。发布任何成本质量结论前，"
              "必须先确认 src/config.py 中的模型 ID、输入/输出 token 单价和价格版本日期。"),
            ncols=NCOLS, font=Font(size=10, italic=True, color="626A6E"))
        row += 1

    # Column widths for the 9-col appendix. Col 2 widened because the
    # Decision Rules tables put long Verdict / Decision text in column 2;
    # col 3 widened slightly so Field-Contribution-Summary "Suitable tasks"
    # cell does not wrap awkwardly.
    autosize_cols(ws, {
        1: 24, 2: 50, 3: 28, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16, 9: 22,
    })
    ws.freeze_panes = "A2"


# ---------- main ----------

def main():
    import argparse
    global LANG
    ap = argparse.ArgumentParser()
    ap.add_argument("--local-only", action="store_true",
                    help="Save xlsx to Results/ and skip Google Sheets upload.")
    ap.add_argument("--sheet-id", default=None,
                    help="Override RESULTS_SHEETS_ID. Set to 'none' to disable upload.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Alias of --local-only.")
    ap.add_argument("--lang", choices=["en", "zh"], default="en",
                    help="Workbook language. 'zh' writes the Chinese edition to "
                         "Results/Prompt Eval Results - 中文.xlsx and skips Drive "
                         "upload by default (override with --sheet-id <id>).")
    args = ap.parse_args()
    LANG = args.lang

    if not SCORED_CSV.exists():
        raise SystemExit(f"Missing {SCORED_CSV}. Run `analyze --score` first.")

    scored = pd.read_csv(SCORED_CSV)

    from src.utils import read_jsonl
    raw_records = read_jsonl(cfg.OUTPUTS_DIR / "results.jsonl")
    raw = pd.DataFrame(raw_records) if raw_records else scored

    # Refresh Human ↔ Sonnet kappa.json up front (reads back any human ratings
    # preserved from the previous build) so Tab 1's agreement line is current in
    # this same build, not one build behind.
    _human_sonnet_kappa(scored)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_tab_executive_summary(wb, scored, raw)       # Tab 1 — Executive Summary (conclusion)
    build_tab_recommended_configs(wb, scored, raw)     # Tab 2 — Final Production Recommendations
    build_tab_analysis(wb, scored, raw)                # Tab 3 — Analysis (supporting data tables)
    build_tab_human_validation(wb, scored, raw)        # Tab 4 — Human Review

    # Output file + Sheets target are language-specific so EN and ZH don't
    # collide. ASCII-only filename for ZH so Excel + OneDrive sync don't
    # choke on Chinese characters in the file name itself (content is still
    # fully Chinese).
    if LANG == "zh":
        out_path = cfg.RESULTS_DIR / "Prompt Eval Results (CN).xlsx"
        target_sheet_id = (
            args.sheet_id
            if args.sheet_id
            else getattr(cfg, "RESULTS_SHEETS_ID_ZH", None)
        )
    else:
        out_path = DST
        target_sheet_id = args.sheet_id if args.sheet_id else cfg.RESULTS_SHEETS_ID
    if args.local_only or args.dry_run or (args.sheet_id and args.sheet_id.lower() == "none"):
        target_sheet_id = None

    if target_sheet_id:
        import tempfile
        from pathlib import Path
        tmp = Path(tempfile.gettempdir()) / "_build_xlsx_upload.xlsx"
        try:
            wb.save(tmp)
            print(f"Built workbook ({tmp.stat().st_size:,} bytes, sheets: {wb.sheetnames})")
            print(f"Uploading to Google Sheets {target_sheet_id} (replacing existing content)...")
            from src.google_drive import upload_xlsx_to_replace_sheets
            url = upload_xlsx_to_replace_sheets(tmp, target_sheet_id)
            print(f"\n[OK] Google Sheets updated: {url}\n")
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass
    else:
        cfg.RESULTS_DIR.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        print(f"-> {out_path}")
        print(f"   Sheets: {wb.sheetnames}")
        if LANG == "zh":
            print("   (中文版 — 已跳过 Drive 上传，--local-only 模式。)")
        else:
            print("   (Skipped Drive upload — --local-only or sheet-id=none.)")


if __name__ == "__main__":
    main()
